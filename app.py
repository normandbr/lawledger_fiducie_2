# -*- coding: utf-8 -*-
"""
LawLedger - Legal Invoice Management System
"""
import warnings
# On ignore l'avertissement de SQLAlchemy sur la version du serveur
warnings.filterwarnings("ignore", message=".*Unrecognized server version info.*")

from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, Response, session
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import inspect as sa_inspect, text as sa_text
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from flask_mail import Mail, Message
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.utils import secure_filename
from itsdangerous import URLSafeTimedSerializer, SignatureExpired, BadSignature
from datetime import datetime, timezone, timedelta, UTC
from decimal import Decimal, ROUND_HALF_UP
import math
from translations import LEXICON
import urllib.parse
import urllib.request
import urllib.error
import json
import re
import logging
import os
import csv
import io
import pandas as pd
import uuid
import licensing as _licensing
import sys
import hashlib
logger = logging.getLogger(__name__)

APP_VERSION = "2026.03.5"
def get_real_ip():
   
    
    ip = request.headers.get('X-ARR-ClientIP')

    if not ip:
        ip = request.headers.get('X-Forwarded-For')

    if ip:
        ip = ip.split(',')[0].strip()
    else:
        ip = request.remote_addr

    return ip


def _round_half_up(value, places=2):
    """Round *value* to *places* decimal places using the traditional
    "round half up" rule (0.5 always rounds away from zero).

    Python's built-in ``round()`` uses banker's rounding (round-half-to-even),
    which can produce counter-intuitive results for monetary amounts
    (e.g. 0.045 rounds to 0.04 instead of 0.05).  This helper always rounds
    0.5 upward, matching standard accounting expectations.
    """
    quant = Decimal('0.' + '0' * places) if places > 0 else Decimal('1')
    return float(Decimal(str(value)).quantize(quant, rounding=ROUND_HALF_UP))


# Initialize Flask app
app = Flask(__name__, static_folder="static", template_folder="templates")
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1)

class _ScriptNameMiddleware:
    """Inject SCRIPT_NAME so that url_for() generates prefix-aware URLs.

    The IIS reverse proxy strips the /lawledger prefix before forwarding
    requests to Flask (via the web.config rewrite rule).  This middleware
    tells Werkzeug/Flask its mount point so that all redirect and link URLs
    are generated with the correct prefix (e.g. /lawledger).
    """

    def __init__(self, wsgi_app, script_name):
        self.app = wsgi_app
        self.script_name = script_name

    def __call__(self, environ, start_response):
        environ['SCRIPT_NAME'] = self.script_name
        return self.app(environ, start_response)


# Load .env file if present (must run before reading env vars like URL_PREFIX)
try:
    from dotenv import load_dotenv
    import logging as _logging

    class _DotenvHelpFilter(_logging.Filter):
        """Enhance python-dotenv parse warnings with actionable guidance."""
        def filter(self, record):
            if 'could not parse statement' in record.getMessage():
                record.msg += (
                    '\n  Your .env file contains an invalid line '
                    '(e.g. a PowerShell here-string delimiter like "@ | Out-File ...).'
                    '\n  The file must contain only KEY=VALUE pairs.'
                    '\n  On Windows: copy .env.example .env  '
                    'then open .env in Notepad and fill in your values.'
                )
            return True

    _logging.getLogger('dotenv.main').addFilter(_DotenvHelpFilter())
    load_dotenv()
except ImportError:
    pass

# When URL_PREFIX is set (e.g. URL_PREFIX=/lawledger), wrap the WSGI app so
# that url_for() and redirects include the prefix automatically.
_url_prefix = os.environ.get('URL_PREFIX', '').rstrip('/')
if _url_prefix:
    app.wsgi_app = _ScriptNameMiddleware(app.wsgi_app, _url_prefix)

# Configure Flask app
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key')
app.config['JSON_AS_ASCII'] = False  # Ensure French characters are not ASCII-escaped in JSON
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = False  # Important car tu n'as pas de HTTPS (SSL) localement
app.config['SESSION_COOKIE_HTTPONLY'] = True
# Cookie path set to root so the session works both with and without the IIS
# reverse-proxy prefix (/lawledger).
app.config['SESSION_COOKIE_PATH'] = '/'
# Sessions are non-persistent: the cookie has no expiry, so it disappears when
# the browser is closed and a fresh login is required every time.
# Note: PERMANENT_SESSION_LIFETIME only applies to *permanent* sessions; the
# 15-minute inactivity timeout is enforced server-side by _check_session_timeout.
app.config['SESSION_PERMANENT'] = False
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=15)
app.config.update(
    SESSION_PERMANENT=False,          # 🔴 session non persistante
    SESSION_TYPE='filesystem',        # ou autre selon ton setup
    PERMANENT_SESSION_LIFETIME=timedelta(minutes=0)  # optionnel
)

#app.config['SECRET_KEY'] = '8513c25c36b9708695e2fc52da2ba23df65839164c5d19fcefd5ea0dc565896f'

# Build SQL Server connection string
server   = os.environ.get('DB_SERVER', 'localhost')
database = os.environ.get('DB_NAME', 'LawLedger')
username = os.environ.get('DB_USER', '')
password = os.environ.get('DB_PASSWORD', '')
driver   = os.environ.get('DB_DRIVER', 'ODBC Driver 17 for SQL Server')

# URL encode the connection string parameters
params = urllib.parse.quote_plus(
    f"DRIVER={{{driver}}};"
    f"SERVER={server};"
    f"DATABASE={database};"
    f"UID={username};"
    f"PWD={password};"
    f"TrustServerCertificate=yes;"
)

app.config['SQLALCHEMY_DATABASE_URI'] = f"mssql+pyodbc:///?odbc_connect={params}"
# Allow DATABASE_URL env var to override (useful for testing with SQLite)
_override_db_url = os.environ.get('DATABASE_URL', '').strip()
if _override_db_url:
    app.config['SQLALCHEMY_DATABASE_URI'] = _override_db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ECHO'] = os.environ.get('DEBUG', '').lower() in ('1', 'true', 'yes')

@app.route('/set_lang/<lang>')
def set_lang(lang):
    session['lang'] = lang if lang in ['fr', 'en'] else 'fr'
    return redirect(request.referrer or url_for('index'))
@app.route('/dashboard')
@login_required
def dashboard():
    return "OK"

# Flask-Mail configuration (from .env or config.ini)
# When BREVO_API_KEY is set the app sends via Brevo's transactional API directly,
# so the SMTP settings below are only used as a fallback.
if os.environ.get('BREVO_API_KEY'):
    app.config['MAIL_SERVER'] = 'smtp-relay.brevo.com'
    app.config['MAIL_PORT'] = 587
    app.config['MAIL_USE_TLS'] = True
    app.config['MAIL_USERNAME'] = os.environ.get('BREVO_SENDER_EMAIL', '')
    app.config['MAIL_PASSWORD'] = os.environ.get('BREVO_API_KEY', '')
    app.config['MAIL_DEFAULT_SENDER'] = (
        os.environ.get('BREVO_SENDER_NAME', 'Law Ledger'),
        os.environ.get('BREVO_SENDER_EMAIL', 'noreply@lawledger.com'),
    )
else:
    app.config['MAIL_SERVER'] = os.environ.get('MAIL_SERVER', 'smtp.gmail.com')
    app.config['MAIL_PORT'] = int(os.environ.get('MAIL_PORT', '587'))
    app.config['MAIL_USE_TLS'] = os.environ.get('MAIL_USE_TLS', 'true').lower() != 'false'
    app.config['MAIL_USERNAME'] = os.environ.get('MAIL_USERNAME', '')
    app.config['MAIL_PASSWORD'] = os.environ.get('MAIL_PASSWORD', '')
    app.config['MAIL_DEFAULT_SENDER'] = os.environ.get('MAIL_DEFAULT_SENDER', 'noreply@lawledger.com')

# ── License configuration ─────────────────────────────────────────────────────
# LICENSE_FILE:       path to license.json (env var or config key)
# LICENSE_PUBLIC_KEY: base64url Ed25519 public key (env var or config key)
app.config.setdefault(
    'LICENSE_FILE',
    os.environ.get(
        'LICENSE_FILE',
        os.path.join(os.path.dirname(__file__), 'config', 'license.json'),
    ),
)
app.config.setdefault(
    'LICENSE_PUBLIC_KEY',
    os.environ.get('LICENSE_PUBLIC_KEY', '') or _licensing._DEFAULT_PUBLIC_KEY_B64,
)

# Upload folder for logos and other media
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'static', 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
ALLOWED_IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'svg', 'webp'}
ALLOWED_TRUST_AUTH_EXTENSIONS = {'.pdf', '.png', '.jpg', '.jpeg', '.gif', '.tiff', '.tif', '.bmp'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Folder for holding imported CSV/Excel files (server backup).
# Configured via IMPORT_FILES_DIR in .env; falls back to the OS data directory.
_programdata_base = (
    os.environ.get('PROGRAMDATA', r'C:\ProgramData')
    if os.name == 'nt'
    else os.path.join(os.path.expanduser('~'), '.local', 'share')
)
_default_import_files = os.path.join(_programdata_base, 'lawledger', 'import')
IMPORT_UPLOAD_FOLDER = os.environ.get('IMPORT_FILES_DIR', '').strip() or _default_import_files
try:
    os.makedirs(IMPORT_UPLOAD_FOLDER, exist_ok=True)
except OSError:
    # If the configured / system path is unusable, fall back to a directory next to the app.
    IMPORT_UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'import_uploads')
    try:
        os.makedirs(IMPORT_UPLOAD_FOLDER, exist_ok=True)
    except OSError as exc:
        logging.warning('Could not create import upload folder %s: %s', IMPORT_UPLOAD_FOLDER, exc)
app.config['IMPORT_UPLOAD_FOLDER'] = IMPORT_UPLOAD_FOLDER

# Folder for import log files.
# Configured via IMPORT_LOG_DIR in .env; falls back to the OS data directory.
_default_import_logs = (
    r'C:\programdata\lawledger\logs'
    if os.name == 'nt'
    else os.path.join(os.path.expanduser('~'), '.local', 'share', 'lawledger', 'logs')
)
IMPORT_LOG_FOLDER = os.environ.get('IMPORT_LOG_DIR', '').strip() or _default_import_logs
try:
    os.makedirs(IMPORT_LOG_FOLDER, exist_ok=True)
except OSError:
    IMPORT_LOG_FOLDER = os.path.join(os.path.dirname(__file__), 'logs')
    os.makedirs(IMPORT_LOG_FOLDER, exist_ok=True)
app.config['IMPORT_LOG_FOLDER'] = IMPORT_LOG_FOLDER

# Folder for trust authorization documents.
# Configured via TRUST_AUTH_DOCS_DIR in .env.
_default_trust_auth_docs = (
    r'C:\programdata\lawledger\trust_auth'
    if os.name == 'nt'
    else os.path.join(os.path.expanduser('~'), '.local', 'share', 'lawledger', 'trust_auth')
)
TRUST_AUTH_DOCS_FOLDER = os.environ.get('TRUST_AUTH_DOCS_DIR', '').strip() or _default_trust_auth_docs
try:
    os.makedirs(TRUST_AUTH_DOCS_FOLDER, exist_ok=True)
except OSError as exc:
    logging.warning('Could not create trust auth docs folder %s: %s', TRUST_AUTH_DOCS_FOLDER, exc)
app.config['TRUST_AUTH_DOCS_FOLDER'] = TRUST_AUTH_DOCS_FOLDER


# Initialize extensions
db = SQLAlchemy(app)
mail = Mail(app)

# Initialize Flask-Login
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Veuillez vous connecter pour accéder à cette page.'
login_manager.login_message_category = 'warning'

# Force UTF-8 charset in all HTML responses
@app.after_request
def set_charset(response):
    if response.mimetype == 'text/html' and 'charset' not in response.content_type:
        response.headers['Content-Type'] = 'text/html; charset=utf-8'
    return response


# Prevent browsers from caching protected pages so that pressing the back
# button after logout does not reveal stale content.  Static assets are
# excluded so their normal browser caching is preserved.
@app.after_request
def add_no_cache_headers(response):
    if not request.path.startswith('/static/'):
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response

# Token serializer for password reset
serializer = URLSafeTimedSerializer(app.config['SECRET_KEY'])


def send_reset_email(user, reset_url, lang='en'):
    """Send a password-reset email via Brevo API or Flask-Mail fallback."""
    t = LEXICON.get(lang, LEXICON.get('en'))
    html_content = render_template('email_reset_password.html', user=user, reset_url=reset_url, t=t, lang=lang)
    subject = t.get('email_reset_subject', 'LawLedger \u2013 Password Reset')
    brevo_api_key = os.environ.get('BREVO_API_KEY')
    if brevo_api_key:
        sender_email = os.environ.get('BREVO_SENDER_EMAIL', 'noreply@lawledger.com')
        sender_name = os.environ.get('BREVO_SENDER_NAME', 'Law Ledger')
        payload = json.dumps({
            'sender': {'name': sender_name, 'email': sender_email},
            'to': [{'email': user.email}],
            'subject': subject,
            'htmlContent': html_content,
            'replyTo': {'email': sender_email},
        }).encode('utf-8')
        req = urllib.request.Request(
            'https://api.brevo.com/v3/smtp/email',
            data=payload,
            headers={
                'Content-Type': 'application/json',
                'api-key': brevo_api_key,
            },
            method='POST',
        )
        try:
            with urllib.request.urlopen(req) as resp:
                logger.info('Brevo API response: %s', resp.status)
        except urllib.error.HTTPError as exc:
            body = exc.read().decode('utf-8', errors='replace')
            logger.error('Brevo API error %s: %s', exc.code, body)
            raise
    else:
        msg = Message(
            subject=subject,
            recipients=[user.email],
            html=html_content,
        )
        mail.send(msg)


def send_mfa_email(user, code, lang='en'):
    """Send an MFA verification code email via Brevo API or Flask-Mail fallback."""
    t = LEXICON.get(lang, LEXICON.get('en'))
    html_content = render_template('email_mfa_code.html', user=user, code=code, t=t, lang=lang)
    subject = t.get('email_mfa_subject', 'LawLedger \u2013 Verification Code')
    brevo_api_key = os.environ.get('BREVO_API_KEY')
    if brevo_api_key:
        sender_email = os.environ.get('BREVO_SENDER_EMAIL', 'noreply@lawledger.com')
        sender_name = os.environ.get('BREVO_SENDER_NAME', 'Law Ledger')
        payload = json.dumps({
            'sender': {'name': sender_name, 'email': sender_email},
            'to': [{'email': user.email}],
            'subject': subject,
            'htmlContent': html_content,
            'replyTo': {'email': sender_email},
        }).encode('utf-8')
        req = urllib.request.Request(
            'https://api.brevo.com/v3/smtp/email',
            data=payload,
            headers={
                'Content-Type': 'application/json',
                'api-key': brevo_api_key,
            },
            method='POST',
        )
        try:
            with urllib.request.urlopen(req) as resp:
                logger.info('Brevo API response (MFA): %s', resp.status)
        except urllib.error.HTTPError as exc:
            body = exc.read().decode('utf-8', errors='replace')
            logger.error('Brevo API error %s (MFA): %s', exc.code, body)
            raise
    else:
        msg = Message(
            subject=subject,
            recipients=[user.email],
            html=html_content,
        )
        mail.send(msg)


@login_manager.user_loader
def load_user(user_id):
    #return Employee.query.get(int(user_id))
    return db.session.get(Employee, int(user_id))


class Client(db.Model):
    __tablename__ = 'clients'

    id = db.Column(db.Integer, primary_key=True)
    client_number = db.Column(db.String(50), unique=True, nullable=False)
    client_name = db.Column(db.String(255), nullable=False)
    street = db.Column(db.String(255), nullable=True)
    city = db.Column(db.String(100), nullable=True)
    state = db.Column(db.String(100), nullable=True)
    postal_code = db.Column(db.String(20), nullable=True)
    country = db.Column(db.String(100), nullable=True)
    contact_name = db.Column(db.String(255), nullable=True)
    phone = db.Column(db.String(50), nullable=True)
    email = db.Column(db.String(255), nullable=True)
    is_active = db.Column(db.Boolean, default=True)
    is_deleted = db.Column(db.Boolean, default=False)
    accounting_code = db.Column(db.String(20), nullable=True, default='1100')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    created_by = db.Column(db.String(80), nullable=True)
    deleted_by = db.Column(db.String(80), nullable=True)
    disabled_by = db.Column(db.String(80), nullable=True)
    reenabled_by = db.Column(db.String(80), nullable=True)

    matters = db.relationship('Matter', backref='client', lazy=True, cascade='all, delete-orphan')

    def to_dict(self):
        return {
            'id': self.id,
            'client_number': self.client_number,
            'client_name': self.client_name,
            'street': self.street or '',
            'city': self.city or '',
            'state': self.state or '',
            'postal_code': self.postal_code or '',
            'country': self.country or '',
            'contact_name': self.contact_name or '',
            'phone': self.phone or '',
            'email': self.email or '',
            'is_active': self.is_active,
            'accounting_code': self.accounting_code or '1100',
            'created_by': self.created_by or '',
            'deleted_by': self.deleted_by or '',
            'disabled_by': self.disabled_by or '',
            'reenabled_by': self.reenabled_by or '',
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'updated_at': self.updated_at.isoformat() if self.updated_at else None
        }


class Matter(db.Model):
    __tablename__ = 'matters'

    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.Integer, db.ForeignKey('clients.id'), nullable=False)
    matter_number = db.Column(db.String(50), nullable=False)
    matter_description = db.Column(db.String(500))
    is_active = db.Column(db.Boolean, default=True)
    is_deleted = db.Column(db.Boolean, default=False)
    # Unique key for duplicate detection during import (client_number_matter_number)
    client_matter_key = db.Column(db.String(120), nullable=True, unique=True)
    # Attorney assignment fields (up to 6)
    attorney1_name = db.Column(db.String(255), nullable=True)
    attorney1_start_date = db.Column(db.Date, nullable=True)
    attorney2_name = db.Column(db.String(255), nullable=True)
    attorney2_start_date = db.Column(db.Date, nullable=True)
    attorney3_name = db.Column(db.String(255), nullable=True)
    attorney3_start_date = db.Column(db.Date, nullable=True)
    attorney4_name = db.Column(db.String(255), nullable=True)
    attorney4_start_date = db.Column(db.Date, nullable=True)
    attorney5_name = db.Column(db.String(255), nullable=True)
    attorney5_start_date = db.Column(db.Date, nullable=True)
    attorney6_name = db.Column(db.String(255), nullable=True)
    attorney6_start_date = db.Column(db.Date, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    created_by = db.Column(db.String(80), nullable=True)
    deleted_by = db.Column(db.String(80), nullable=True)
    disabled_by = db.Column(db.String(80), nullable=True)
    reenabled_by = db.Column(db.String(80), nullable=True)

    expenses = db.relationship('Expense', backref='matter', lazy=True, cascade='all, delete-orphan')
    invoices = db.relationship('Invoice', backref='matter', lazy=True)

    def to_dict(self):
        return {
            'id': self.id,
            'client_id': self.client_id,
            'matter_number': self.matter_number,
            'matter_description': self.matter_description,
            'is_active': self.is_active,
            'attorney1_name': self.attorney1_name or '',
            'attorney1_start_date': self.attorney1_start_date.isoformat() if self.attorney1_start_date else '',
            'attorney2_name': self.attorney2_name or '',
            'attorney2_start_date': self.attorney2_start_date.isoformat() if self.attorney2_start_date else '',
            'attorney3_name': self.attorney3_name or '',
            'attorney3_start_date': self.attorney3_start_date.isoformat() if self.attorney3_start_date else '',
            'attorney4_name': self.attorney4_name or '',
            'attorney4_start_date': self.attorney4_start_date.isoformat() if self.attorney4_start_date else '',
            'attorney5_name': self.attorney5_name or '',
            'attorney5_start_date': self.attorney5_start_date.isoformat() if self.attorney5_start_date else '',
            'attorney6_name': self.attorney6_name or '',
            'attorney6_start_date': self.attorney6_start_date.isoformat() if self.attorney6_start_date else '',
            'created_by': self.created_by or '',
            'deleted_by': self.deleted_by or '',
            'disabled_by': self.disabled_by or '',
            'reenabled_by': self.reenabled_by or '',
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'updated_at': self.updated_at.isoformat() if self.updated_at else None
        }


class CostCode(db.Model):
    __tablename__ = 'cost_codes'

    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(50), unique=True, nullable=False)
    description = db.Column(db.String(255), nullable=False)
    charge_type = db.Column(db.String(100), nullable=True)
    rate = db.Column(db.Numeric(10, 2), default=0.00)
    account_code = db.Column(db.String(20), nullable=True)
    is_active = db.Column(db.Boolean, default=True)
    changed_by = db.Column(db.String(80), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def to_dict(self):
        return {
            'id': self.id,
            'code': self.code,
            'description': self.description,
            'charge_type': self.charge_type or '',
            'rate': float(self.rate) if self.rate else 0.00,
            'account_code': self.account_code or '',
            'is_active': self.is_active if self.is_active is not None else True,
            'changed_by': self.changed_by or '',
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'updated_at': self.updated_at.isoformat() if self.updated_at else None
        }


class Expense(db.Model):
    __tablename__ = 'expenses'

    id = db.Column(db.Integer, primary_key=True)
    matter_id = db.Column(db.Integer, db.ForeignKey('matters.id'), nullable=False)
    code = db.Column(db.String(100), nullable=True)
    employee_id = db.Column(db.Integer, db.ForeignKey('employees.id'), nullable=True)
    username = db.Column(db.String(80), nullable=True)
    description = db.Column(db.String(500), nullable=False)
    amount = db.Column(db.Numeric(10, 2), default=0.00)
    expense_date = db.Column(db.Date, default=datetime.utcnow)
    is_billed = db.Column(db.Boolean, default=False)
    invoice_id = db.Column(db.Integer, db.ForeignKey('invoices.id'), nullable=True)
    invoice_number = db.Column(db.String(100), nullable=True)
    invoice_date = db.Column(db.Date, nullable=True)
    import_id = db.Column(db.String(32), nullable=True)
    quantity = db.Column(db.Numeric(10, 2), default=1.00)
    user_pin = db.Column(db.String(20), nullable=True)
    is_deleted = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    employee = db.relationship('Employee', backref='expenses', lazy=True)

    def to_dict(self):
        return {
            'id': self.id,
            'matter_id': self.matter_id,
            'code': self.code,
            'employee_id': self.employee_id,
            'username': self.username or (self.employee.username if self.employee else None),
            'first_name': self.employee.first_name if self.employee else None,
            'last_name': self.employee.last_name if self.employee else None,
            'description': self.description,
            'amount': float(self.amount) if self.amount else 0.00,
            'expense_date': self.expense_date.isoformat() if self.expense_date else None,
            'is_billed': self.is_billed,
            'invoice_id': self.invoice_id,
            'invoice_number': self.invoice_number,
            'invoice_date': self.invoice_date.isoformat() if self.invoice_date else None,
            'import_id': self.import_id,
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'updated_at': self.updated_at.isoformat() if self.updated_at else None
        }


class Employee(UserMixin, db.Model):
    __tablename__ = 'employees'

    id = db.Column(db.Integer, primary_key=True)
    first_name = db.Column(db.String(100), nullable=True)
    last_name = db.Column(db.String(100), nullable=True)
    username = db.Column(db.String(80), unique=True, nullable=True)
    email = db.Column(db.String(255), nullable=True)
    personal_email = db.Column(db.String(255), nullable=True)
    password_hash = db.Column(db.String(255), nullable=True)
    title = db.Column(db.String(100))
    address = db.Column(db.String(500), nullable=True)
    phone_number = db.Column(db.String(50), nullable=True)
    social_insurance_number = db.Column(db.String(20))
    salary_type = db.Column(db.String(20))
    salary = db.Column(db.Numeric(10, 2), default=0.00)
    hiring_date = db.Column(db.Date, nullable=True)
    leave_date = db.Column(db.Date, nullable=True)
    emergency_contact = db.Column(db.String(255), nullable=True)
    emergency_phone = db.Column(db.String(50), nullable=True)
    notes = db.Column(db.Text, nullable=True)
    is_active = db.Column(db.Boolean, default=True)
    is_deleted = db.Column(db.Boolean, default=False)
    is_manager = db.Column(db.Boolean, default=False)
    is_user = db.Column(db.Boolean, default=False)
    timer_user = db.Column(db.Boolean, default=False)
    is_accounting = db.Column(db.Boolean, default=False)
    hourly_rate = db.Column(db.Numeric(10, 2), nullable=True)
    pin = db.Column(db.String(20), nullable=True)
    group_name = db.Column(db.String(100), nullable=True)
    network_id = db.Column(db.String(100), nullable=True)
    office_phone = db.Column(db.String(50), nullable=True)
    supervisor = db.Column(db.String(255), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    # Session management: single-session enforcement and inactivity tracking
    session_token = db.Column(db.String(36), nullable=True)
    last_login = db.Column(db.DateTime, nullable=True)
    login_ip = db.Column(db.String(45), nullable=True)
    must_change_password = db.Column(db.Boolean, default=False)
    added_by = db.Column(db.String(80), nullable=True)
    changed_by = db.Column(db.String(80), nullable=True)
    deleted_by = db.Column(db.String(80), nullable=True)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        if not self.password_hash:
            return False
        try:
            return check_password_hash(self.password_hash, password)
        except (ValueError, TypeError):
            return False

    @property
    def display_name(self):
        """Return the employee's full name, falling back to username."""
        name = f"{self.first_name or ''} {self.last_name or ''}".strip()
        return name or self.username

    def get_reset_token(self):
        return serializer.dumps(self.email, salt='password-reset')

    @staticmethod
    def verify_reset_token(token, max_age=3600):
        try:
            email = serializer.loads(token, salt='password-reset', max_age=max_age)
        except (SignatureExpired, BadSignature):
            return None
        return Employee.query.filter_by(email=email).first()

    def to_dict(self):
        sin = self.social_insurance_number
        masked_sin = ('***-***-' + sin[-3:]) if sin and len(sin) >= 3 else ('***' if sin else None)
        return {
            'id': self.id,
            'first_name': self.first_name,
            'last_name': self.last_name,
            'username': self.username,
            'email': self.email,
            'personal_email': self.personal_email,
            'title': self.title,
            'address': self.address,
            'phone_number': self.phone_number,
            'social_insurance_number': masked_sin,
            'salary_type': self.salary_type,
            'salary': float(self.salary) if self.salary else 0.00,
            'hiring_date': self.hiring_date.isoformat() if self.hiring_date else None,
            'leave_date': self.leave_date.isoformat() if self.leave_date else None,
            'emergency_contact': self.emergency_contact,
            'emergency_phone': self.emergency_phone,
            'notes': self.notes,
            'is_active': self.is_active,
            'is_manager': bool(self.is_manager),
            'is_user': bool(self.is_user),
            'timer_user': bool(self.timer_user),
            'is_accounting': bool(self.is_accounting),
            'hourly_rate': float(self.hourly_rate) if self.hourly_rate else None,
            'pin': self.pin,
            'group_name': self.group_name,
            'network_id': self.network_id,
            'office_phone': self.office_phone,
            'supervisor': self.supervisor,
            'added_by': self.added_by or '',
            'changed_by': self.changed_by or '',
            'deleted_by': self.deleted_by or '',
            'created_at': self.created_at.isoformat() if self.created_at else None
        }


class ImportLog(db.Model):
    __tablename__ = 'import_logs'

    id = db.Column(db.Integer, primary_key=True)
    import_id = db.Column(db.String(32), nullable=True)
    filename = db.Column(db.String(255), nullable=False)
    file_hash = db.Column(db.String(64), nullable=True)
    import_date = db.Column(db.DateTime, default=datetime.utcnow)
    records_imported = db.Column(db.Integer, default=0)
    records_failed = db.Column(db.Integer, default=0)
    status = db.Column(db.String(50), default='success')
    error_message = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def to_dict(self):
        return {
            'id': self.id,
            'import_id': self.import_id,
            'filename': self.filename,
            'file_hash': self.file_hash,
            'import_date': self.import_date.isoformat() if self.import_date else None,
            'records_imported': self.records_imported,
            'records_failed': self.records_failed,
            'status': self.status,
            'error_message': self.error_message,
            'created_at': self.created_at.isoformat() if self.created_at else None
        }


class Invoice(db.Model):
    __tablename__ = 'invoices'

    id = db.Column(db.Integer, primary_key=True)
    matter_id = db.Column(db.Integer, db.ForeignKey('matters.id'), nullable=True)
    client_id = db.Column(db.Integer, db.ForeignKey('clients.id'), nullable=True)
    invoice_number = db.Column(db.String(100), unique=True, nullable=False)
    invoice_date = db.Column(db.Date, nullable=False)
    due_date = db.Column(db.Date, nullable=True)
    subtotal = db.Column(db.Numeric(12, 2), default=0.00)
    gst_rate = db.Column(db.Numeric(6, 3), default=5.000)
    gst_amount = db.Column(db.Numeric(12, 2), default=0.00)
    qst_rate = db.Column(db.Numeric(6, 3), default=9.975)
    qst_amount = db.Column(db.Numeric(12, 2), default=0.00)
    total_amount = db.Column(db.Numeric(12, 2), default=0.00)
    credit_applied = db.Column(db.Numeric(12, 2), default=0.00)
    trust_applied = db.Column(db.Numeric(12, 2), default=0.00)
    status = db.Column(db.String(20), default='draft')
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    expenses = db.relationship('Expense', backref='invoice', lazy=True,
                                foreign_keys='Expense.invoice_id')
    client_ref = db.relationship('Client', foreign_keys=[client_id], lazy=True)

    @property
    def resolved_client(self):
        """Return the Client for this invoice regardless of how it was created."""
        if self.matter:
            return self.matter.client
        return self.client_ref

    def to_dict(self):
        matter = self.matter
        client = self.resolved_client
        return {
            'id': self.id,
            'matter_id': self.matter_id,
            'client_id': self.client_id or (matter.client_id if matter else None),
            'matter_number': matter.matter_number if matter else None,
            'client_name': client.client_name if client else None,
            'client_number': client.client_number if client else None,
            'invoice_number': self.invoice_number,
            'invoice_date': self.invoice_date.isoformat() if self.invoice_date else None,
            'due_date': self.due_date.isoformat() if self.due_date else None,
            'subtotal': float(self.subtotal) if self.subtotal else 0.00,
            'gst_rate': float(self.gst_rate) if self.gst_rate else 0.00,
            'gst_amount': float(self.gst_amount) if self.gst_amount else 0.00,
            'qst_rate': float(self.qst_rate) if self.qst_rate else 0.00,
            'qst_amount': float(self.qst_amount) if self.qst_amount else 0.00,
            'total_amount': float(self.total_amount) if self.total_amount else 0.00,
            'credit_applied': float(self.credit_applied) if self.credit_applied else 0.00,
            'trust_applied': float(self.trust_applied) if self.trust_applied else 0.00,
            'status': self.status,
            'notes': self.notes,
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'updated_at': self.updated_at.isoformat() if self.updated_at else None
        }


class FirmInfo(db.Model):
    __tablename__ = 'firm_info'

    id = db.Column(db.Integer, primary_key=True)
    firm_name = db.Column(db.String(255), nullable=False, default='Your Law Firm')
    address_line1 = db.Column(db.String(255))
    address_line2 = db.Column(db.String(255))
    city = db.Column(db.String(100))
    province = db.Column(db.String(100))
    postal_code = db.Column(db.String(20))
    phone = db.Column(db.String(50))
    email = db.Column(db.String(255))
    tax_number = db.Column(db.String(100))
    logo_filename = db.Column(db.String(255))
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    # Configurable tax labels (empty/NULL = hide that tax on invoices)
    tax1_name = db.Column(db.String(100), default='GST')
    tax2_name = db.Column(db.String(100), default=None)
    # When True, tax2 is calculated on (subtotal + tax1) instead of subtotal alone
    tax2_compound = db.Column(db.Boolean, default=False)
    # Tax rates (%) stored here so invoices use consistent firm-wide rates
    tax1_rate = db.Column(db.Numeric(6, 3), default=0.000)
    tax2_rate = db.Column(db.Numeric(6, 3), default=0.000)
    # Multi-factor authentication: when True, users must verify a 6-digit email code after login
    mfa_enabled = db.Column(db.Boolean, default=False)

    def to_dict(self):
        return {
            'id': self.id,
            'firm_name': self.firm_name or '',
            'address_line1': self.address_line1 or '',
            'address_line2': self.address_line2 or '',
            'city': self.city or '',
            'province': self.province or '',
            'postal_code': self.postal_code or '',
            'phone': self.phone or '',
            'email': self.email or '',
            'tax_number': self.tax_number or '',
            'logo_filename': self.logo_filename or '',
            'tax1_name': self.tax1_name or 'GST',
            'tax2_name': self.tax2_name or '',
            'tax2_compound': bool(self.tax2_compound) if self.tax2_compound is not None else False,
            'tax1_rate': float(self.tax1_rate) if self.tax1_rate is not None else 0.0,
            'tax2_rate': float(self.tax2_rate) if self.tax2_rate is not None else 0.0,
            'mfa_enabled': bool(self.mfa_enabled) if self.mfa_enabled is not None else False,
        }


class SecondWork(db.Model):
    """Records the raw seconds worked per timer session, linked to a matter and employee."""
    __tablename__ = 'second_works'

    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey('employees.id'), nullable=False)
    matter_id = db.Column(db.Integer, db.ForeignKey('matters.id'), nullable=False)
    seconds_worked = db.Column(db.Integer, nullable=False)
    expense_id = db.Column(db.Integer, db.ForeignKey('expenses.id'), nullable=True)
    recorded_at = db.Column(db.DateTime, default=datetime.utcnow)

    def to_dict(self):
        return {
            'id': self.id,
            'employee_id': self.employee_id,
            'matter_id': self.matter_id,
            'seconds_worked': self.seconds_worked,
            'expense_id': self.expense_id,
            'recorded_at': self.recorded_at.isoformat() if self.recorded_at else None,
        }


class HrRecord(db.Model):
    """HR record for an employee: PTO balance, last review date, and review comment."""
    __tablename__ = 'hr_records'

    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey('employees.id'), nullable=False)
    balance_pto = db.Column(db.Numeric(8, 2), default=0.00)
    date_last_review = db.Column(db.Date, nullable=True)
    review_comment = db.Column(db.Text, nullable=True)
    is_deleted = db.Column(db.Boolean, default=False)
    changed_by = db.Column(db.String(80), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    employee = db.relationship('Employee', backref=db.backref('hr_records', lazy=True))

    def to_dict(self):
        return {
            'id': self.id,
            'employee_id': self.employee_id,
            'balance_pto': float(self.balance_pto) if self.balance_pto is not None else 0.0,
            'date_last_review': self.date_last_review.isoformat() if self.date_last_review else None,
            'review_comment': self.review_comment,
            'changed_by': self.changed_by or '',
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'updated_at': self.updated_at.isoformat() if self.updated_at else None,
        }


class ActiveTimerSession(db.Model):
    """Tracks a running timer session for an employee.

    Enforces the one-timer-at-a-time rule: at most one row per employee_id
    (enforced by the unique constraint on the column).  Rows older than 24 hours
    are considered stale and are replaced automatically when a new session starts.
    """
    __tablename__ = 'active_timer_sessions'

    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey('employees.id'),
                            nullable=False, unique=True)
    matter_id = db.Column(db.Integer, db.ForeignKey('matters.id'), nullable=False)
    session_token = db.Column(db.String(36), nullable=False)
    started_at = db.Column(db.DateTime, default=datetime.utcnow)


class CreditNote(db.Model):
    """Credit note that can be applied to a client's next invoice."""
    __tablename__ = 'credit_notes'

    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.Integer, db.ForeignKey('clients.id'), nullable=False)
    amount = db.Column(db.Numeric(12, 2), nullable=False)
    applied_amount = db.Column(db.Numeric(12, 2), default=0.00)
    reason = db.Column(db.String(500))
    applied_invoice_id = db.Column(db.Integer, db.ForeignKey('invoices.id'), nullable=True)
    added_by = db.Column(db.String(80), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    client = db.relationship('Client', backref='credit_notes', lazy=True)
    applied_invoice = db.relationship('Invoice', lazy=True)

    @property
    def remaining(self):
        return float(self.amount or 0) - float(self.applied_amount or 0)

    def to_dict(self):
        return {
            'id': self.id,
            'client_id': self.client_id,
            'client_name': self.client.client_name if self.client else None,
            'amount': float(self.amount) if self.amount else 0.00,
            'applied_amount': float(self.applied_amount) if self.applied_amount else 0.00,
            'remaining': self.remaining,
            'reason': self.reason,
            'applied_invoice_id': self.applied_invoice_id,
            'applied_invoice_number': self.applied_invoice.invoice_number if self.applied_invoice else None,
            'added_by': self.added_by or '',
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'updated_at': self.updated_at.isoformat() if self.updated_at else None,
        }


class Supplier(db.Model):
    """Supplier (fournisseur) with contact information and service details."""
    __tablename__ = 'suppliers'

    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(255), nullable=False)
    account_number = db.Column(db.String(100), nullable=True)
    address = db.Column(db.String(500), nullable=True)
    phone = db.Column(db.String(50), nullable=True)
    email = db.Column(db.String(255), nullable=True)
    service_provided = db.Column(db.String(500), nullable=True)
    notes = db.Column(db.Text, nullable=True)
    is_active = db.Column(db.Boolean, default=True)
    is_deleted = db.Column(db.Boolean, default=False)
    accounting_code = db.Column(db.String(20), nullable=True)
    added_by = db.Column(db.String(80), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    payments = db.relationship('SupplierPayment', backref='supplier', lazy=True,
                                foreign_keys='SupplierPayment.supplier_id')

    def to_dict(self):
        return {
            'id': self.id,
            'name': self.name,
            'account_number': self.account_number or '',
            'address': self.address or '',
            'phone': self.phone or '',
            'email': self.email or '',
            'service_provided': self.service_provided or '',
            'notes': self.notes or '',
            'is_active': bool(self.is_active),
            'accounting_code': self.accounting_code or '',
            'added_by': self.added_by or '',
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'updated_at': self.updated_at.isoformat() if self.updated_at else None,
        }


class SupplierPayment(db.Model):
    """Payment made to a supplier, tracking invoice number and payment method."""
    __tablename__ = 'supplier_payments'

    id = db.Column(db.Integer, primary_key=True)
    supplier_id = db.Column(db.Integer, db.ForeignKey('suppliers.id'), nullable=False)
    invoice_number = db.Column(db.String(100), nullable=True)
    invoice_date = db.Column(db.Date, nullable=True)
    amount = db.Column(db.Numeric(12, 2), nullable=False, default=0.00)
    description = db.Column(db.String(500), nullable=True)
    payment_date = db.Column(db.Date, nullable=True)
    payment_method = db.Column(db.String(50), nullable=True)  # cheque, virement, etc.
    cheque_number = db.Column(db.String(100), nullable=True)
    bank_transaction = db.Column(db.String(255), nullable=True)
    created_by = db.Column(db.String(80), nullable=True)
    is_deleted = db.Column(db.Boolean, default=False)
    deleted_by = db.Column(db.String(80), nullable=True)   # who soft-deleted this payment
    is_paid = db.Column(db.Boolean, default=False)   # explicitly marked paid by a user
    date_paid = db.Column(db.Date, nullable=True)    # date when invoice was marked as paid
    paid_by = db.Column(db.String(80), nullable=True)   # who marked as paid
    is_posted = db.Column(db.Boolean, default=False)
    posted_by = db.Column(db.String(80), nullable=True)  # who posted to GL
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def to_dict(self):
        return {
            'id': self.id,
            'supplier_id': self.supplier_id,
            'supplier_name': self.supplier.name if self.supplier else None,
            'invoice_number': self.invoice_number or '',
            'invoice_date': self.invoice_date.isoformat() if self.invoice_date else None,
            'amount': float(self.amount) if self.amount else 0.00,
            'description': self.description or '',
            'payment_date': self.payment_date.isoformat() if self.payment_date else None,
            'payment_method': self.payment_method or '',
            'cheque_number': self.cheque_number or '',
            'bank_transaction': self.bank_transaction or '',
            'created_by': self.created_by or '',
            'is_paid': bool(self.is_paid),
            'date_paid': self.date_paid.isoformat() if self.date_paid else None,
            'paid_by': self.paid_by or '',
            'is_posted': bool(self.is_posted),
            'posted_by': self.posted_by or '',
            'deleted_by': self.deleted_by or '',
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'updated_at': self.updated_at.isoformat() if self.updated_at else None,
        }


# ── Chart of Accounts / Journal models (double-entry GL) ──────────────────────

class Account(db.Model):
    """Chart of accounts (plan de comptes).

    Accounts are organised into a hierarchy via parent_id and categorised by
    type: asset, liability, equity, revenue, expense.
    """
    __tablename__ = 'GL_accounts'

    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(20), unique=True, nullable=False)
    name = db.Column(db.String(255), nullable=False)
    # asset | liability | equity | revenue | expense
    account_type = db.Column(db.String(20), nullable=False, default='expense')
    parent_id = db.Column(db.Integer, db.ForeignKey('GL_accounts.id'), nullable=True)
    is_active = db.Column(db.Boolean, default=True)
    is_system = db.Column(db.Boolean, default=False)   # protected from deletion
    is_deleted = db.Column(db.Boolean, default=False)  # soft-delete flag
    changed_by = db.Column(db.String(80), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    parent = db.relationship('Account', remote_side=[id], backref='children')
    journal_lines = db.relationship('JournalLine', backref='account', lazy=True)

    def to_dict(self):
        return {
            'id': self.id,
            'code': self.code,
            'name': self.name,
            'account_type': self.account_type,
            'parent_id': self.parent_id,
            'is_active': self.is_active,
            'is_system': self.is_system,
            'is_deleted': self.is_deleted,
            'changed_by': self.changed_by or '',
        }


class JournalEntry(db.Model):
    """Header record for a double-entry journal posting."""
    __tablename__ = 'journal_entries'

    id = db.Column(db.Integer, primary_key=True)
    entry_date = db.Column(db.Date, nullable=False, default=datetime.utcnow)
    description = db.Column(db.String(500), nullable=True)
    # invoice | trust_deposit | trust_withdrawal | trust_payment | manual
    source_type = db.Column(db.String(50), nullable=True)
    source_id = db.Column(db.Integer, nullable=True)
    is_reversed = db.Column(db.Boolean, default=False)
    reversed_by_id = db.Column(db.Integer, db.ForeignKey('journal_entries.id'), nullable=True)
    created_by = db.Column(db.String(255), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    lines = db.relationship('JournalLine', backref='entry', lazy=True,
                            cascade='all, delete-orphan')
    reversal_entry = db.relationship('JournalEntry', remote_side=[id])

    def to_dict(self):
        return {
            'id': self.id,
            'entry_date': self.entry_date.isoformat() if self.entry_date else None,
            'description': self.description or '',
            'source_type': self.source_type or '',
            'source_id': self.source_id,
            'is_reversed': self.is_reversed,
            'reversed_by_id': self.reversed_by_id,
            'created_by': self.created_by or '',
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'lines': [l.to_dict() for l in self.lines],
        }


class JournalLine(db.Model):
    """Individual debit or credit line within a journal entry."""
    __tablename__ = 'journal_lines'

    id = db.Column(db.Integer, primary_key=True)
    entry_id = db.Column(db.Integer, db.ForeignKey('journal_entries.id'), nullable=False)
    account_id = db.Column(db.Integer, db.ForeignKey('GL_accounts.id'), nullable=False)
    debit = db.Column(db.Numeric(12, 2), default=0.00)
    credit = db.Column(db.Numeric(12, 2), default=0.00)
    client_id = db.Column(db.Integer, db.ForeignKey('clients.id'), nullable=True)
    matter_id = db.Column(db.Integer, db.ForeignKey('matters.id'), nullable=True)
    is_trust = db.Column(db.Boolean, default=False)
    memo = db.Column(db.String(500), nullable=True)

    def to_dict(self):
        return {
            'id': self.id,
            'entry_id': self.entry_id,
            'account_id': self.account_id,
            'account_code': self.account.code if self.account else '',
            'account_name': self.account.name if self.account else '',
            'account_type': self.account.account_type if self.account else '',
            'debit': float(self.debit or 0),
            'credit': float(self.credit or 0),
            'client_id': self.client_id,
            'matter_id': self.matter_id,
            'is_trust': self.is_trust,
            'memo': self.memo or '',
        }


class TrustReconciliation(db.Model):
    """Monthly trust-account reconciliation record.

    Stores the bank-statement balance entered by the user and the computed GL
    balance so that discrepancies are tracked over time.
    """
    __tablename__ = 'trust_reconciliations'

    id = db.Column(db.Integer, primary_key=True)
    statement_date = db.Column(db.Date, nullable=False)
    bank_balance = db.Column(db.Numeric(12, 2), nullable=False, default=0.00)
    gl_balance = db.Column(db.Numeric(12, 2), nullable=False, default=0.00)
    difference = db.Column(db.Numeric(12, 2), nullable=False, default=0.00)
    notes = db.Column(db.Text, nullable=True)
    created_by = db.Column(db.String(255), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def to_dict(self):
        return {
            'id': self.id,
            'statement_date': self.statement_date.isoformat() if self.statement_date else None,
            'bank_balance': float(self.bank_balance or 0),
            'gl_balance': float(self.gl_balance or 0),
            'difference': float(self.difference or 0),
            'notes': self.notes or '',
            'created_by': self.created_by or '',
            'created_at': self.created_at.isoformat() if self.created_at else None,
        }


class TrustAuthorization(db.Model):
    """Signed client authorization allowing the firm to take funds from trust.

    Authorization is per-matter and can cover a specific date range (date_from /
    date_to) or be indefinite (is_indefinite = True / date_to = NULL).  A PDF/image
    of the signed document is stored on disk under TRUST_AUTH_DOCS_FOLDER.
    """
    __tablename__ = 'trust_authorizations'

    id = db.Column(db.Integer, primary_key=True)
    matter_id = db.Column(db.Integer, db.ForeignKey('matters.id'), nullable=False)
    client_id = db.Column(db.Integer, db.ForeignKey('clients.id'), nullable=True)
    date_from = db.Column(db.Date, nullable=True)
    date_to = db.Column(db.Date, nullable=True)   # NULL = indefinite
    is_indefinite = db.Column(db.Boolean, default=False)
    max_amount = db.Column(db.Numeric(12, 2), nullable=True)
    notes = db.Column(db.Text, nullable=True)
    doc_filename = db.Column(db.String(500), nullable=True)
    doc_original_name = db.Column(db.String(255), nullable=True)
    is_active = db.Column(db.Boolean, default=True)
    is_deleted = db.Column(db.Boolean, default=False)    # explicit soft-delete flag
    deleted_at = db.Column(db.DateTime, nullable=True)   # NULL = not deleted (soft-delete timestamp)
    deleted_by = db.Column(db.String(255), nullable=True)  # who performed the soft-delete
    created_by = db.Column(db.String(255), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    matter = db.relationship('Matter', backref='trust_authorizations', lazy=True)

    def is_active_on(self, check_date=None):
        """Return True if this authorization is valid on *check_date* (today by default).

        An authorization is valid only when:
        - It has not been soft-deleted (``deleted_at`` is NULL *and* ``is_active`` is True)
        - Today falls within the configured date range (``date_from`` / ``date_to``)
        """
        # Treat deleted_at IS NOT NULL (explicit deletion timestamp) as soft-deleted.
        # Also honour the legacy is_active flag in case deleted_at was not yet set,
        # and the explicit is_deleted flag added later.
        if self.deleted_at is not None or not self.is_active or bool(self.is_deleted):
            return False
        if check_date is None:
            check_date = datetime.now(UTC).date()
        if self.date_from and check_date < self.date_from:
            return False
        if self.date_to and check_date > self.date_to:
            return False
        return True

    def to_dict(self):
        return {
            'id': self.id,
            'matter_id': self.matter_id,
            'client_id': self.client_id,
            'date_from': self.date_from.isoformat() if self.date_from else None,
            'date_to': self.date_to.isoformat() if self.date_to else None,
            'is_indefinite': bool(self.is_indefinite) if self.is_indefinite is not None else False,
            'max_amount': float(self.max_amount) if self.max_amount is not None else None,
            'notes': self.notes or '',
            'doc_filename': self.doc_filename or '',
            'doc_original_name': self.doc_original_name or '',
            'created_by': self.created_by or '',
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'is_active': bool(self.is_active) and self.deleted_at is None and not bool(self.is_deleted),
            'is_deleted': bool(self.is_deleted),
            'deleted_at': self.deleted_at.isoformat() if self.deleted_at else None,
            'deleted_by': self.deleted_by or '',
            'is_valid': self.is_active_on(),
        }


class CalendarEvent(db.Model):
    """Calendar event (hearing, deadline, meeting…) optionally linked to a matter."""
    __tablename__ = 'calendar_events'

    id = db.Column(db.Integer, primary_key=True)
    matter_id = db.Column(db.Integer, db.ForeignKey('matters.id'), nullable=True)
    title = db.Column(db.String(255), nullable=False)
    event_type = db.Column(db.String(50), nullable=True)   # hearing | deadline | meeting | other
    event_date = db.Column(db.Date, nullable=False)
    event_time = db.Column(db.String(10), nullable=True)   # HH:MM (optional)
    location = db.Column(db.String(255), nullable=True)
    notes = db.Column(db.Text, nullable=True)
    is_done = db.Column(db.Boolean, default=False)
    assigned_to = db.Column(db.String(80), nullable=True)
    created_by = db.Column(db.String(80), nullable=True)
    is_deleted = db.Column(db.Boolean, default=False)
    room_id = db.Column(db.Integer, db.ForeignKey('room_configs.id'), nullable=True)
    duration_minutes = db.Column(db.Integer, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    matter = db.relationship('Matter', backref='calendar_events', lazy=True)
    room = db.relationship('RoomConfig', backref='reservations', lazy=True)

    def to_dict(self):
        client_number = None
        matter_number = None
        if self.matter:
            matter_number = self.matter.matter_number
            if self.matter.client:
                client_number = self.matter.client.client_number
        return {
            'id': self.id,
            'matter_id': self.matter_id,
            'client_number': client_number or '',
            'matter_number': matter_number or '',
            'title': self.title,
            'event_type': self.event_type or '',
            'event_date': self.event_date.isoformat() if self.event_date else None,
            'event_time': self.event_time or '',
            'location': self.location or '',
            'notes': self.notes or '',
            'is_done': bool(self.is_done),
            'assigned_to': self.assigned_to or '',
            'created_by': self.created_by or '',
            'room_id': self.room_id,
            'room_name': self.room.room_name if self.room else '',
            'duration_minutes': self.duration_minutes,
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'updated_at': self.updated_at.isoformat() if self.updated_at else None,
        }


class RoomConfig(db.Model):
    """Configuration for a meeting/hearing room (up to 10 per firm).

    Each entry has a user-defined room name and can be toggled active/inactive.
    These rooms will be used for future room-reservation features.
    """
    __tablename__ = 'room_configs'

    id = db.Column(db.Integer, primary_key=True)
    room_index = db.Column(db.Integer, nullable=False, unique=True)   # 1-10
    room_name = db.Column(db.String(255), nullable=False)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def to_dict(self):
        return {
            'id': self.id,
            'room_index': self.room_index,
            'room_name': self.room_name,
            'is_active': bool(self.is_active),
        }


class SalaryConfig(db.Model):
    """Configuration for a salary field (up to 10 per firm).

    Each field has a user-defined label and an associated accounting code.
    """
    __tablename__ = 'salary_configs'

    id = db.Column(db.Integer, primary_key=True)
    field_index = db.Column(db.Integer, nullable=False)   # 1-10
    field_name = db.Column(db.String(255), nullable=False)
    account_code = db.Column(db.String(20), nullable=True)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    entries = db.relationship('SalaryEntry', backref='config', lazy=True,
                               cascade='all, delete-orphan')

    def to_dict(self):
        return {
            'id': self.id,
            'field_index': self.field_index,
            'field_name': self.field_name,
            'account_code': self.account_code or '',
            'is_active': bool(self.is_active),
        }


class SalaryEntry(db.Model):
    """A monetary entry for one salary configuration field."""
    __tablename__ = 'salary_entries'

    id = db.Column(db.Integer, primary_key=True)
    config_id = db.Column(db.Integer, db.ForeignKey('salary_configs.id'), nullable=False)
    entry_date = db.Column(db.Date, nullable=False)
    amount = db.Column(db.Numeric(12, 2), nullable=False, default=0.00)
    description = db.Column(db.String(500), nullable=True)
    created_by = db.Column(db.String(80), nullable=True)
    is_deleted = db.Column(db.Boolean, default=False)
    is_posted = db.Column(db.Boolean, default=False)
    posted_by = db.Column(db.String(80), nullable=True)   # who posted to GL
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def to_dict(self):
        return {
            'id': self.id,
            'config_id': self.config_id,
            'field_name': self.config.field_name if self.config else '',
            'account_code': self.config.account_code if self.config else '',
            'entry_date': self.entry_date.isoformat() if self.entry_date else None,
            'amount': float(self.amount) if self.amount else 0.00,
            'description': self.description or '',
            'created_by': self.created_by or '',
            'is_posted': bool(self.is_posted),
            'posted_by': self.posted_by or '',
            'created_at': self.created_at.isoformat() if self.created_at else None,
        }


# ── Schema migrations (no Alembic – add missing columns on startup) ───────────

# Columns that may be missing from existing tables after a schema upgrade.
# Each entry is (column_name, SQL Server column definition).
_COLUMN_MIGRATIONS = {
    'cost_codes': [
        ('charge_type',   'NVARCHAR(100) NULL'),
        ('is_active',     'BIT NOT NULL DEFAULT 1'),
        ('changed_by',    'NVARCHAR(80) NULL'),
        ('account_code',  'NVARCHAR(20) NULL'),
    ],
    'clients': [
        ('is_active',    'BIT NOT NULL DEFAULT 1'),
        ('is_deleted',   'BIT NOT NULL DEFAULT 0'),
        ('address',      'NVARCHAR(500) NULL'),
        ('phone',        'NVARCHAR(50) NULL'),
        ('email',        'NVARCHAR(255) NULL'),
        ('street',       'NVARCHAR(255) NULL'),
        ('city',         'NVARCHAR(100) NULL'),
        ('state',        'NVARCHAR(100) NULL'),
        ('postal_code',  'NVARCHAR(20) NULL'),
        ('country',      'NVARCHAR(100) NULL'),
        ('contact_name', 'NVARCHAR(255) NULL'),
        ('accounting_code', 'NVARCHAR(20) NULL DEFAULT \'1100\''),
        ('created_by',   'NVARCHAR(80) NULL'),
        ('deleted_by',   'NVARCHAR(80) NULL'),
        ('disabled_by',  'NVARCHAR(80) NULL'),
        ('reenabled_by', 'NVARCHAR(80) NULL'),
    ],
    'matters': [
        ('is_active',            'BIT NOT NULL DEFAULT 1'),
        ('is_deleted',           'BIT NOT NULL DEFAULT 0'),
        ('attorney1_name',       'NVARCHAR(255) NULL'),
        ('attorney1_start_date', 'DATE NULL'),
        ('attorney2_name',       'NVARCHAR(255) NULL'),
        ('attorney2_start_date', 'DATE NULL'),
        ('attorney3_name',       'NVARCHAR(255) NULL'),
        ('attorney3_start_date', 'DATE NULL'),
        ('attorney4_name',       'NVARCHAR(255) NULL'),
        ('attorney4_start_date', 'DATE NULL'),
        ('attorney5_name',       'NVARCHAR(255) NULL'),
        ('attorney5_start_date', 'DATE NULL'),
        ('attorney6_name',       'NVARCHAR(255) NULL'),
        ('attorney6_start_date', 'DATE NULL'),
        ('client_matter_key',    'NVARCHAR(120) NULL'),
        ('created_by',           'NVARCHAR(80) NULL'),
        ('deleted_by',           'NVARCHAR(80) NULL'),
        ('disabled_by',          'NVARCHAR(80) NULL'),
        ('reenabled_by',         'NVARCHAR(80) NULL'),
    ],
    'expenses': [
        ('is_deleted',   'BIT NOT NULL DEFAULT 0'),
        ('employee_id',    'INT NULL'),
        ('username',       'NVARCHAR(80) NULL'),
        ('code',           'NVARCHAR(100) NULL'),
        ('import_id',      'NVARCHAR(32) NULL'),
        ('invoice_number', 'NVARCHAR(100) NULL'),
        ('invoice_date',   'DATE NULL'),
        ('quantity',       'DECIMAL(10,2) NOT NULL DEFAULT 1'),
        ('user_pin',       'NVARCHAR(20) NULL'),
    ],
    'employees': [
        ('is_active',               'BIT NOT NULL DEFAULT 1'),
        ('is_deleted',              'BIT NOT NULL DEFAULT 0'),
        ('is_manager',              'BIT NOT NULL DEFAULT 0'),
        ('is_user',                 'BIT NOT NULL DEFAULT 0'),
        ('timer_user',              'BIT NOT NULL DEFAULT 0'),
        ('is_accounting',           'BIT NOT NULL DEFAULT 0'),
        ('hourly_rate',             'DECIMAL(10,2) NULL'),
        ('first_name',              'NVARCHAR(100) NULL'),
        ('last_name',               'NVARCHAR(100) NULL'),
        ('title',                   'NVARCHAR(100) NULL'),
        ('address',                 'NVARCHAR(500) NULL'),
        ('phone_number',            'NVARCHAR(50) NULL'),
        ('emergency_contact',       'NVARCHAR(255) NULL'),
        ('emergency_phone',         'NVARCHAR(50) NULL'),
        ('notes',                   'NVARCHAR(MAX) NULL'),
        ('salary_type',             'NVARCHAR(20) NULL'),
        ('hiring_date',             'DATE NULL'),
        ('leave_date',              'DATE NULL'),
        ('social_insurance_number', 'NVARCHAR(20) NULL'),
        ('personal_email',          'NVARCHAR(255) NULL'),
        ('pin',                     'NVARCHAR(20) NULL'),
        ('group_name',              'NVARCHAR(100) NULL'),
        ('network_id',              'NVARCHAR(100) NULL'),
        ('office_phone',            'NVARCHAR(50) NULL'),
        ('supervisor',              'NVARCHAR(255) NULL'),
        # Session management columns (Issues 3 & 4)
        ('session_token',           'NVARCHAR(36) NULL'),
        ('last_login',              'DATETIME2(7) NULL'),
        ('login_ip',                'NVARCHAR(45) NULL'),
        ('must_change_password',    'BIT NOT NULL DEFAULT 0'),
        ('added_by',                'NVARCHAR(80) NULL'),
        ('changed_by',              'NVARCHAR(80) NULL'),
        ('deleted_by',              'NVARCHAR(80) NULL'),
    ],
    'firm_info': [
        ('address_line1', 'NVARCHAR(255) NULL'),
        ('address_line2', 'NVARCHAR(255) NULL'),
        ('city',          'NVARCHAR(100) NULL'),
        ('province',      'NVARCHAR(100) NULL'),
        ('postal_code',   'NVARCHAR(20) NULL'),
        ('phone',         'NVARCHAR(50) NULL'),
        ('email',         'NVARCHAR(255) NULL'),
        ('tax_number',    'NVARCHAR(100) NULL'),
        ('logo_filename', 'NVARCHAR(255) NULL'),
        ('updated_at',    'DATETIME2(7) NULL'),
        ('tax1_name',     "NVARCHAR(100) NOT NULL DEFAULT 'GST'"),
        ('tax2_name',     'NVARCHAR(100) NULL'),
        ('tax2_compound', 'BIT NOT NULL DEFAULT 0'),
        ('tax1_rate',     'DECIMAL(6,3) NOT NULL DEFAULT 0'),
        ('tax2_rate',     'DECIMAL(6,3) NOT NULL DEFAULT 0'),
        ('mfa_enabled',   'BIT NOT NULL DEFAULT 0'),
    ],
    'invoices': [
        ('client_id',      'INT NULL'),
        ('credit_applied', 'DECIMAL(12,2) NOT NULL DEFAULT 0'),
        ('trust_applied',  'DECIMAL(12,2) NOT NULL DEFAULT 0'),
    ],
    'second_works': [
        ('seconds_worked', 'INT NULL'),
        ('expense_id',     'INT NULL'),
        ('recorded_at',    'DATETIME2(7) NULL'),
    ],
    'import_logs': [
        ('import_id', 'NVARCHAR(32) NULL'),
        ('file_hash', 'NVARCHAR(64) NULL'),
    ],
    'hr_records': [
        ('is_deleted', 'BIT NOT NULL DEFAULT 0'),
        ('changed_by', 'NVARCHAR(80) NULL'),
    ],
    'credit_notes': [
        ('added_by', 'NVARCHAR(80) NULL'),
    ],
    'TransactionsFiducie': [
        ('beneficiaire',    'NVARCHAR(255) NULL'),
        ('motif',           'NVARCHAR(MAX) NULL'),
        ('created_by',      'NVARCHAR(255) NULL'),
        ('invoice_number',  'NVARCHAR(100) NULL'),
    ],
    'suppliers': [
        ('account_number',    'NVARCHAR(100) NULL'),
        ('email',             'NVARCHAR(255) NULL'),
        ('notes',             'NVARCHAR(MAX) NULL'),
        ('is_active',         'BIT NOT NULL DEFAULT 1'),
        ('is_deleted',        'BIT NOT NULL DEFAULT 0'),
        ('accounting_code',   'NVARCHAR(20) NULL'),
        ('added_by',          'NVARCHAR(80) NULL'),
    ],
    'supplier_payments': [
        ('description',       'NVARCHAR(500) NULL'),
        ('payment_method',    'NVARCHAR(50) NULL'),
        ('cheque_number',     'NVARCHAR(100) NULL'),
        ('bank_transaction',  'NVARCHAR(255) NULL'),
        ('created_by',        'NVARCHAR(80) NULL'),
        ('is_deleted',        'BIT NOT NULL DEFAULT 0'),
        ('deleted_by',        'NVARCHAR(80) NULL'),
        ('is_paid',           'BIT NOT NULL DEFAULT 0'),
        ('date_paid',         'DATE NULL'),
        ('paid_by',           'NVARCHAR(80) NULL'),
        ('is_posted',         'BIT NOT NULL DEFAULT 0'),
        ('posted_by',         'NVARCHAR(80) NULL'),
    ],
    'GL_accounts': [
        ('is_deleted', 'BIT NOT NULL DEFAULT 0'),
        ('changed_by', 'NVARCHAR(80) NULL'),
    ],
    'calendar_events': [
        ('event_type',       'NVARCHAR(50) NULL'),
        ('event_time',       'NVARCHAR(10) NULL'),
        ('location',         'NVARCHAR(255) NULL'),
        ('notes',            'NVARCHAR(MAX) NULL'),
        ('is_done',          'BIT NOT NULL DEFAULT 0'),
        ('assigned_to',      'NVARCHAR(80) NULL'),
        ('created_by',       'NVARCHAR(80) NULL'),
        ('is_deleted',       'BIT NOT NULL DEFAULT 0'),
        ('updated_at',       'DATETIME2(7) NULL'),
        ('room_id',          'INT NULL'),
        ('duration_minutes', 'INT NULL'),
    ],
    'salary_entries': [
        ('is_posted',  'BIT NOT NULL DEFAULT 0'),
        ('posted_by',  'NVARCHAR(80) NULL'),
    ],
    'trust_authorizations': [
        ('client_id',         'INT NULL'),
        ('date_from',         'DATE NULL'),
        ('date_to',           'DATE NULL'),
        ('is_indefinite',     'BIT NOT NULL DEFAULT 0'),
        ('max_amount',        'DECIMAL(12,2) NULL'),
        ('notes',             'NTEXT NULL'),
        ('doc_filename',      'NVARCHAR(500) NULL'),
        ('doc_original_name', 'NVARCHAR(255) NULL'),
        ('is_active',         'BIT NOT NULL DEFAULT 1'),
        ('is_deleted',        'BIT NOT NULL DEFAULT 0'),
        ('deleted_at',        'DATETIME NULL'),
        ('deleted_by',        'NVARCHAR(255) NULL'),
        ('created_by',        'NVARCHAR(255) NULL'),
    ],
}

_schema_migrations_applied = False


def _apply_schema_migrations():
    """Create missing tables and add missing columns to existing tables.

    Called once at startup.  Uses ``db.create_all()`` to create any tables that
    do not exist yet, then runs ``ALTER TABLE … ADD …`` for each column that is
    present in the SQLAlchemy model but absent from the live database.  Errors
    are logged as warnings so that a temporary DB-connectivity issue does not
    prevent the app from starting.
    """
    global _schema_migrations_applied
    if _schema_migrations_applied:
        return
    try:
        # Create tables that don't exist yet (leaves existing tables untouched)
        db.create_all()

        inspector = sa_inspect(db.engine)
        existing_tables = set(inspector.get_table_names())

        # Rename legacy 'accounts' table to 'GL_accounts' if needed
        if 'accounts' in existing_tables and 'GL_accounts' not in existing_tables:
            try:
                with db.engine.begin() as conn:
                    conn.execute(sa_text("EXEC sp_rename 'dbo.accounts', 'GL_accounts'"))
                existing_tables.discard('accounts')
                existing_tables.add('GL_accounts')
                logger.info("Schema migration: renamed table accounts -> GL_accounts")
            except Exception as ren_exc:
                logger.warning("Could not rename accounts to GL_accounts: %s", ren_exc)

        for table_name, columns in _COLUMN_MIGRATIONS.items():
            if table_name not in existing_tables:
                continue
            existing_cols = {col['name'].lower() for col in inspector.get_columns(table_name)}
            with db.engine.begin() as conn:
                for col_name, col_def in columns:
                    if col_name.lower() not in existing_cols:
                        # Use quoted identifiers ([…]) to prevent any future
                        # misuse if the dict keys ever come from a dynamic source.
                        conn.execute(sa_text(
                            f"ALTER TABLE [{table_name}] ADD [{col_name}] {col_def}"
                        ))
                        logger.info("Schema migration: added %s.%s", table_name, col_name)

        # Ensure existing employees can still log in after permission columns were
        # added with DEFAULT 0.  Backfill is_user=1 for all active, non-deleted
        # employees where it is still 0 – whether is_user was just added by this
        # migration run or was pre-existing (e.g. added manually with DEFAULT 0,
        # or reset to 0 by a GUI tool like SSMS when recreating the table).
        # Idempotent: rows already at is_user=1 are not touched.
        if 'employees' in existing_tables:
            with db.engine.begin() as conn:
                conn.execute(sa_text(
                    "UPDATE [employees] SET [is_user] = 1 "
                    "WHERE [is_active] = 1 AND [is_deleted] = 0 AND [is_user] = 0"
                ))
                logger.info("Schema migration: backfilled is_user=1 for active employees")

        # Backfill is_paid and date_paid for supplier_payments that already have
        # a payment_date – those were paid before the is_paid column was added.
        if 'supplier_payments' in existing_tables:
            with db.engine.begin() as conn:
                conn.execute(sa_text(
                    "UPDATE [supplier_payments] "
                    "SET [is_paid] = 1, [date_paid] = [payment_date] "
                    "WHERE [payment_date] IS NOT NULL AND [is_paid] = 0"
                ))
                logger.info("Schema migration: backfilled is_paid/date_paid for paid supplier_payments")

        # Backfill deleted_at for trust_authorizations that were soft-deleted via
        # is_active = 0 before the deleted_at column existed.  Use updated_at as the
        # approximate deletion timestamp (falls back to created_at, then GETDATE()).
        # Also backfill is_deleted=1 for records that already have deleted_at set.
        if 'trust_authorizations' in existing_tables:
            try:
                with db.engine.begin() as conn:
                    conn.execute(sa_text(
                        "UPDATE [trust_authorizations] "
                        "SET [deleted_at] = COALESCE([updated_at], [created_at], GETDATE()) "
                        "WHERE [is_active] = 0 AND [deleted_at] IS NULL"
                    ))
                    conn.execute(sa_text(
                        "UPDATE [trust_authorizations] "
                        "SET [is_deleted] = 1 "
                        "WHERE [deleted_at] IS NOT NULL AND [is_deleted] = 0"
                    ))
                    logger.info("Schema migration: backfilled deleted_at and is_deleted for soft-deleted trust_authorizations")
            except Exception as ta_exc:
                logger.warning("Could not backfill trust_authorizations.deleted_at/is_deleted: %s", ta_exc)

        # Add CHECK constraint for timer_user / hourly_rate requirement
        if 'employees' in existing_tables:
            try:
                with db.engine.begin() as conn:
                    row = conn.execute(sa_text(
                        "SELECT COUNT(*) FROM sys.check_constraints "
                        "WHERE parent_object_id = OBJECT_ID('dbo.employees') "
                        "AND name = 'CK_employees_timer_hourly_rate'"
                    )).scalar()
                    if not row:
                        conn.execute(sa_text(
                            "ALTER TABLE [employees] ADD CONSTRAINT [CK_employees_timer_hourly_rate] "
                            "CHECK (timer_user = 0 OR (hourly_rate IS NOT NULL AND hourly_rate > 0))"
                        ))
                        logger.info("Schema migration: added CK_employees_timer_hourly_rate")
            except Exception as ck_exc:
                logger.warning("Could not add timer_hourly_rate check constraint: %s", ck_exc)

        # Make invoices.matter_id nullable to support "All Matters" client-level invoices
        if 'invoices' in existing_tables:
            try:
                with db.engine.begin() as conn:
                    row = conn.execute(sa_text(
                        "SELECT IS_NULLABLE FROM INFORMATION_SCHEMA.COLUMNS "
                        "WHERE TABLE_NAME='invoices' AND COLUMN_NAME='matter_id'"
                    )).fetchone()
                    if row and row[0] == 'NO':
                        conn.execute(sa_text(
                            "ALTER TABLE [invoices] ALTER COLUMN [matter_id] INT NULL"
                        ))
                        logger.info("Schema migration: made invoices.matter_id nullable")
            except Exception as mi_exc:
                logger.warning("Could not make invoices.matter_id nullable: %s", mi_exc)

        # Seed the default chart of accounts if the table is empty
        _seed_default_accounts()

        # Only mark as done when every migration step succeeds so that a
        # partial failure (e.g. temporary DB outage) retries on next request.
        _schema_migrations_applied = True
    except Exception as exc:
        logger.warning("Schema migration check failed (will retry on next request): %s", exc)


# Default plan de comptes for a Quebec law firm.
# Each entry is: (code, name, account_type, parent_code, is_system)
_DEFAULT_ACCOUNTS = [
    # ── ACTIF (1xxx) ──────────────────────────────────────────────
    ('1000', 'ACTIF',                       'asset',     None,   True),
    ('1010', 'Banque – compte opérationnel','asset',     '1000', True),
    ('1020', 'Banque – compte en fiducie',  'asset',     '1000', True),
    ('1100', 'Comptes à recevoir',          'asset',     '1000', True),
    ('1200', 'Avances aux employés',        'asset',     '1000', False),
    # ── PASSIF (2xxx) ─────────────────────────────────────────────
    ('2000', 'PASSIF',                      'liability', None,   True),
    ('2010', 'Comptes à payer',             'liability', '2000', True),
    ('2100', 'Salaires à payer',            'liability', '2000', False),
    ('2110', 'DAS à payer',                 'liability', '2000', False),
    ('2200', 'TPS/TVQ à remettre',          'liability', '2000', True),
    ('2300', 'Fiducie clients (passif)',     'liability', '2000', True),
    # ── CAPITAUX PROPRES (3xxx) ───────────────────────────────────
    ('3000', 'CAPITAUX PROPRES',            'equity',    None,   False),
    ('3010', 'Capital',                     'equity',    '3000', False),
    ('3020', 'Bénéfices non répartis',      'equity',    '3000', False),
    # ── REVENUS (4xxx) ────────────────────────────────────────────
    ('4000', 'REVENUS',                     'revenue',   None,   True),
    ('4010', 'Honoraires juridiques',       'revenue',   '4000', True),
    ('4020', 'Débours refacturés',          'revenue',   '4000', True),
    # ── DÉPENSES (5xxx) ───────────────────────────────────────────
    ('5000', 'DÉPENSES',                    'expense',   None,   False),
    ('5010', 'Salaires',                    'expense',   '5000', False),
    ('5020', 'Charges sociales',            'expense',   '5000', False),
    ('5030', 'Loyer',                       'expense',   '5000', False),
    ('5040', 'Fournitures de bureau',       'expense',   '5000', False),
    ('5050', 'Débours avancés (frais)',     'expense',   '5000', False),
]


def _seed_default_accounts():
    """Insert the default plan de comptes if the accounts table is empty."""
    try:
        if Account.query.count() > 0:
            return
        # First pass: create accounts without parent links
        code_to_obj = {}
        for code, name, acct_type, parent_code, is_system in _DEFAULT_ACCOUNTS:
            acct = Account(code=code, name=name, account_type=acct_type,
                           is_system=is_system, is_active=True)
            db.session.add(acct)
            code_to_obj[code] = acct
        db.session.flush()
        # Second pass: set parent_id
        for code, name, acct_type, parent_code, is_system in _DEFAULT_ACCOUNTS:
            if parent_code and parent_code in code_to_obj:
                code_to_obj[code].parent_id = code_to_obj[parent_code].id
        db.session.commit()
        logger.info("Seeded %d default accounts", len(_DEFAULT_ACCOUNTS))
    except Exception as exc:
        db.session.rollback()
        logger.warning("Could not seed default accounts: %s", exc)


# ── Journal entry helpers ─────────────────────────────────────────────────────

def _get_account_by_code(code):
    """Return the Account with the given code, or None if not found."""
    return Account.query.filter_by(code=code).first()


def _create_journal_entry(entry_date, description, source_type, source_id,
                           lines, created_by=None):
    """Create a balanced journal entry with the given lines.

    ``lines`` is a list of dicts with keys:
        account_code  – e.g. '1100'
        debit         – float (0 if credit side)
        credit        – float (0 if debit side)
        client_id     – optional int
        matter_id     – optional int
        is_trust      – optional bool
        memo          – optional str

    Returns the saved JournalEntry, or None on failure.
    """
    try:
        entry = JournalEntry(
            entry_date=entry_date,
            description=description,
            source_type=source_type,
            source_id=source_id,
            created_by=created_by,
        )
        db.session.add(entry)
        db.session.flush()

        for ln in lines:
            account = _get_account_by_code(ln['account_code'])
            if account is None:
                logger.warning("Journal entry: unknown account code '%s' — skipping line",
                               ln['account_code'])
                continue
            line = JournalLine(
                entry_id=entry.id,
                account_id=account.id,
                debit=round(float(ln.get('debit', 0)), 2),
                credit=round(float(ln.get('credit', 0)), 2),
                client_id=ln.get('client_id'),
                matter_id=ln.get('matter_id'),
                is_trust=bool(ln.get('is_trust', False)),
                memo=ln.get('memo'),
            )
            db.session.add(line)

        return entry
    except Exception as exc:
        logger.warning("Could not create journal entry: %s", exc)
        return None


def _post_invoice_journal(invoice, client):
    """Auto-post a double-entry journal record when an invoice is finalised.

    Debit  1100 Comptes à recevoir
    Credit 4010 Honoraires juridiques  (subtotal)
    Credit 2200 TPS/TVQ à remettre     (taxes, if any)
    """
    try:
        amount = float(invoice.total_amount or 0)
        if amount <= 0:
            return
        matter = Matter.query.get(invoice.matter_id) if invoice.matter_id else None
        lines = [
            {
                'account_code': '1100',
                'debit': amount,
                'credit': 0,
                'client_id': client.id if client else None,
                'matter_id': invoice.matter_id,
            },
        ]
        subtotal = float(invoice.subtotal or 0)
        gst = float(invoice.gst_amount or 0)
        qst = float(invoice.qst_amount or 0)
        if subtotal > 0:
            lines.append({
                'account_code': '4010',
                'debit': 0,
                'credit': round(subtotal, 2),
                'client_id': client.id if client else None,
                'matter_id': invoice.matter_id,
            })
        tax_total = round(gst + qst, 2)
        if tax_total > 0:
            lines.append({
                'account_code': '2200',
                'debit': 0,
                'credit': tax_total,
                'client_id': client.id if client else None,
                'matter_id': invoice.matter_id,
            })
        actor = current_user.display_name if current_user and current_user.is_authenticated else 'system'
        _create_journal_entry(
            entry_date=invoice.invoice_date,
            description=f'Facture {invoice.invoice_number}',
            source_type='invoice',
            source_id=invoice.id,
            lines=lines,
            created_by=actor,
        )
    except Exception as exc:
        logger.warning("Could not auto-post invoice journal: %s", exc)


def _post_payment_journal(invoice, client, payment_amount):
    """Auto-post the payment journal entry when an invoice is marked paid.

    Debit  1010 Banque – compte opérationnel
    Credit 1100 Comptes à recevoir
    """
    try:
        if payment_amount <= 0:
            return
        actor = current_user.display_name if current_user and current_user.is_authenticated else 'system'
        _create_journal_entry(
            entry_date=invoice.invoice_date,
            description=f'Paiement facture {invoice.invoice_number}',
            source_type='payment',
            source_id=invoice.id,
            lines=[
                {
                    'account_code': '1010',
                    'debit': round(payment_amount, 2),
                    'credit': 0,
                    'client_id': client.id if client else None,
                    'matter_id': invoice.matter_id,
                },
                {
                    'account_code': '1100',
                    'debit': 0,
                    'credit': round(payment_amount, 2),
                    'client_id': client.id if client else None,
                    'matter_id': invoice.matter_id,
                },
            ],
            created_by=actor,
        )
    except Exception as exc:
        logger.warning("Could not auto-post payment journal: %s", exc)


def _post_trust_journal(transaction, matter):
    """Auto-post a trust transaction to the double-entry journal.

    DEPOT:
        Debit  1020 Banque fiducie
        Credit 2300 Fiducie clients (passif)

    RETRAIT / payment from trust:
        Debit  2300 Fiducie clients (passif)
        Credit 1020 Banque fiducie
    """
    try:
        amount = float(transaction.montant or 0)
        if amount <= 0:
            return
        client = matter.client if matter else None
        actor = transaction.created_by or 'system'
        if transaction.type_trans == 'DEPOT':
            lines = [
                {
                    'account_code': '1020',
                    'debit': amount,
                    'credit': 0,
                    'client_id': client.id if client else None,
                    'matter_id': matter.id if matter else None,
                    'is_trust': True,
                },
                {
                    'account_code': '2300',
                    'debit': 0,
                    'credit': amount,
                    'client_id': client.id if client else None,
                    'matter_id': matter.id if matter else None,
                    'is_trust': True,
                },
            ]
            description = f'Dépôt fiducie – {matter.matter_number if matter else ""}'
            source_type = 'trust_deposit'
        else:
            lines = [
                {
                    'account_code': '2300',
                    'debit': amount,
                    'credit': 0,
                    'client_id': client.id if client else None,
                    'matter_id': matter.id if matter else None,
                    'is_trust': True,
                },
                {
                    'account_code': '1020',
                    'debit': 0,
                    'credit': amount,
                    'client_id': client.id if client else None,
                    'matter_id': matter.id if matter else None,
                    'is_trust': True,
                },
            ]
            description = (f'Retrait fiducie – {matter.matter_number if matter else ""}'
                           + (f' (Facture {transaction.invoice_number})' if transaction.invoice_number else ''))
            source_type = 'trust_withdrawal'

        entry_date = transaction.date_trans.date() if isinstance(transaction.date_trans, datetime) else transaction.date_trans
        _create_journal_entry(
            entry_date=entry_date,
            description=description,
            source_type=source_type,
            source_id=transaction.fid_id,
            lines=lines,
            created_by=actor,
        )
    except Exception as exc:
        logger.warning("Could not auto-post trust journal: %s", exc)


def _post_supplier_payment_journal(payment, supplier):
    """Auto-post a supplier payment to the double-entry journal.

    Debit  <supplier accounting_code or 2010 Comptes à payer>
    Credit 1010 Banque – compte opérationnel
    """
    try:
        amount = float(payment.amount or 0)
        if amount <= 0:
            return
        actor = current_user.display_name if current_user and current_user.is_authenticated else 'system'
        expense_code = (supplier.accounting_code or '2010') if supplier else '2010'
        entry_date = payment.payment_date or payment.invoice_date or datetime.utcnow().date()
        if isinstance(entry_date, datetime):
            entry_date = entry_date.date()
        _create_journal_entry(
            entry_date=entry_date,
            description=f'Paiement fournisseur {supplier.name if supplier else ""}'
                        + (f' – {payment.invoice_number}' if payment.invoice_number else ''),
            source_type='supplier_payment',
            source_id=payment.id,
            lines=[
                {
                    'account_code': expense_code,
                    'debit': round(amount, 2),
                    'credit': 0,
                },
                {
                    'account_code': '1010',
                    'debit': 0,
                    'credit': round(amount, 2),
                },
            ],
            created_by=actor,
        )
    except Exception as exc:
        logger.warning("Could not auto-post supplier payment journal: %s", exc)


@app.before_request
def _ensure_schema_migrated():
    """Run schema migrations once before the first request is handled.

    Registered as the FIRST before_request hook so the database schema is
    always up-to-date before any other hook or route handler queries the DB.
    This prevents column-not-found errors when Flask-Login's user_loader is
    triggered by subsequent hooks (e.g. license check) before migrations run.
    """
    if not _schema_migrations_applied:
        _apply_schema_migrations()


@app.before_request
def _enforce_license_restrictions():
    """Role-based license enforcement on every request.

    Access matrix (all non-VALID statuses treated identically):

    A) Manager + VALID        → all modules accessible
    B) Manager + INVALID      → Home, Clients & Matters, Timer allowed;
                                 Cost Codes, Invoices, Employees, Import blocked
    C) Timer-only user        → handled by _enforce_timer_user_restriction();
                                 this function passes through unchanged
    D) User (non-manager) + VALID   → Clients & Matters, Invoices accessible;
                                       Cost Codes/Employees/Import blocked by
                                       existing role checks in those views
    E) User (non-manager) + INVALID → only Home + Timer allowed; everything
                                       else blocked

    Blocked HTML routes redirect to index with a flash warning.
    Blocked API routes return JSON 403 with consistent payload.
    """
    path = request.path

    # ── Always allow static files ─────────────────────────────────────────
    if path.startswith('/static/'):
        return

    # ── Always allow auth + public setup routes ───────────────────────────
    _always_allowed = {
        '/login', '/logout', '/register',
        '/forgot-password', '/reset-password',
        '/license/acknowledge', '/mfa-verify',
    }
    if path in _always_allowed or path.startswith('/reset-password/') or path.startswith('/set_lang/'):
        return

    # ── Always allow public APIs (used from login page setup dialog) ──────
    _public_apis = {
        '/api/firm-info', '/api/firm-info/logo',
        '/api/manager', '/api/license-info',
    }
    if path in _public_apis:
        return

    license_path, public_key_b64 = _licensing.get_license_config(app.config)
    result = _licensing.get_cached_license_result(license_path, public_key_b64)

    # Valid license → role checks in individual views handle the rest
    if result.is_valid:
        return

    # Invalid license – unauthenticated requests pass through so Flask-Login
    # can redirect them to the login page normally.
    if not current_user.is_authenticated:
        return

    is_manager  = getattr(current_user, 'is_manager', False)
    is_user     = getattr(current_user, 'is_user', False)
    timer_user  = getattr(current_user, 'timer_user', False)

    # Timer-only users: let _enforce_timer_user_restriction() handle them
    if timer_user and not is_user and not is_manager:
        return

    def _api_403():
        return jsonify({
            'error':   'license_invalid',
            'status':  result.status.value,
            'message': 'This module requires a valid license.',
        }), 403

    if is_manager:
        # B) Manager + invalid license: block specific modules
        _blocked = (
            path.startswith('/cost-codes') or
            path.startswith('/api/cost-codes') or
            path.startswith('/invoices') or
            path.startswith('/api/invoices') or
            path.startswith('/employees') or
            path.startswith('/api/employees') or
            path.startswith('/hr-records') or
            path.startswith('/api/hr-records') or
            path.startswith('/import') or
            path.startswith('/api/import')
        )
        if _blocked:
            if path.startswith('/api/'):
                return _api_403()
            flash('This module requires a valid license.', 'warning')
            return redirect(url_for('index'))
        return  # Allow home, clients, timer, export, etc.

    if is_user:
        # E) Non-manager user + invalid license: allow only Home + Timer routes
        _allowed = (
            path == '/' or
            path.startswith('/timer') or
            path.startswith('/api/timer')
        )
        if not _allowed:
            if path.startswith('/api/'):
                return _api_403()
            flash('A valid license is required to access this module.', 'warning')
            return redirect(url_for('index'))
        return


# ── Session timeout & single-session enforcement ─────────────────────────────

_SESSION_TIMEOUT_SECONDS = 15 * 60  # 15 minutes of inactivity

_SESSION_EXEMPT_PATHS = {'/login', '/logout', '/register', '/forgot-password',
                         '/reset-password', '/license/acknowledge', '/mfa-verify'}


@app.before_request
def _check_session_timeout():
    """Force-logout a user who has been inactive for more than 15 minutes."""
    if not current_user.is_authenticated:
        return
    path = request.path
    if path.startswith('/static/') or path in _SESSION_EXEMPT_PATHS or path.startswith('/reset-password/'):
        return
    last_activity_raw = session.get('last_activity')
    if last_activity_raw:
        try:
            last_activity = datetime.fromisoformat(last_activity_raw)
            #idle_seconds = (datetime.utcnow() - last_activity).total_seconds()
            idle_seconds = (datetime.now(timezone.utc) - last_activity).total_seconds()
            if idle_seconds > _SESSION_TIMEOUT_SECONDS:       
          
                logout_user()
                session.clear()
                if path.startswith('/api/'):
                    return jsonify({'error': 'session_expired',
                                    'message': 'Session expirée après 15 minutes d\'inactivité.'}), 401
                flash('Votre session a expiré après 15 minutes d\'inactivité. Veuillez vous reconnecter.', 'warning')
                return redirect(url_for('login'))
        except (ValueError, TypeError):
            pass
    # Refresh last-activity timestamp on every authenticated request
    #session['last_activity'] = datetime.utcnow().isoformat()
    session['last_activity'] = datetime.now(UTC).isoformat()
    session.modified = True


@app.before_request
def _enforce_single_session():
    
    """If the user logged in from another device, invalidate the current session.

    On login a unique token is stored both in the server-side DB row and in
    the session cookie.  If the tokens no longer match (because a new login
    replaced the DB token), this device's session is rejected.

    The check is skipped when either token is absent (e.g. the session_token
    column does not exist yet, or the DB commit failed during login) so that
    transient infrastructure issues never lock users out permanently.
    """
    if not current_user.is_authenticated:
        return
    path = request.path
    if path.startswith('/static/') or path in _SESSION_EXEMPT_PATHS or path.startswith('/reset-password/'):
        return
    try:
        cookie_token = session.get('login_token')
        db_token = getattr(current_user, 'session_token', None)
    except Exception:
        # If we cannot read the tokens for any reason, skip enforcement rather
        # than accidentally logging the user out.
        return
    if cookie_token and db_token and cookie_token != db_token:
        logout_user()
        session.clear()
        if path.startswith('/api/'):
            return jsonify({'error': 'session_invalidated',
                            'message': 'Vous avez été déconnecté car un autre appareil s\'est connecté avec ce compte.'}), 401
        flash('Vous avez été déconnecté car un autre appareil s\'est connecté avec votre compte.', 'warning')
        return redirect(url_for('login'))


@app.before_request
def _enforce_timer_user_restriction():
    """Block timer-only users from accessing non-timer pages/APIs."""
    if not current_user.is_authenticated:
        return
    if not getattr(current_user, 'timer_user', False):
        return
    # Users who also have regular or manager access are unrestricted
    if current_user.is_user or current_user.is_manager:
        return
    path = request.path
    allowed = (
        path.startswith('/timer') or
        path.startswith('/time-logs') or
        path.startswith('/api/timer/') or
        path.startswith('/api/time-logs') or
        path.startswith('/static/') or
        path.startswith('/set_lang/') or
        path in ('/logout', '/reset-password', '/forgot-password') or
        path.startswith('/reset-password/')
    )
    if not allowed:
        if path.startswith('/api/'):
            return jsonify({'error': 'Access restricted to timer functionality'}), 403
        return redirect(url_for('timer_page'))


# ── License context processor & acknowledgement ───────────────────────────────

@app.context_processor
def inject_translations():
    # 1. On récupère la langue (fr par défaut si vide)
    lang = session.get('lang', 'fr') 
    
    # 2. On s'assure que LEXICON contient bien cette langue
    texts = LEXICON.get(lang, LEXICON.get('fr'))

    # 3. On rend 't' et 'current_lang' disponibles pour TOUS les fichiers .html
    return dict(t=texts, current_lang=lang, app_version=APP_VERSION)


@app.context_processor
def _inject_license_status():
    """Expose license status to every template.

    ``license_modules_restricted`` is True for *any* non-VALID status so that
    templates can grey-out restricted modules regardless of the specific reason
    the license is invalid.
    """
    license_path, public_key_b64 = _licensing.get_license_config(app.config)
    result = _licensing.get_cached_license_result(license_path, public_key_b64)
    return {
        'license_valid':              result.is_valid,
        'license_status':             result.status.value,
        'license_modules_restricted': not result.is_valid,
    }


@app.route('/license/acknowledge', methods=['POST'])
def license_acknowledge():
    """Legacy acknowledge endpoint – kept for backward compatibility.

    Acknowledgement mode has been removed; this route now redirects to login
    without setting any session flag.
    """
    return redirect(url_for('login'))


# ── Authentication routes ─────────────────────────────────────────────────────
@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    if request.method == 'POST':
        uname = request.form.get('username', '').strip()
        pwd = request.form.get('password', '')
        user = Employee.query.filter_by(username=uname).first()
        if user and user.is_active and not getattr(user, 'is_deleted', False) and user.check_password(pwd):
            if user.is_user or user.is_manager or user.timer_user:
                # Check if MFA is enabled at the firm level
                firm = FirmInfo.query.first()
                if firm and getattr(firm, 'mfa_enabled', False):
                    lang = session.get('lang', 'fr')
                    # MFA is enabled — send a 6-digit code and redirect to verification
                    import random
                    mfa_code = '{:06d}'.format(random.randint(0, 999999))
                    expires_at = (datetime.now(timezone.utc) + timedelta(minutes=4)).isoformat()
                    next_page = request.args.get('next', '')
                    # Store pending MFA state in the session (not yet logged in)
                    session['mfa_pending_user_id'] = user.id
                    session['mfa_code'] = mfa_code
                    session['mfa_expires_at'] = expires_at
                    session['mfa_next_url'] = next_page
                    try:
                        send_mfa_email(user, mfa_code, lang=lang)
                    except Exception as mail_exc:
                        logger.error("Failed to send MFA email to '%s': %s", uname, mail_exc)
                        texts = LEXICON.get(lang, LEXICON.get('fr'))
                        flash(texts.get('msg_mfa_email_failed', 'Unable to send verification code. Please contact your administrator.'), 'danger')
                        session.pop('mfa_pending_user_id', None)
                        session.pop('mfa_code', None)
                        session.pop('mfa_expires_at', None)
                        session.pop('mfa_next_url', None)
                        return render_template('login.html')
                    return redirect(url_for('mfa_verify'))
                # No MFA — proceed with normal login
                # Generate a new session token (invalidates any previous session)
                new_token = str(uuid.uuid4())
                user.session_token = new_token
                user.last_login = datetime.now()
                user.login_ip = request.headers.get("X-Real-IP", request.remote_addr)
                try:
                    db.session.commit()
                except Exception as commit_exc:
                    # If we cannot persist the session token (e.g. transient DB
                    # error), roll back and continue the login without token
                    # enforcement so the user is not silently locked out.
                    logger.warning(
                        "Login: failed to persist session token for '%s': %s",
                        uname, commit_exc,
                    )
                    db.session.rollback()
                    new_token = None
                login_user(user, remember=False)
                session.permanent = False  # Session cookie expires when browser closes
                # Discard any stale login_token from a previous session before
                # writing the new one so _enforce_single_session never sees a
                # mismatched pair if the cookie is recycled from a previous login.
                session.pop('login_token', None)
                if new_token:
                    session['login_token'] = new_token
                session['last_activity'] = datetime.now(timezone.utc).isoformat()
                # If the admin set must_change_password, redirect to the change-password page
                if getattr(user, 'must_change_password', False):
                    return redirect(url_for('reset_password'))
                next_page = request.args.get('next')
                # Timer-only users go to the timer page
                if user.timer_user and not user.is_user and not user.is_manager:
                    return redirect(next_page or url_for('timer_page'))
                return redirect(next_page or url_for('index'))
            flash("You don't have permission to connect to this application.", 'danger')
            return render_template('login.html')
        # Check if the account exists but is deactivated/deleted
        if user and (not user.is_active or getattr(user, 'is_deleted', False)):
            lang = session.get('lang', 'fr')
            texts = LEXICON.get(lang, LEXICON.get('fr'))
            flash(texts.get('msg_account_deactivated', 'Your account has been deactivated. Please contact your supervisor.'), 'warning')
            return render_template('login.html')
        flash('Invalid username or password.', 'danger')
    return render_template('login.html')


@app.route('/mfa-verify', methods=['GET', 'POST'])
def mfa_verify():
    """MFA verification page — shown after a successful password check when MFA is enabled."""
    # If already authenticated, go home
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    # There must be a pending MFA session
    pending_id = session.get('mfa_pending_user_id')
    if not pending_id:
        return redirect(url_for('login'))
    lang = session.get('lang', 'fr')
    t = LEXICON.get(lang, LEXICON.get('fr'))
    if request.method == 'POST':
        entered_code = request.form.get('mfa_code', '').strip()
        stored_code = session.get('mfa_code', '')
        expires_at_str = session.get('mfa_expires_at', '')
        # Check expiry
        try:
            expires_at = datetime.fromisoformat(expires_at_str)
            if expires_at.tzinfo is None:
                expires_at = expires_at.replace(tzinfo=timezone.utc)
        except (ValueError, TypeError):
            expires_at = datetime.now(timezone.utc) - timedelta(seconds=1)
        if datetime.now(timezone.utc) > expires_at:
            session.pop('mfa_pending_user_id', None)
            session.pop('mfa_code', None)
            session.pop('mfa_expires_at', None)
            session.pop('mfa_next_url', None)
            flash(t.get('msg_mfa_expired', 'The verification code has expired. Please log in again.'), 'warning')
            return redirect(url_for('login'))
        if entered_code != stored_code:
            flash(t.get('msg_mfa_invalid', 'Invalid verification code. Please try again.'), 'danger')
            return render_template('mfa_verify.html', t=t, lang=lang)
        # Code is valid — complete the login
        user = db.session.get(Employee, pending_id)
        if not user or not user.is_active or getattr(user, 'is_deleted', False):
            session.pop('mfa_pending_user_id', None)
            session.pop('mfa_code', None)
            session.pop('mfa_expires_at', None)
            session.pop('mfa_next_url', None)
            flash(t.get('msg_account_deactivated', 'Your account has been deactivated. Please contact your supervisor.'), 'warning')
            return redirect(url_for('login'))
        next_page = session.pop('mfa_next_url', '') or ''
        # Clear MFA session keys
        session.pop('mfa_pending_user_id', None)
        session.pop('mfa_code', None)
        session.pop('mfa_expires_at', None)
        # Complete the login (same as normal login flow)
        new_token = str(uuid.uuid4())
        user.session_token = new_token
        user.last_login = datetime.now()
        user.login_ip = request.headers.get("X-Real-IP", request.remote_addr)
        try:
            db.session.commit()
        except Exception as commit_exc:
            logger.warning(
                "MFA login: failed to persist session token for user id %s: %s",
                pending_id, commit_exc,
            )
            db.session.rollback()
            new_token = None
        login_user(user, remember=False)
        session.permanent = False
        session.pop('login_token', None)
        if new_token:
            session['login_token'] = new_token
        session['last_activity'] = datetime.now(timezone.utc).isoformat()
        if getattr(user, 'must_change_password', False):
            return redirect(url_for('reset_password'))
        if user.timer_user and not user.is_user and not user.is_manager:
            return redirect(next_page or url_for('timer_page'))
        return redirect(next_page or url_for('index'))
    return render_template('mfa_verify.html', t=t, lang=lang)

   

@app.route('/register', methods=['GET', 'POST'])
def register():
    # Registration is only for the initial manager setup.
    # Once the first user exists, new employees are created by a manager
    # through the employee management module.
    if Employee.query.count() > 0:
        return redirect(url_for('login'))
    if request.method == 'POST':
        uname = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip()
        pwd = request.form.get('password', '')
        confirm = request.form.get('confirm_password', '')
        if not uname or not email or not pwd:
            flash('All fields are required.', 'danger')
        elif pwd != confirm:
            flash('Passwords do not match.', 'danger')
        elif Employee.query.filter_by(username=uname).first():
            flash('Username already exists.', 'danger')
        elif Employee.query.filter_by(email=email).first():
            flash('Email already registered.', 'danger')
        else:
            # First user is always the manager
            employee = Employee(username=uname, email=email,
                                is_manager=True, is_user=True, is_active=True)
            employee.set_password(pwd)
            db.session.add(employee)
            db.session.flush()  # obtain employee.id before commit
            # Create the matching HR record
            _ensure_hr_record(employee)
            db.session.commit()
            flash(f'Gestionnaire "{uname}" créé avec succès. Veuillez vous connecter.', 'success')
            return redirect(url_for('login'))
    return render_template('register.html')

    # --- MODULE FIDUCIE ---
# 1. D'ABORD LA DÉFINITION DE LA TABLE (La classe)
class TransactionsFiducie(db.Model):
    __tablename__ = 'TransactionsFiducie'
    fid_id = db.Column(db.Integer, primary_key=True)
    matter_id = db.Column(db.Integer, db.ForeignKey('matters.id'), nullable=False)
    date_trans = db.Column(db.DateTime, default=datetime.utcnow)
    type_trans = db.Column(db.String(20))
    montant = db.Column(db.Float, nullable=False)
    beneficiaire = db.Column(db.String(255), nullable=True)
    motif = db.Column(db.Text, nullable=True)
    ref_bancaire = db.Column(db.String(100))
    invoice_number = db.Column(db.String(100), nullable=True)
    est_annulee = db.Column(db.Boolean, default=False)
    created_by = db.Column(db.String(255), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow)

@app.route('/fiducie', methods=['GET'])
@login_required
def fiducie_select():
    """Landing page: show client/matter selector for the fiducie module."""
    clients = Client.query.filter(Client.is_deleted == False, Client.is_active == True).order_by(Client.client_name).all()
    return render_template('fiducie.html',
                           dossier=None,
                           solde=0,
                           transactions=[],
                           clients=clients)

# 1. On ajoute les "methods" pour permettre d'enregistrer des données
@app.route('/fiducie/<int:id>', methods=['GET', 'POST'])
@login_required
def module_fiducie(id):
    # 1. Préparation des données de base
    dossier = Matter.query.get_or_404(id)
    

    # 2. Récupérer les transactions existantes pour calculer le solde
    transactions = TransactionsFiducie.query.filter_by(matter_id=id).order_by(TransactionsFiducie.date_trans.desc()).all()
    
    # On calcule le solde (Dépôts - Retraits)
    solde_final = sum(t.montant if t.type_trans == 'DEPOT' else -t.montant for t in transactions if not t.est_annulee)

    # 3. Gérer l'enregistrement d'une nouvelle transaction (POST)
    if request.method == 'POST':
        try:
            montant = float(request.form.get('montant', 0))
            type_trans = request.form.get('type')

            if type_trans == 'RETRAIT' and montant > solde_final:
                flash(f"🚫 Fonds insuffisants ! Solde : {solde_final:,.2f}$", "danger")
            else:
                nouvelle_t = TransactionsFiducie(
                    matter_id=id,
                    type_trans=type_trans,
                    montant=montant,
                    beneficiaire=request.form.get('beneficiaire'),
                    motif=request.form.get('motif'),
                    ref_bancaire=request.form.get('ref'),
                    est_annulee=False,
                    created_by=current_user.display_name
                )
                db.session.add(nouvelle_t)
                db.session.commit()
                flash("✅ Transaction enregistrée.", "success")
            
            return redirect(url_for('module_fiducie', id=id))
            
        except Exception as e:
            flash(f"Erreur : {str(e)}", "danger")
            return redirect(url_for('module_fiducie', id=id))

    # 4. Affichage final (GET)
    clients = Client.query.filter(Client.is_deleted == False, Client.is_active == True).order_by(Client.client_name).all()
    return render_template('fiducie.html', 
                           solde=solde_final, 
                           transactions=transactions, 
                           dossier=dossier,
                           clients=clients)


# ── API: Fiducie / Trust Accounting ──────────────────────────────────────────

@app.route('/api/fiducie/<int:matter_id>', methods=['GET'])
@login_required
def api_fiducie_list(matter_id):
    """Return all trust transactions for a given matter as JSON."""
    Matter.query.get_or_404(matter_id)
    transactions = TransactionsFiducie.query.filter_by(matter_id=matter_id).order_by(
        TransactionsFiducie.date_trans.desc()
    ).all()
    solde = sum(
        t.montant if t.type_trans == 'DEPOT' else -t.montant
        for t in transactions if not t.est_annulee
    )
    return jsonify({
        'matter_id': matter_id,
        'balance': round(solde, 2),
        'transactions': [
            {
                'id': t.fid_id,
                'date': t.date_trans.strftime('%Y-%m-%d') if t.date_trans else None,
                'type': t.type_trans,
                'montant': float(t.montant),
                'beneficiaire': t.beneficiaire,
                'motif': t.motif,
                'ref_bancaire': t.ref_bancaire,
                'invoice_number': t.invoice_number or '',
                'est_annulee': t.est_annulee,
                'created_by': t.created_by,
            }
            for t in transactions
        ]
    })


@app.route('/api/fiducie/<int:matter_id>', methods=['POST'])
@login_required
def api_fiducie_create(matter_id):
    """Create a new trust transaction via JSON API."""
    matter = Matter.query.get_or_404(matter_id)
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No JSON data provided'}), 400

    type_trans = (data.get('type_trans') or '').strip().upper()
    if type_trans not in ('DEPOT', 'RETRAIT', 'REMBOURSEMENT'):
        return jsonify({'error': 'type_trans must be DEPOT, RETRAIT or REMBOURSEMENT'}), 400

    try:
        montant = float(data.get('montant', 0))
        if montant <= 0:
            return jsonify({'error': 'montant must be positive'}), 400
    except (TypeError, ValueError):
        return jsonify({'error': 'Invalid montant value'}), 400

    # Check sufficient funds for withdrawals
    if type_trans in ('RETRAIT', 'REMBOURSEMENT'):
        transactions = TransactionsFiducie.query.filter_by(matter_id=matter_id).all()
        solde = sum(
            t.montant if t.type_trans == 'DEPOT' else -t.montant
            for t in transactions if not t.est_annulee
        )
        if montant > solde:
            return jsonify({'error': f'Insufficient funds. Balance: {solde:,.2f}$'}), 400

        # Require an active client authorization to withdraw/reimburse funds from trust
        today = datetime.now(UTC).date()
        auths = TrustAuthorization.query.filter_by(matter_id=matter_id).all()
        has_active_auth = any(a.is_active_on(today) for a in auths)
        if not has_active_auth:
            return jsonify({
                'error': 'no_authorization',
                'message': 'Aucune autorisation client active pour ce dossier. '
                           'Veuillez ajouter une autorisation signée avant de procéder à un retrait ou remboursement.'
            }), 403

    nouvelle_t = TransactionsFiducie(
        matter_id=matter_id,
        type_trans=type_trans,
        montant=montant,
        beneficiaire=(data.get('beneficiaire') or '').strip() or None,
        motif=(data.get('motif') or '').strip() or None,
        ref_bancaire=(data.get('ref_bancaire') or '').strip() or None,
        invoice_number=(data.get('invoice_number') or '').strip() or None,
        est_annulee=False,
        created_by=current_user.display_name
    )
    db.session.add(nouvelle_t)
    db.session.commit()
    # Auto-post corresponding journal entry
    _post_trust_journal(nouvelle_t, matter)
    db.session.commit()
    return jsonify({'success': True, 'id': nouvelle_t.fid_id}), 201


@app.route('/api/fiducie/transaction/<int:trans_id>/cancel', methods=['PUT'])
@login_required
def api_fiducie_cancel(trans_id):
    """Cancel (void) a trust transaction."""
    if not current_user.is_manager:
        return jsonify({'error': 'Manager access required'}), 403
    trans = TransactionsFiducie.query.get_or_404(trans_id)
    trans.est_annulee = True
    db.session.commit()
    return jsonify({'success': True, 'id': trans_id, 'est_annulee': True})


@app.route('/api/fiducie/transaction/<int:trans_id>', methods=['DELETE'])
@login_required
def api_fiducie_delete(trans_id):
    """Delete a trust transaction (manager only)."""
    if not current_user.is_manager:
        return jsonify({'error': 'Manager access required'}), 403
    trans = TransactionsFiducie.query.get_or_404(trans_id)
    db.session.delete(trans)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/fiducie/<int:matter_id>/balance', methods=['GET'])
@login_required
def api_fiducie_balance(matter_id):
    """Return the current trust balance for a matter."""
    Matter.query.get_or_404(matter_id)
    transactions = TransactionsFiducie.query.filter_by(matter_id=matter_id).all()
    solde = sum(
        t.montant if t.type_trans == 'DEPOT' else -t.montant
        for t in transactions if not t.est_annulee
    )
    return jsonify({'matter_id': matter_id, 'balance': round(solde, 2)})


# ── Trust Authorization routes ────────────────────────────────────────────────

@app.route('/api/fiducie/<int:matter_id>/authorizations', methods=['GET'])
@login_required
def api_trust_auth_list(matter_id):
    """List all trust authorizations for a matter."""
    try:
        Matter.query.get_or_404(matter_id)
        auths = TrustAuthorization.query.filter_by(matter_id=matter_id).order_by(
            TrustAuthorization.created_at.desc()
        ).all()
        return jsonify([a.to_dict() for a in auths])
    except Exception as exc:
        logger.exception('Error listing trust authorizations for matter %s: %s', matter_id, exc)
        return jsonify({'error': str(exc)}), 500


@app.route('/api/fiducie/<int:matter_id>/authorizations', methods=['POST'])
@login_required
def api_trust_auth_create(matter_id):
    """Create a new trust authorization for a matter."""
    if not current_user.is_manager:
        return jsonify({'error': 'Manager access required'}), 403
    matter = Matter.query.get_or_404(matter_id)
    data = request.get_json() or {}

    # Reject if there is already an active authorization for this matter.
    # Pre-filter in the DB: exclude soft-deleted records (is_active=False OR deleted_at IS NOT NULL)
    # and those whose date range does not overlap with today, then apply is_active_on() for all
    # edge cases.
    today = datetime.now(UTC).date()
    candidate_auths = TrustAuthorization.query.filter(
        TrustAuthorization.matter_id == matter_id,
        TrustAuthorization.is_active == True,  # noqa: E712 - exclude soft-deleted (legacy flag)
        TrustAuthorization.deleted_at == None,  # noqa: E711 - exclude soft-deleted (explicit timestamp)
        db.or_(
            TrustAuthorization.date_from == None,   # noqa: E711
            TrustAuthorization.date_from <= today,
        ),
        db.or_(
            TrustAuthorization.date_to == None,   # noqa: E711 (SQLAlchemy requires ==)
            TrustAuthorization.date_to >= today,
        ),
    ).all()
    if any(a.is_active_on(today) for a in candidate_auths):
        return jsonify({'error': 'Une autorisation active existe déjà pour ce dossier. Veuillez d\'abord supprimer l\'autorisation existante.'}), 409

    date_from = None
    date_to = None
    if data.get('date_from'):
        try:
            date_from = datetime.strptime(data['date_from'], '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'Invalid date_from format, expected YYYY-MM-DD'}), 400
    if data.get('date_to'):
        try:
            date_to = datetime.strptime(data['date_to'], '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'Invalid date_to format, expected YYYY-MM-DD'}), 400

    try:
        auth = TrustAuthorization(
            matter_id=matter_id,
            client_id=matter.client_id,
            date_from=date_from,
            date_to=date_to,
            is_indefinite=not bool(date_to),
            max_amount=float(data['max_amount']) if data.get('max_amount') not in (None, '') else None,
            notes=(data.get('notes') or '').strip() or None,
            created_by=current_user.display_name,
        )
        db.session.add(auth)
        db.session.commit()
        return jsonify(auth.to_dict()), 201
    except Exception as exc:
        db.session.rollback()
        logger.exception('Error creating trust authorization for matter %s: %s', matter_id, exc)
        return jsonify({'error': str(exc)}), 500


@app.route('/api/fiducie/authorizations/<int:auth_id>', methods=['PUT'])
@login_required
def api_trust_auth_update(auth_id):
    """Update a trust authorization (manager only)."""
    if not current_user.is_manager:
        return jsonify({'error': 'Manager access required'}), 403
    auth = TrustAuthorization.query.get_or_404(auth_id)
    data = request.get_json() or {}

    if 'date_from' in data:
        if data['date_from']:
            try:
                auth.date_from = datetime.strptime(data['date_from'], '%Y-%m-%d').date()
            except ValueError:
                return jsonify({'error': 'Invalid date_from format'}), 400
        else:
            auth.date_from = None
    if 'date_to' in data:
        if data['date_to']:
            try:
                auth.date_to = datetime.strptime(data['date_to'], '%Y-%m-%d').date()
            except ValueError:
                return jsonify({'error': 'Invalid date_to format'}), 400
        else:
            auth.date_to = None
        auth.is_indefinite = not bool(auth.date_to)
    if 'max_amount' in data:
        auth.max_amount = float(data['max_amount']) if data['max_amount'] not in (None, '') else None
    if 'notes' in data:
        auth.notes = (data['notes'] or '').strip() or None

    db.session.commit()
    return jsonify(auth.to_dict())


@app.route('/api/fiducie/authorizations/<int:auth_id>', methods=['DELETE'])
@login_required
def api_trust_auth_delete(auth_id):
    """Soft-delete a trust authorization (manager only).

    The record and any associated document file are preserved on disk.
    Both ``is_active`` and ``deleted_at`` are set so that the deletion is
    visible regardless of which column exists in an older database schema.
    """
    if not current_user.is_manager:
        return jsonify({'error': 'Manager access required'}), 403
    auth = TrustAuthorization.query.get_or_404(auth_id)
    try:
        auth.is_active = False
        auth.is_deleted = True
        auth.deleted_at = datetime.utcnow()
        auth.deleted_by = current_user.display_name
        db.session.commit()
        return jsonify({'success': True})
    except Exception as exc:
        db.session.rollback()
        logger.exception('Error soft-deleting trust authorization %s: %s', auth_id, exc)
        return jsonify({'error': str(exc)}), 500


@app.route('/api/fiducie/authorizations/<int:auth_id>/document', methods=['POST'])
@login_required
def api_trust_auth_upload(auth_id):
    """Upload a signed authorization document (PDF/image) for an authorization record."""
    if not current_user.is_manager:
        return jsonify({'error': 'Manager access required'}), 403
    auth = TrustAuthorization.query.get_or_404(auth_id)
    if 'file' not in request.files or not request.files['file'].filename:
        return jsonify({'error': 'No file uploaded'}), 400

    file_obj = request.files['file']
    original_name = secure_filename(file_obj.filename)
    ext = os.path.splitext(original_name)[1].lower()
    allowed_exts = ALLOWED_TRUST_AUTH_EXTENSIONS
    if ext not in allowed_exts:
        return jsonify({'error': f'File type not allowed: {ext}'}), 400

    # Remove old document if it exists
    if auth.doc_filename:
        old_path = os.path.join(app.config['TRUST_AUTH_DOCS_FOLDER'], auth.doc_filename)
        if os.path.exists(old_path):
            try:
                os.remove(old_path)
            except OSError:
                pass

    # Build client_number/matter_number subfolder path
    import uuid as _uuid
    client_number = ''
    matter_number = ''
    if auth.matter:
        matter_number = secure_filename(auth.matter.matter_number or '')
        if auth.matter.client:
            client_number = secure_filename(auth.matter.client.client_number or '')
    subfolder = os.path.join(client_number, matter_number) if client_number and matter_number else ''
    unique_filename = f'{_uuid.uuid4().hex}_{original_name}'
    relative_path = os.path.join(subfolder, unique_filename) if subfolder else unique_filename
    save_dir = os.path.join(app.config['TRUST_AUTH_DOCS_FOLDER'], subfolder) if subfolder else app.config['TRUST_AUTH_DOCS_FOLDER']
    save_path = os.path.join(save_dir, unique_filename)
    try:
        os.makedirs(save_dir, exist_ok=True)
        file_obj.save(save_path)
    except OSError as exc:
        return jsonify({'error': f'Could not save document: {exc}'}), 500

    auth.doc_filename = relative_path
    auth.doc_original_name = original_name
    db.session.commit()
    return jsonify(auth.to_dict()), 200


@app.route('/api/fiducie/authorizations/<int:auth_id>/document', methods=['GET'])
@login_required
def api_trust_auth_document(auth_id):
    """Serve the authorization document file for viewing."""
    import mimetypes
    from flask import send_from_directory
    auth = TrustAuthorization.query.get_or_404(auth_id)
    if not auth.doc_filename:
        return jsonify({'error': 'No document on file'}), 404
    base_dir = os.path.abspath(app.config['TRUST_AUTH_DOCS_FOLDER'])
    doc_path = os.path.join(base_dir, auth.doc_filename)
    if not os.path.exists(doc_path):
        return jsonify({'error': 'Document file not found on server'}), 404
    mime_type = mimetypes.guess_type(auth.doc_filename)[0] or 'application/octet-stream'
    # Validate that the MIME type is in the set of allowed document types
    _allowed_mimes = {
        'application/pdf',
        'image/png', 'image/jpeg', 'image/gif',
        'image/tiff', 'image/bmp',
    }
    if mime_type not in _allowed_mimes:
        mime_type = 'application/octet-stream'
    # doc_filename may contain subdirectory components (client_number/matter_number/file)
    sub_dir = os.path.dirname(auth.doc_filename)
    filename_only = os.path.basename(auth.doc_filename)
    serve_dir = os.path.join(base_dir, sub_dir) if sub_dir else base_dir
    return send_from_directory(serve_dir, filename_only, mimetype=mime_type, as_attachment=False)


@app.route('/api/fiducie/<int:matter_id>/authorizations/active', methods=['GET'])
@login_required
def api_trust_auth_active(matter_id):
    """Return the active authorization for a matter on today's date, if any."""
    Matter.query.get_or_404(matter_id)
    today = datetime.now(UTC).date()
    auths = TrustAuthorization.query.filter_by(matter_id=matter_id).all()
    active = [a for a in auths if a.is_active_on(today)]
    return jsonify([a.to_dict() for a in active])


@app.route('/api/clients/<int:client_id>/trust-balance', methods=['GET'])
@login_required
def api_client_trust_balance(client_id):
    """Return the total trust balance across all matters for a client, with per-matter breakdown."""
    client = Client.query.get_or_404(client_id)
    matters = Matter.query.filter_by(client_id=client_id).filter(Matter.is_deleted == False).all()
    total = 0.0
    breakdown = []
    for matter in matters:
        txns = TransactionsFiducie.query.filter_by(matter_id=matter.id).all()
        balance = sum(
            float(t.montant) if t.type_trans == 'DEPOT' else -float(t.montant)
            for t in txns if not t.est_annulee
        )
        if balance != 0.0:
            breakdown.append({
                'matter_id': matter.id,
                'matter_number': matter.matter_number,
                'matter_description': matter.matter_description or '',
                'balance': round(balance, 2),
            })
        total += balance
    return jsonify({'client_id': client_id, 'trust_balance': round(total, 2), 'matters': breakdown})


@app.route('/api/fiducie/<int:matter_id>/export', methods=['GET'])
@login_required
def api_fiducie_export(matter_id):
    """Export trust transactions for a matter as a CSV file.

    Optional query params: date_from (YYYY-MM-DD), date_to (YYYY-MM-DD)
    """
    matter = Matter.query.get_or_404(matter_id)
    client = matter.client

    date_from_str = request.args.get('date_from', '').strip()
    date_to_str = request.args.get('date_to', '').strip()

    query = TransactionsFiducie.query.filter_by(matter_id=matter_id)
    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
            query = query.filter(TransactionsFiducie.date_trans >= date_from)
        except ValueError:
            pass
    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
            query = query.filter(TransactionsFiducie.date_trans <= date_to)
        except ValueError:
            pass

    transactions = query.order_by(TransactionsFiducie.date_trans.asc()).all()

    output = io.StringIO()
    writer = csv.writer(output, delimiter=',', quoting=csv.QUOTE_MINIMAL)

    client_number = client.client_number if client else ''
    matter_number = matter.matter_number

    writer.writerow(['Client', 'Dossier', 'Date', 'Type', 'Montant', 'Bénéficiaire / Payeur', 'Motif', 'Réf. bancaire', 'Annulée', 'Créé par'])
    for t in transactions:
        writer.writerow([
            client_number,
            matter_number,
            t.date_trans.strftime('%Y-%m-%d') if t.date_trans else '',
            t.type_trans,
            f'{float(t.montant):.2f}',
            t.beneficiaire or '',
            t.motif or '',
            t.ref_bancaire or '',
            'Oui' if t.est_annulee else 'Non',
            t.created_by or '',
        ])

    csv_content = output.getvalue()
    filename = f'fiducie_{client_number}_{matter_number}.csv'
    return Response(
        csv_content.encode('utf-8-sig'),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


@app.route('/api/fiducie/summary', methods=['GET'])
@login_required
def api_fiducie_summary():
    """Return trust summary: one entry per matter with opening balance, transactions, closing balance.
    Supports optional date_from / date_to filters, optional client_id filter, and optional matter_id filter.
    """
    date_from_str = request.args.get('date_from', '').strip()
    date_to_str = request.args.get('date_to', '').strip()
    client_id = request.args.get('client_id')
    matter_id = request.args.get('matter_id')

    date_from = None
    date_to = None
    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
        except ValueError:
            pass

    # Get all matters that have trust transactions
    matter_q = Matter.query.filter(Matter.is_deleted == False)
    if client_id:
        matter_q = matter_q.filter(Matter.client_id == int(client_id))
    if matter_id:
        matter_q = matter_q.filter(Matter.id == int(matter_id))
    matters = matter_q.order_by(Matter.client_id, Matter.matter_number).all()

    results = []
    for matter in matters:
        client = matter.client
        all_txns = TransactionsFiducie.query.filter_by(matter_id=matter.id).filter(
            TransactionsFiducie.est_annulee == False
        ).order_by(TransactionsFiducie.date_trans).all()

        if not all_txns:
            continue  # skip matters with no transactions

        # Opening balance: sum of transactions BEFORE date_from
        if date_from:
            prior_txns = [t for t in all_txns if t.date_trans.date() < date_from]
        else:
            prior_txns = []
        opening_balance = sum(
            float(t.montant) if t.type_trans == 'DEPOT' else -float(t.montant)
            for t in prior_txns
        )

        # Period transactions
        if date_from or date_to:
            period_txns = [
                t for t in all_txns
                if (not date_from or t.date_trans.date() >= date_from)
                and (not date_to or t.date_trans.date() <= date_to)
            ]
        else:
            period_txns = all_txns

        period_deposits = sum(float(t.montant) for t in period_txns if t.type_trans == 'DEPOT')
        period_withdrawals = sum(float(t.montant) for t in period_txns if t.type_trans in ('RETRAIT', 'REMBOURSEMENT'))
        closing_balance = opening_balance + period_deposits - period_withdrawals

        # Collect unique invoice numbers for RETRAIT/REMBOURSEMENT transactions in the period
        invoice_numbers = list(dict.fromkeys(
            t.invoice_number for t in period_txns
            if t.type_trans in ('RETRAIT', 'REMBOURSEMENT') and t.invoice_number
        ))

        results.append({
            'client_number': client.client_number if client else '',
            'client_name': client.client_name if client else '',
            'matter_number': matter.matter_number,
            'matter_description': matter.matter_description or '',
            'opening_balance': round(opening_balance, 2),
            'period_deposits': round(period_deposits, 2),
            'period_withdrawals': round(period_withdrawals, 2),
            'closing_balance': round(closing_balance, 2),
            'txn_count': len(period_txns),
            'invoice_numbers': invoice_numbers,
        })

    return jsonify(results)


@app.route('/api/fiducie/summary-by-account', methods=['GET'])
@login_required
def api_fiducie_summary_by_account():
    """Return trust transaction totals grouped by GL account code.

    Supports optional date_from / date_to filters and optional client_id filter.
    Each row contains: account_code, account_name, total_debit, total_credit, net.
    """
    date_from_str = request.args.get('date_from', '').strip()
    date_to_str = request.args.get('date_to', '').strip()
    client_id = request.args.get('client_id')

    date_from = None
    date_to = None
    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
        except ValueError:
            pass

    # Build journal line query filtered to trust lines
    line_q = JournalLine.query.join(JournalEntry, JournalLine.entry_id == JournalEntry.id).filter(
        JournalLine.is_trust == True
    )
    if date_from:
        line_q = line_q.filter(JournalEntry.entry_date >= date_from)
    if date_to:
        line_q = line_q.filter(JournalEntry.entry_date <= date_to)
    if client_id:
        line_q = line_q.filter(JournalLine.client_id == int(client_id))

    lines = line_q.all()

    # Aggregate by account
    account_totals = {}
    for line in lines:
        account = db.session.get(Account, line.account_id) if line.account_id else None
        code = account.code if account else 'UNKNOWN'
        name = account.name if account else ''
        if code not in account_totals:
            account_totals[code] = {'account_code': code, 'account_name': name, 'total_debit': 0.0, 'total_credit': 0.0}
        account_totals[code]['total_debit'] += float(line.debit or 0)
        account_totals[code]['total_credit'] += float(line.credit or 0)

    results = []
    for row in sorted(account_totals.values(), key=lambda r: r['account_code']):
        row['total_debit'] = round(row['total_debit'], 2)
        row['total_credit'] = round(row['total_credit'], 2)
        row['net'] = round(row['total_debit'] - row['total_credit'], 2)
        results.append(row)

    return jsonify(results)


@login_required
def fiducie_print(matter_id):
    """Render a printable trust statement for a matter with optional date filtering."""
    matter = Matter.query.get_or_404(matter_id)
    client = matter.client
    firm = FirmInfo.query.first()

    date_from_str = request.args.get('date_from', '').strip()
    date_to_str = request.args.get('date_to', '').strip()

    query = TransactionsFiducie.query.filter_by(matter_id=matter_id)
    if date_from_str:
        try:
            df = datetime.strptime(date_from_str, '%Y-%m-%d').date()
            query = query.filter(TransactionsFiducie.date_trans >= df)
        except ValueError:
            pass
    if date_to_str:
        try:
            dt = datetime.strptime(date_to_str, '%Y-%m-%d').date()
            query = query.filter(TransactionsFiducie.date_trans <= dt)
        except ValueError:
            pass

    transactions = query.order_by(TransactionsFiducie.date_trans.asc()).all()

    total_deposits = sum(float(t.montant) for t in transactions if t.type_trans == 'DEPOT' and not t.est_annulee)
    total_withdrawals = sum(float(t.montant) for t in transactions if t.type_trans in ('RETRAIT', 'REMBOURSEMENT') and not t.est_annulee)
    closing_balance = total_deposits - total_withdrawals

    return render_template('fiducie_print.html',
                           matter=matter,
                           client=client,
                           firm=firm,
                           transactions=transactions,
                           total_deposits=total_deposits,
                           total_withdrawals=total_withdrawals,
                           closing_balance=closing_balance,
                           date_from=date_from_str,
                           date_to=date_to_str)


@app.route('/unbilled/print')
@login_required
def unbilled_print():
    """Render a printable accounts-receivable report."""
    firm = FirmInfo.query.first()
    client_id = request.args.get('client_id')

    inv_q = Invoice.query.filter(
        Invoice.status == 'sent'
    )

    if client_id:
        matter_ids = [m.id for m in Matter.query.filter_by(client_id=int(client_id)).all()]
        inv_q = inv_q.filter(
            db.or_(
                Invoice.matter_id.in_(matter_ids),
                Invoice.client_id == int(client_id)
            )
        )

    invoices = inv_q.order_by(Invoice.invoice_date.asc()).all()

    today = datetime.now(UTC).date()
    rows = []
    total_invoice = 0.0
    total_paid = 0.0
    aging_current = 0.0
    aging_30 = 0.0
    aging_60 = 0.0
    aging_60_plus = 0.0

    for inv in invoices:
        client = inv.resolved_client
        amount = float(inv.total_amount or 0)
        paid = float(inv.total_amount or 0) if inv.status == 'paid' else 0.0
        balance = amount - paid
        total_invoice += amount
        total_paid += paid

        days_old = (today - inv.invoice_date).days if inv.invoice_date else 0
        if days_old <= 0:
            aging_current += balance
        elif days_old <= 30:
            aging_30 += balance
        elif days_old <= 60:
            aging_60 += balance
        else:
            aging_60_plus += balance

        rows.append({
            'client_name': client.client_name if client else '',
            'invoice_number': inv.invoice_number,
            'invoice_date': inv.invoice_date,
            'due_date': inv.due_date,
            'amount': amount,
            'paid': paid,
            'balance': balance,
            'days_old': days_old,
        })

    return render_template('ar_print.html',
                           firm=firm,
                           rows=rows,
                           total_invoice=total_invoice,
                           total_paid=total_paid,
                           total_balance=total_invoice - total_paid,
                           aging_current=aging_current,
                           aging_30=aging_30,
                           aging_60=aging_60,
                           aging_60_plus=aging_60_plus,
                           today=today)


@app.route('/statement/print')
@login_required
def statement_print():
    """Render a printable account statement (état de compte) for a client."""
    import calendar as _calendar
    firm = FirmInfo.query.first()
    client_id = request.args.get('client_id')
    date_from_str = request.args.get('date_from', '').strip()
    date_to_str = request.args.get('date_to', '').strip()
    statuses_param = request.args.get('statuses', '').strip()
    allowed_statuses = {'sent', 'paid', 'draft', 'cancelled'}
    if statuses_param:
        filter_statuses = [s for s in statuses_param.split(',') if s in allowed_statuses]
    else:
        filter_statuses = []

    if not date_from_str and not date_to_str:
        now = datetime.now(UTC)
        _, last_day = _calendar.monthrange(now.year, now.month)
        date_from_str = f'{now.year}-{now.month:02d}-01'
        date_to_str = f'{now.year}-{now.month:02d}-{last_day:02d}'

    client = None
    if client_id:
        client = Client.query.get(int(client_id))

    # Get invoices in period
    inv_q = Invoice.query.filter(
        Invoice.invoice_date >= date_from_str,
        Invoice.invoice_date <= date_to_str
    )
    if client_id:
        matter_ids = [m.id for m in Matter.query.filter_by(client_id=int(client_id)).all()]
        inv_q = inv_q.filter(
            db.or_(
                Invoice.matter_id.in_(matter_ids),
                Invoice.client_id == int(client_id)
            )
        )
    invoices = inv_q.order_by(Invoice.invoice_date.asc()).all()

    # Apply status filter if specified
    if filter_statuses:
        invoices = [i for i in invoices if i.status in filter_statuses]

    total_invoiced = sum(float(i.total_amount or 0) for i in invoices)
    total_paid = sum(float(i.total_amount or 0) for i in invoices if i.status == 'paid')
    balance = total_invoiced - total_paid

    # Previous balance (invoices before date_from that are not paid)
    prev_q = Invoice.query.filter(Invoice.invoice_date < date_from_str)
    if client_id:
        prev_q = prev_q.filter(
            db.or_(
                Invoice.matter_id.in_(matter_ids),
                Invoice.client_id == int(client_id)
            )
        )
    prev_invoices = prev_q.all()
    prev_balance = sum(
        float(i.total_amount or 0) for i in prev_invoices if i.status not in ('paid', 'cancelled')
    )

    today = datetime.now(UTC).date()
    return render_template(
        'statement_print.html',
        firm=firm,
        client=client,
        invoices=invoices,
        total_invoiced=total_invoiced,
        total_paid=total_paid,
        balance=balance,
        prev_balance=prev_balance,
        date_from=date_from_str,
        date_to=date_to_str,
        today=today,
    )


@app.route('/logout')
@login_required
def logout():
    # Clear session token to invalidate this session on all devices
    try:
        current_user.session_token = None
        # Clear any active timer session for this user
        ActiveTimerSession.query.filter_by(employee_id=current_user.id).delete()
        db.session.commit()
    except Exception:
        db.session.rollback()
    logout_user()
    # Remove login_token so it cannot linger in the session cookie and
    # interfere with single-session enforcement on the next login.
    session.pop('login_token', None)
    flash('Vous avez été déconnecté.', 'info')
    return redirect(url_for('login'))


@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    lang = session.get('lang', 'fr')
    t = LEXICON.get(lang, LEXICON.get('fr'))
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        user = Employee.query.filter_by(email=email).first()
        if user:
            token = user.get_reset_token()
            email_link_host = os.environ.get('EMAILLINKK', '').strip().rstrip('/')
            if not email_link_host:
                _host = os.environ.get('HOST', '127.0.0.1')
                _prefix = os.environ.get('URL_PREFIX', '').rstrip('/')
                _scheme = 'https' if os.environ.get('USE_HTTPS', '').lower() in ('1', 'true', 'yes') else 'http'
                _default_port = '443' if _scheme == 'https' else '80'
                # EXTERNAL_PORT is the port visible to end-users (e.g. 80
                # behind IIS).  Do NOT fall back to PORT which is the
                # internal Waitress port (5000).
                _port = os.environ.get('EXTERNAL_PORT', _default_port)
                _port_part = '' if _port == _default_port else f':{_port}'
                email_link_host = f"{_scheme}://{_host}{_port_part}{_prefix}"
            reset_url = f"{email_link_host}/reset-password/{token}"
            try:
                send_reset_email(user, reset_url, lang=lang)
            except Exception as mail_err:
                logger.warning('Password reset email failed: %s', mail_err)
            flash(t.get('msg_reset_sent', 'Password reset instructions have been sent to your email address.'), 'success')
        else:
            flash(t.get('msg_no_account', 'No account found with this email address.'), 'danger')
        return redirect(url_for('forgot_password'))
    return render_template('forgot_password.html')


@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password_via_token(token):
    lang = session.get('lang', 'fr')
    t = LEXICON.get(lang, LEXICON.get('fr'))
    user = Employee.verify_reset_token(token)
    if not user:
        flash(t.get('msg_reset_link_invalid', 'The password reset link is invalid or has expired.'), 'danger')
        return redirect(url_for('forgot_password'))
    if request.method == 'POST':
        new_password = request.form.get('new_password', '')
        confirm = request.form.get('confirm_password', '')
        if not new_password:
            flash(t.get('msg_password_empty', 'New password cannot be empty.'), 'danger')
        elif new_password != confirm:
            flash(t.get('msg_passwords_no_match', 'Passwords do not match.'), 'danger')
        else:
            user.set_password(new_password)
            db.session.commit()
            flash(t.get('msg_password_reset_done', 'Your password has been reset. Please sign in.'), 'success')
            return redirect(url_for('login'))
    return render_template('reset_password_token.html', token=token)


@app.route('/reset-password', methods=['GET', 'POST'])
@login_required
def reset_password():
    if request.method == 'POST':
        current_password = request.form.get('current_password', '')
        new_password = request.form.get('new_password', '')
        confirm = request.form.get('confirm_password', '')
        if not current_user.check_password(current_password):
            flash('Current password is incorrect.', 'danger')
        elif not new_password:
            flash('New password cannot be empty.', 'danger')
        elif new_password != confirm:
            flash('New passwords do not match.', 'danger')
        else:
            current_user.set_password(new_password)
            current_user.must_change_password = False
            db.session.commit()
            flash('Password changed successfully.', 'success')
            return redirect(url_for('index'))
    return render_template('reset_password.html')


# ── Page routes ───────────────────────────────────────────────────────────────

@app.route('/')
@login_required
def index():
   return render_template('index.html')


@app.route('/clients')
@login_required
def list_clients():
    return render_template('clients.html')


@app.route('/cost-codes')
@login_required
def list_cost_codes():
    if not current_user.is_manager:
        flash('Access restricted to managers.', 'danger')
        return redirect(url_for('index'))
    codes = CostCode.query.all()
    return render_template('cost_codes.html', codes=codes)


@app.route('/employees')
@login_required
def list_employees():
    if not current_user.is_manager:
        flash('Access restricted to managers.', 'danger')
        return redirect(url_for('index'))
    show_deactivated = request.args.get('show_deactivated', 'false').lower() == 'true'
    q = Employee.query.filter(Employee.is_deleted == False)
    if not show_deactivated:
        q = q.filter(Employee.is_active == True)
    employees = q.order_by(Employee.last_name, Employee.first_name).all()
    return render_template('employees.html', employees=employees, show_deactivated=show_deactivated)


@app.route('/invoices')
@login_required
def list_invoices():
    return render_template('invoices.html')


@app.route('/unbilled')
@login_required
def unbilled_page():
    return render_template('unbilled.html')


@app.route('/api/ar-invoices', methods=['GET'])
@login_required
def api_ar_invoices():
    """Return invoices that are sent but not yet paid (accounts receivable)."""
    q = Invoice.query.filter(Invoice.status == 'sent')
    client_id = request.args.get('client_id')
    if client_id:
        matter_ids = [m.id for m in Matter.query.filter_by(client_id=int(client_id)).all()]
        q = q.filter(
            db.or_(
                Invoice.matter_id.in_(matter_ids),
                Invoice.client_id == int(client_id)
            )
        )
    date_from = request.args.get('date_from')
    if date_from:
        q = q.filter(Invoice.invoice_date >= date_from)
    date_to = request.args.get('date_to')
    if date_to:
        q = q.filter(Invoice.invoice_date <= date_to)
    invoices = q.order_by(Invoice.invoice_date.desc()).all()
    return jsonify([i.to_dict() for i in invoices])


@app.route('/api/draft-invoices', methods=['GET'])
@login_required
def api_draft_invoices():
    """Return invoices in draft status (created but not yet sent)."""
    q = Invoice.query.filter(Invoice.status == 'draft')
    client_id = request.args.get('client_id')
    if client_id:
        matter_ids = [m.id for m in Matter.query.filter_by(client_id=int(client_id)).all()]
        q = q.filter(
            db.or_(
                Invoice.matter_id.in_(matter_ids),
                Invoice.client_id == int(client_id)
            )
        )
    date_from = request.args.get('date_from')
    if date_from:
        q = q.filter(Invoice.invoice_date >= date_from)
    date_to = request.args.get('date_to')
    if date_to:
        q = q.filter(Invoice.invoice_date <= date_to)
    invoices = q.order_by(Invoice.invoice_date.desc()).all()
    return jsonify([i.to_dict() for i in invoices])


@app.route('/api/unbilled', methods=['GET'])
@login_required
def api_unbilled():
    """Return all unbilled expenses, optionally filtered by client_id and date range."""
    q = Expense.query.filter(Expense.is_billed == False, Expense.is_deleted == False)
    client_id = request.args.get('client_id')
    if client_id:
        matter_ids = [m.id for m in Matter.query.filter_by(client_id=int(client_id)).all()]
        q = q.filter(Expense.matter_id.in_(matter_ids))
    date_from = request.args.get('date_from')
    if date_from:
        q = q.filter(Expense.expense_date >= date_from)
    date_to = request.args.get('date_to')
    if date_to:
        q = q.filter(Expense.expense_date <= date_to)
    expenses = q.order_by(Expense.expense_date.desc()).all()
    results = []
    for exp in expenses:
        d = exp.to_dict()
        #matter = Matter.query.get(exp.matter_id) if exp.matter_id else None
        matter = db.session.get(Matter, exp.matter_id) if exp.matter_id else None
        client = matter.client if matter else None
        d['matter_number'] = matter.matter_number if matter else None
        d['client_name'] = client.client_name if client else None
        d['client_id'] = client.id if client else None
        results.append(d)
    return jsonify(results)


@app.route('/ar-invoices')
@login_required
def ar_invoices_page():
    return render_template('ar_invoices.html')


@app.route('/statement')
@login_required
def statement_page():
    return render_template('statement.html')


@app.route('/api/statement', methods=['GET'])
@login_required
def api_statement():
    """Return account statement data for a date range, optionally filtered by client.
    Supports date_from/date_to params (YYYY-MM-DD), falling back to month/year for backward compat."""
    import calendar
    client_id = request.args.get('client_id')

    # Prefer explicit date_from / date_to; fall back to month/year
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    if not date_from and not date_to:
        month = int(request.args.get('month', datetime.now(UTC).month))
        year = int(request.args.get('year', datetime.now(UTC).year))
        _, last_day = calendar.monthrange(year, month)
        date_from = f'{year}-{month:02d}-01'
        date_to = f'{year}-{month:02d}-{last_day:02d}'
    else:
        # If only one is provided, default the other to a reasonable range
        if not date_from:
            date_from = date_to
        if not date_to:
            date_to = date_from

    # Invoiced in period
    inv_q = Invoice.query.filter(
        Invoice.invoice_date >= date_from,
        Invoice.invoice_date <= date_to
    )
    if client_id:
        matter_ids = [m.id for m in Matter.query.filter_by(client_id=int(client_id)).all()]
        inv_q = inv_q.filter(
            db.or_(
                Invoice.matter_id.in_(matter_ids),
                Invoice.client_id == int(client_id)
            )
        )
    invoices = inv_q.all()
    total_invoiced = sum(float(i.total_amount or 0) for i in invoices)
    total_paid = sum(float(i.total_amount or 0) for i in invoices if i.status == 'paid')

    # Unbilled in period
    exp_q = Expense.query.filter(
        Expense.is_billed == False,
        Expense.is_deleted == False,
        Expense.expense_date >= date_from,
        Expense.expense_date <= date_to
    )
    if client_id:
        exp_q = exp_q.filter(Expense.matter_id.in_(matter_ids))
    unbilled_expenses = exp_q.all()
    total_unbilled = sum(float(e.amount or 0) for e in unbilled_expenses)

    return jsonify({
        'date_from': date_from,
        'date_to': date_to,
        'total_invoiced': round(total_invoiced, 2),
        'total_paid': round(total_paid, 2),
        'total_unbilled': round(total_unbilled, 2),
        'outstanding_balance': round(total_invoiced - total_paid, 2),
        'invoices': [i.to_dict() for i in invoices],
    })


# ── Page & API: General Ledger (GL) ──────────────────────────────────────────

@app.route('/gl')
@login_required
def gl_page():
    if not current_user.is_manager:
        flash('Access restricted to managers.', 'danger')
        return redirect(url_for('index'))
    firm = FirmInfo.query.first()
    return render_template('gl.html', firm=firm)


@app.route('/api/gl', methods=['GET'])
@login_required
def api_gl():
    """Return general-ledger style entries for a date range.

    Query params: date_from (YYYY-MM-DD), date_to (YYYY-MM-DD).
    Returns invoices, salary entries and supplier payments as debits/credits.
    Trust/fiducie entries are excluded — they are managed in a separate module.
    Requires manager access.
    """
    if not current_user.is_manager:
        return jsonify({'error': 'access_denied', 'message': 'Manager access required.'}), 403
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()

    if not date_from or not date_to:
        return jsonify({'error': 'date_from and date_to are required'}), 400

    raw_entries = []

    # ── 1. Invoices ───────────────────────────────────────────────────────────
    inv_q = Invoice.query.filter(
        Invoice.invoice_date >= date_from,
        Invoice.invoice_date <= date_to,
        Invoice.status != 'cancelled'
    )
    invoices = inv_q.order_by(Invoice.invoice_date.asc()).all()

    for inv in invoices:
        client = inv.resolved_client
        client_name = client.client_name if client else ''
        accounting_code = (client.accounting_code or '1100') if client else '1100'
        amount = float(inv.total_amount or 0)
        inv_date = inv.invoice_date.isoformat() if inv.invoice_date else ''

        raw_entries.append({
            'date': inv_date,
            'invoice_number': inv.invoice_number,
            'client_name': client_name,
            'accounting_code': accounting_code,
            'description': f'Facture {inv.invoice_number}',
            'debit': round(amount, 2),
            'credit': 0,
        })

        if inv.status == 'paid':
            raw_entries.append({
                'date': inv_date,
                'invoice_number': inv.invoice_number,
                'client_name': client_name,
                'accounting_code': accounting_code,
                'description': f'Paiement – {inv.invoice_number}',
                'debit': 0,
                'credit': round(amount, 2),
            })

    # ── 2. Salary entries ─────────────────────────────────────────────────────
    salary_q = SalaryEntry.query.filter(
        SalaryEntry.is_deleted == False,
        SalaryEntry.entry_date >= date_from,
        SalaryEntry.entry_date <= date_to,
    ).order_by(SalaryEntry.entry_date.asc())

    for entry in salary_q.all():
        cfg = entry.config
        account_code = (cfg.account_code or '') if cfg else ''
        field_name = (cfg.field_name or 'Salaire') if cfg else 'Salaire'
        desc = f'Salaire – {field_name}'
        if entry.description:
            desc += f' – {entry.description}'
        raw_entries.append({
            'date': entry.entry_date.isoformat() if entry.entry_date else '',
            'invoice_number': '',
            'client_name': 'Salaire',
            'accounting_code': account_code,
            'description': desc,
            'debit': round(float(entry.amount or 0), 2),
            'credit': 0,
        })

    # ── 3. Supplier payments ──────────────────────────────────────────────────
    # Use payment_date when available, fall back to invoice_date for filtering.
    sup_all = SupplierPayment.query.filter(
        SupplierPayment.is_deleted == False,
    ).all()

    for payment in sup_all:
        pay_date = payment.payment_date or payment.invoice_date
        if not pay_date:
            continue
        pay_date_str = pay_date.isoformat()
        if pay_date_str < date_from or pay_date_str > date_to:
            continue
        supplier = payment.supplier
        sup_name = supplier.name if supplier else ''
        account_code = (supplier.accounting_code or '2010') if supplier else '2010'
        desc = payment.description or f'Paiement fournisseur {sup_name}'
        raw_entries.append({
            'date': pay_date_str,
            'invoice_number': payment.invoice_number or '',
            'client_name': sup_name,
            'accounting_code': account_code,
            'description': desc,
            'debit': round(float(payment.amount or 0), 2),
            'credit': 0,
        })

    # ── Sort by date, then compute running balance ────────────────────────────
    raw_entries.sort(key=lambda x: x.get('date', ''))

    running_balance = 0.0
    total_debit = 0.0
    total_credit = 0.0
    entries = []
    for e in raw_entries:
        running_balance += e['debit'] - e['credit']
        total_debit += e['debit']
        total_credit += e['credit']
        e['balance'] = round(running_balance, 2)
        entries.append(e)

    return jsonify({
        'entries': entries,
        'total_debit': round(total_debit, 2),
        'total_credit': round(total_credit, 2),
        'balance': round(running_balance, 2),
    })


# ── API: Chart of Accounts ────────────────────────────────────────────────────

@app.route('/accounts')
@login_required
def accounts_page():
    """Codes comptable management page (manager only)."""
    if not current_user.is_manager:
        flash('Access restricted to managers.', 'danger')
        return redirect(url_for('index'))
    firm = FirmInfo.query.first()
    return render_template('accounts.html', firm=firm)


@app.route('/api/accounts', methods=['GET', 'POST'])
@login_required
def api_accounts():
    """List or create chart-of-accounts entries."""
    if not current_user.is_manager:
        return jsonify({'error': 'access_denied', 'message': 'Manager access required.'}), 403
    if request.method == 'GET':
        accounts = Account.query.filter(Account.is_deleted == False).order_by(Account.code).all()
        return jsonify([a.to_dict() for a in accounts])
    # POST – create new account
    data = request.get_json()
    if not data or not data.get('code') or not data.get('name'):
        return jsonify({'error': 'code and name are required'}), 400
    code = data['code'].strip()
    if Account.query.filter_by(code=code).first():
        return jsonify({'error': 'Account code already exists'}), 409
    account = Account(
        code=code,
        name=data['name'].strip(),
        account_type=data.get('account_type', 'expense'),
        parent_id=data.get('parent_id') or None,
        is_active=bool(data.get('is_active', True)),
        is_system=False,
    )
    db.session.add(account)
    db.session.commit()
    return jsonify(account.to_dict()), 201


@app.route('/api/accounts/<int:account_id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def api_account_detail(account_id):
    """Get, update or delete a single account."""
    account = Account.query.get_or_404(account_id)
    if request.method == 'GET':
        return jsonify(account.to_dict())
    if not current_user.is_manager:
        return jsonify({'error': 'access_denied', 'message': 'Manager access required.'}), 403
    if request.method == 'PUT':
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        if 'code' in data and data['code']:
            new_code = data['code'].strip()
            existing = Account.query.filter(Account.code == new_code, Account.id != account_id).first()
            if existing:
                return jsonify({'error': 'Account code already exists'}), 409
            account.code = new_code
        if 'name' in data:
            account.name = data['name'].strip()
        if 'account_type' in data:
            account.account_type = data['account_type']
        if 'parent_id' in data:
            account.parent_id = data['parent_id'] or None
        if 'is_active' in data:
            account.is_active = bool(data['is_active'])
        account.changed_by = current_user.display_name
        account.updated_at = datetime.utcnow()
        db.session.commit()
        return jsonify(account.to_dict())
    # DELETE – forbid deleting system accounts; soft-delete non-system accounts
    if account.is_system:
        return jsonify({'error': 'Cannot delete a system account'}), 409
    account.is_deleted = True
    account.is_active = False
    db.session.commit()
    return '', 204


# ── API: Journal Entries ──────────────────────────────────────────────────────

@app.route('/api/journal-entries', methods=['GET', 'POST'])
@login_required
def api_journal_entries():
    """List journal entries with optional filters, or create a manual entry."""
    if not current_user.is_manager:
        return jsonify({'error': 'access_denied', 'message': 'Manager access required.'}), 403

    if request.method == 'GET':
        date_from = request.args.get('date_from', '').strip()
        date_to = request.args.get('date_to', '').strip()
        client_id = request.args.get('client_id')
        matter_id = request.args.get('matter_id')
        source_type = request.args.get('source_type', '').strip()
        is_trust = request.args.get('is_trust', '').strip().lower()

        q = JournalEntry.query
        if date_from:
            try:
                q = q.filter(JournalEntry.entry_date >= datetime.strptime(date_from, '%Y-%m-%d').date())
            except ValueError:
                pass
        if date_to:
            try:
                q = q.filter(JournalEntry.entry_date <= datetime.strptime(date_to, '%Y-%m-%d').date())
            except ValueError:
                pass
        if source_type:
            q = q.filter(JournalEntry.source_type == source_type)

        # Filter by client or matter via journal lines sub-query
        if client_id or matter_id or is_trust == 'true':
            line_q = JournalLine.query
            if client_id:
                line_q = line_q.filter(JournalLine.client_id == int(client_id))
            if matter_id:
                line_q = line_q.filter(JournalLine.matter_id == int(matter_id))
            if is_trust == 'true':
                line_q = line_q.filter(JournalLine.is_trust == True)
            entry_ids = [l.entry_id for l in line_q.all()]
            if entry_ids:
                q = q.filter(JournalEntry.id.in_(entry_ids))
            else:
                return jsonify([])

        entries = q.order_by(JournalEntry.entry_date.desc()).all()
        return jsonify([e.to_dict() for e in entries])

    # POST – create a manual journal entry
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data provided'}), 400
    lines_data = data.get('lines', [])
    if len(lines_data) < 2:
        return jsonify({'error': 'A journal entry requires at least 2 lines'}), 400

    # Validate balanced entry
    total_debit = round(sum(float(l.get('debit', 0)) for l in lines_data), 2)
    total_credit = round(sum(float(l.get('credit', 0)) for l in lines_data), 2)
    if abs(total_debit - total_credit) > 0.005:
        return jsonify({'error': f'Journal entry is not balanced (debit {total_debit} ≠ credit {total_credit})'}), 400

    try:
        entry_date = datetime.fromisoformat(data['entry_date']).date()
    except (KeyError, ValueError):
        entry_date = datetime.utcnow().date()

    entry = _create_journal_entry(
        entry_date=entry_date,
        description=data.get('description', ''),
        source_type='manual',
        source_id=None,
        lines=lines_data,
        created_by=current_user.display_name,
    )
    if entry is None:
        return jsonify({'error': 'Could not create journal entry'}), 500
    db.session.commit()
    return jsonify(entry.to_dict()), 201


@app.route('/api/journal-entries/<int:entry_id>', methods=['GET'])
@login_required
def api_journal_entry_detail(entry_id):
    """Return a single journal entry with all its lines."""
    if not current_user.is_manager:
        return jsonify({'error': 'access_denied', 'message': 'Manager access required.'}), 403
    entry = JournalEntry.query.get_or_404(entry_id)
    return jsonify(entry.to_dict())


# ── API: Enhanced GL from journal entries ─────────────────────────────────────

@app.route('/api/gl/journal', methods=['GET'])
@login_required
def api_gl_journal():
    """Return GL lines from journal_entries for a given period.

    Supports optional filters: date_from, date_to, client_id, matter_id,
    account_id, is_trust.
    Returns a flat list of journal lines with running balance per line,
    plus summary totals.
    """
    if not current_user.is_manager:
        return jsonify({'error': 'access_denied', 'message': 'Manager access required.'}), 403

    date_from_str = request.args.get('date_from', '').strip()
    date_to_str = request.args.get('date_to', '').strip()
    client_id = request.args.get('client_id')
    matter_id = request.args.get('matter_id')
    account_id = request.args.get('account_id')
    is_trust = request.args.get('is_trust', '').strip().lower()

    # Build a joined query: journal_lines JOIN journal_entries
    q = (db.session.query(JournalLine, JournalEntry)
         .join(JournalEntry, JournalLine.entry_id == JournalEntry.id)
         .order_by(JournalEntry.entry_date.asc(), JournalEntry.id.asc(), JournalLine.id.asc()))

    if date_from_str:
        try:
            q = q.filter(JournalEntry.entry_date >= datetime.strptime(date_from_str, '%Y-%m-%d').date())
        except ValueError:
            pass
    if date_to_str:
        try:
            q = q.filter(JournalEntry.entry_date <= datetime.strptime(date_to_str, '%Y-%m-%d').date())
        except ValueError:
            pass
    if client_id:
        q = q.filter(JournalLine.client_id == int(client_id))
    if matter_id:
        q = q.filter(JournalLine.matter_id == int(matter_id))
    if account_id:
        q = q.filter(JournalLine.account_id == int(account_id))
    if is_trust == 'true':
        q = q.filter(JournalLine.is_trust == True)

    rows = q.all()

    entries_out = []
    running_balance = 0.0
    total_debit = 0.0
    total_credit = 0.0

    # Group lines by entry
    from itertools import groupby
    grouped = {}
    for line, entry in rows:
        if entry.id not in grouped:
            grouped[entry.id] = {'entry': entry, 'lines': []}
        grouped[entry.id]['lines'].append(line)

    for entry_id_key in sorted(grouped.keys()):
        grp = grouped[entry_id_key]
        entry = grp['entry']
        entry_lines = grp['lines']

        entry_block = {
            'entry_id': entry.id,
            'entry_date': entry.entry_date.isoformat() if entry.entry_date else '',
            'description': entry.description or '',
            'source_type': entry.source_type or '',
            'source_id': entry.source_id,
            'lines': [],
        }

        for line in entry_lines:
            d = float(line.debit or 0)
            c = float(line.credit or 0)
            running_balance += d - c
            total_debit += d
            total_credit += c

            # Resolve client / matter names
            client_obj = Client.query.get(line.client_id) if line.client_id else None
            matter_obj = Matter.query.get(line.matter_id) if line.matter_id else None

            entry_block['lines'].append({
                'line_id': line.id,
                'account_code': line.account.code if line.account else '',
                'account_name': line.account.name if line.account else '',
                'account_type': line.account.account_type if line.account else '',
                'debit': round(d, 2),
                'credit': round(c, 2),
                'balance': round(running_balance, 2),
                'client_name': client_obj.client_name if client_obj else '',
                'matter_number': matter_obj.matter_number if matter_obj else '',
                'is_trust': line.is_trust,
                'memo': line.memo or '',
            })

        entries_out.append(entry_block)

    return jsonify({
        'entries': entries_out,
        'total_debit': round(total_debit, 2),
        'total_credit': round(total_credit, 2),
        'balance': round(running_balance, 2),
    })


@app.route('/api/gl/export', methods=['GET'])
@login_required
def api_gl_export():
    """Export GL data as CSV.

    Query params: date_from, date_to, client_id, view (classic|journal).
    Returns a downloadable CSV file.
    """
    if not current_user.is_manager:
        return jsonify({'error': 'access_denied', 'message': 'Manager access required.'}), 403

    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()
    client_id = request.args.get('client_id')
    view = request.args.get('view', 'journal')

    if not date_from or not date_to:
        return jsonify({'error': 'date_from and date_to are required'}), 400

    output = io.StringIO()
    writer = csv.writer(output)

    if view == 'classic':
        inv_q = Invoice.query.filter(
            Invoice.invoice_date >= date_from,
            Invoice.invoice_date <= date_to,
            Invoice.status != 'cancelled',
        )
        if client_id:
            matter_ids = [m.id for m in Matter.query.filter_by(client_id=int(client_id)).all()]
            inv_q = inv_q.filter(
                db.or_(Invoice.matter_id.in_(matter_ids), Invoice.client_id == int(client_id))
            )
        invoices = inv_q.order_by(Invoice.invoice_date.asc()).all()

        writer.writerow(['Date', 'No Facture', 'Client', 'Description', 'Débit', 'Crédit', 'Solde'])
        running = 0.0
        for inv in invoices:
            client = inv.resolved_client
            client_name = client.client_name if client else ''
            amt = float(inv.total_amount or 0)
            running += amt
            writer.writerow([
                inv.invoice_date.isoformat() if inv.invoice_date else '',
                inv.invoice_number,
                client_name,
                f'Facture {inv.invoice_number}',
                f'{amt:.2f}', '', f'{running:.2f}',
            ])
            if inv.status == 'paid':
                running -= amt
                writer.writerow([
                    inv.invoice_date.isoformat() if inv.invoice_date else '',
                    inv.invoice_number,
                    client_name,
                    f'Paiement – {inv.invoice_number}',
                    '', f'{amt:.2f}', f'{running:.2f}',
                ])
    else:
        # Journal view
        from sqlalchemy.orm import joinedload
        q = (
            JournalLine.query
            .join(JournalEntry, JournalLine.entry_id == JournalEntry.id)
            .filter(
                JournalEntry.entry_date >= date_from,
                JournalEntry.entry_date <= date_to,
            )
            .options(joinedload(JournalLine.account))
            .order_by(JournalEntry.entry_date.asc(), JournalEntry.id.asc(), JournalLine.id.asc())
        )
        if client_id:
            q = q.filter(JournalLine.client_id == int(client_id))

        writer.writerow([
            'Date', 'Description', 'Code comptable', 'Compte',
            'Client', 'Dossier', 'Débit', 'Crédit', 'Fiducie',
        ])
        for line in q.all():
            entry = line.entry
            client_obj = Client.query.get(line.client_id) if line.client_id else None
            matter_obj = Matter.query.get(line.matter_id) if line.matter_id else None
            writer.writerow([
                entry.entry_date.isoformat() if entry.entry_date else '',
                entry.description or '',
                line.account.code if line.account else '',
                line.account.name if line.account else '',
                client_obj.client_name if client_obj else '',
                matter_obj.matter_number if matter_obj else '',
                f'{float(line.debit):.2f}' if line.debit else '',
                f'{float(line.credit):.2f}' if line.credit else '',
                'Oui' if line.is_trust else 'Non',
            ])

    output.seek(0)
    filename = f'grand_livre_{date_from}_{date_to}.csv'
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


# ── API: Trust Reconciliation ─────────────────────────────────────────────────

@app.route('/api/trust/reconciliation', methods=['GET', 'POST'])
@login_required
def api_trust_reconciliation():
    """List all trust reconciliations or create a new one."""
    if not current_user.is_manager:
        return jsonify({'error': 'access_denied', 'message': 'Manager access required.'}), 403

    if request.method == 'GET':
        recs = TrustReconciliation.query.order_by(TrustReconciliation.statement_date.desc()).all()
        return jsonify([r.to_dict() for r in recs])

    # POST – create new reconciliation
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data provided'}), 400

    try:
        stmt_date = datetime.strptime(data['statement_date'], '%Y-%m-%d').date()
    except (KeyError, ValueError):
        return jsonify({'error': 'statement_date (YYYY-MM-DD) is required'}), 400

    try:
        bank_balance = round(float(data['bank_balance']), 2)
    except (KeyError, ValueError, TypeError):
        return jsonify({'error': 'bank_balance is required'}), 400

    # Compute GL trust balance from journal lines flagged as trust
    trust_lines = JournalLine.query.filter_by(is_trust=True).all()
    gl_balance = round(sum(float(l.debit or 0) - float(l.credit or 0) for l in trust_lines), 2)

    # If no journal lines exist yet, fall back to TransactionsFiducie
    if not trust_lines:
        all_txns = TransactionsFiducie.query.filter_by(est_annulee=False).all()
        gl_balance = round(sum(
            float(t.montant) if t.type_trans == 'DEPOT' else -float(t.montant)
            for t in all_txns
        ), 2)

    difference = round(bank_balance - gl_balance, 2)

    rec = TrustReconciliation(
        statement_date=stmt_date,
        bank_balance=bank_balance,
        gl_balance=gl_balance,
        difference=difference,
        notes=(data.get('notes') or '').strip() or None,
        created_by=current_user.display_name,
    )
    db.session.add(rec)
    db.session.commit()
    return jsonify(rec.to_dict()), 201


@app.route('/api/trust/reconciliation/<int:rec_id>', methods=['DELETE'])
@login_required
def api_trust_reconciliation_delete(rec_id):
    """Delete a trust reconciliation record (manager only)."""
    if not current_user.is_manager:
        return jsonify({'error': 'access_denied', 'message': 'Manager access required.'}), 403
    rec = TrustReconciliation.query.get_or_404(rec_id)
    db.session.delete(rec)
    db.session.commit()
    return '', 204


@app.route('/invoices/create')
@login_required
def create_invoice_page():
    return render_template('invoice_create.html')


@app.route('/invoices/<int:invoice_id>/print')
@login_required
def print_invoice(invoice_id):
    invoice = Invoice.query.get_or_404(invoice_id)
    firm = FirmInfo.query.first()
    applied_credits = CreditNote.query.filter_by(applied_invoice_id=invoice_id).all()
    return render_template('invoice_print.html', invoice=invoice, firm=firm, applied_credits=applied_credits)


@app.route('/invoices/by-number/<path:invoice_number>/print')
@login_required
def print_invoice_by_number(invoice_number):
    """Look up an invoice by its number and redirect to the print page."""
    invoice = Invoice.query.filter_by(invoice_number=invoice_number).first_or_404()
    return redirect(url_for('print_invoice', invoice_id=invoice.id))


@app.route('/test-db')
@login_required
def test_db():
    try:
        codes = CostCode.query.all()
        return jsonify({'status': 'success', 'cost_codes_count': len(codes)})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/timer')
@login_required
def timer_page():
    if not getattr(current_user, 'timer_user', False):
        flash('Access restricted to timer users.', 'danger')
        return redirect(url_for('index'))
    return render_template('timer.html')


# ── API: Users (removed – all users are managed via the Employees table) ─────


# ── API: Timer ────────────────────────────────────────────────────────────────

@app.route('/api/timer/start', methods=['POST'])
@login_required
def api_timer_start():
    """Register the start of a timer session.

    Enforces the one-timer-at-a-time rule: returns 409 if the user already
    has an active session that is less than 24 hours old.  Stale sessions
    (older than 24 h) are silently replaced so a browser crash cannot lock a
    user out forever.
    """
    if not getattr(current_user, 'timer_user', False):
        return jsonify({'error': 'Access restricted to timer users'}), 403
    data = request.get_json() or {}
    matter_id = data.get('matter_id')
    if not matter_id:
        return jsonify({'error': 'matter_id is required'}), 400

    existing = ActiveTimerSession.query.filter_by(employee_id=current_user.id).first()
    if existing:
        age_hours = (datetime.utcnow() - existing.started_at).total_seconds() / 3600
        if age_hours < 24:
            return jsonify({
                'error': 'Un chronomètre est déjà en cours. Arrêtez-le avant d\'en démarrer un nouveau.'
            }), 409
        # Stale session – clear it
        db.session.delete(existing)

    token = str(uuid.uuid4())
    session_record = ActiveTimerSession(
        employee_id=current_user.id,
        matter_id=matter_id,
        session_token=token,
        started_at=datetime.utcnow()
    )
    db.session.add(session_record)
    db.session.commit()
    return jsonify({'session_token': token}), 201


@app.route('/api/timer/cancel', methods=['DELETE'])
@login_required
def api_timer_cancel():
    """Cancel the current timer session without saving time."""
    if not getattr(current_user, 'timer_user', False):
        return jsonify({'error': 'Access restricted to timer users'}), 403
    ActiveTimerSession.query.filter_by(employee_id=current_user.id).delete()
    db.session.commit()
    return jsonify({'message': 'Chronomètre annulé.'}), 200


@app.route('/api/timer/clients')
@login_required
def api_timer_clients():
    """Return active clients for the timer client-selection list."""
    if not getattr(current_user, 'timer_user', False):
        return jsonify({'error': 'Access restricted to timer users'}), 403
    search = request.args.get('search', '').strip()
    query = Client.query.filter(Client.is_active == True, Client.is_deleted == False)
    if search:
        like = f'%{search}%'
        query = query.filter(
            db.or_(
                Client.client_name.ilike(like),
                Client.client_number.ilike(like)
            )
        )
    results = []
    for client in query.order_by(Client.client_number).all():
        results.append({
            'id': client.id,
            'client_number': client.client_number,
            'client_name': client.client_name,
        })
    return jsonify(results)


@app.route('/api/timer/matters')
@login_required
def api_timer_matters():
    """Return active matters matching an optional search term (for timer UI).

    Optional query params:
      - ``client_id``: restrict results to a single client
      - ``search``: free-text filter across client/matter fields
    """
    if not getattr(current_user, 'timer_user', False):
        return jsonify({'error': 'Access restricted to timer users'}), 403
    search = request.args.get('search', '').strip()
    client_id = request.args.get('client_id', type=int)
    query = (
        db.session.query(Matter, Client)
        .join(Client, Matter.client_id == Client.id)
        .filter(Matter.is_active == True, Matter.is_deleted == False,
                Client.is_deleted == False)
    )
    if client_id:
        query = query.filter(Matter.client_id == client_id)
    if search:
        like = f'%{search}%'
        query = query.filter(
            db.or_(
                Client.client_name.ilike(like),
                Client.client_number.ilike(like),
                Matter.matter_number.ilike(like),
                Matter.matter_description.ilike(like)
            )
        )
    results = []
    for matter, client in query.order_by(Client.client_number, Matter.matter_number).limit(50).all():
        results.append({
            'id': matter.id,
            'matter_number': matter.matter_number,
            'matter_description': matter.matter_description or '',
            'client_number': client.client_number,
            'client_name': client.client_name,
        })
    return jsonify(results)


@app.route('/api/timer/time-entry', methods=['POST'])
@login_required
def api_timer_time_entry():
    """Create an expense row from a completed timer session."""
    if not getattr(current_user, 'timer_user', False):
        return jsonify({'error': 'Access restricted to timer users'}), 403
    if not current_user.hourly_rate or float(current_user.hourly_rate) <= 0:
        return jsonify({'error': 'Hourly rate is not configured. Please contact your manager.'}), 400
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data provided'}), 400
    matter_id = data.get('matter_id')
    if not matter_id:
        return jsonify({'error': 'matter_id is required'}), 400
    #matter = Matter.query.get(matter_id)
    matter = db.session.get(Matter, matter_id)
    if not matter:
        return jsonify({'error': 'Matter not found'}), 404
    duration_seconds = data.get('duration_seconds')
    if not duration_seconds or float(duration_seconds) <= 0:
        return jsonify({'error': 'duration_seconds must be a positive number'}), 400
    # Resolve the active "Time Entry" cost code dynamically
    time_entry_code = CostCode.query.filter(
        db.func.lower(db.func.ltrim(db.func.rtrim(CostCode.description))) == 'time entry',
        CostCode.is_active == True
    ).first()
    if not time_entry_code:
        return jsonify({
            'error': 'Active "Time Entry" cost code not found. '
                     'Please ask your manager to add a cost code with description "Time Entry".'
        }), 400
    # Compute hours and amount — round UP to nearest 0.1 h (6-min legal billing increment)
    hours = math.ceil(float(duration_seconds) / 360.0) / 10.0
    amount = _round_half_up(hours * float(current_user.hourly_rate))
    # Build description
    description = f'Time Entry ({hours:.1f}h)'
    note = (data.get('note') or '').strip()
    if note:
        description += f' \u2013 {note}'
    # Parse optional expense date
    expense_date_raw = data.get('expense_date')
    if expense_date_raw:
        try:
            expense_date = datetime.fromisoformat(expense_date_raw).date()
        except ValueError:
            expense_date = datetime.utcnow().date()
    else:
        #expense_date = datetime.utcnow().date()
        expense_date = datetime.now(timezone.utc).date()
    #now = datetime.utcnow()
    now = datetime.now(timezone.utc)
    expense = Expense(
        matter_id=matter_id,
        code=time_entry_code.code,
        employee_id=current_user.id,
        username=current_user.username,
        description=description,
        amount=amount,
        expense_date=expense_date,
        is_billed=False,
        invoice_id=None,
        created_at=now,
        updated_at=now
    )
    db.session.add(expense)
    db.session.flush()  # get expense.id before commit
    second_work = SecondWork(
        employee_id=current_user.id,
        matter_id=matter_id,
        seconds_worked=int(duration_seconds),
        expense_id=expense.id,
        recorded_at=now
    )
    db.session.add(second_work)
    # Clear the active timer session now that time has been saved
    ActiveTimerSession.query.filter_by(employee_id=current_user.id).delete()
    db.session.commit()
    return jsonify(expense.to_dict()), 201


# ── API: Firm Info ─────────────────────────────────────────────────────────────

@app.route('/api/firm-info', methods=['GET', 'PUT'])
def api_firm_info():
    """Get or update firm information. GET is public; PUT is available without auth
    so it can be used from the login page firm-info dialog."""
    firm = FirmInfo.query.first()
    if request.method == 'GET':
        if firm:
            return jsonify(firm.to_dict())
        return jsonify({'id': None, 'firm_name': '', 'address_line1': '', 'address_line2': '',
                        'city': '', 'province': '', 'postal_code': '', 'phone': '', 'email': '',
                        'tax_number': '', 'logo_filename': '',
                        'tax1_name': 'GST', 'tax2_name': '', 'tax2_compound': False,
                        'tax1_rate': 0.0, 'tax2_rate': 0.0, 'mfa_enabled': False})
    # PUT – create or update
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data provided'}), 400
    if not firm:
        firm = FirmInfo()
        db.session.add(firm)
    for field in ('firm_name', 'address_line1', 'address_line2', 'city', 'province',
                  'postal_code', 'phone', 'email', 'tax_number'):
        if field in data:
            setattr(firm, field, data[field] or None)
    # Tax name fields: empty string means "hide this tax"
    if 'tax1_name' in data:
        firm.tax1_name = data['tax1_name'] or None
    if 'tax2_name' in data:
        firm.tax2_name = data['tax2_name'] or None
    if 'tax2_compound' in data:
        firm.tax2_compound = bool(data['tax2_compound'])
    if 'tax1_rate' in data:
        try:
            firm.tax1_rate = float(data['tax1_rate']) if data['tax1_rate'] is not None else 0.0
        except (ValueError, TypeError):
            firm.tax1_rate = 0.0
    if 'tax2_rate' in data:
        try:
            firm.tax2_rate = float(data['tax2_rate']) if data['tax2_rate'] is not None else 0.0
        except (ValueError, TypeError):
            firm.tax2_rate = 0.0
    if 'mfa_enabled' in data:
        firm.mfa_enabled = bool(data['mfa_enabled'])
    if not firm.firm_name:
        firm.firm_name = 'Your Law Firm'
    db.session.commit()
    return jsonify(firm.to_dict())


@app.route('/api/firm-info/logo', methods=['POST'])
def api_firm_logo():
    """Upload or replace the firm logo. Available without auth so it can be
    used from the login page setup dialog."""
    if 'logo' not in request.files or not request.files['logo'].filename:
        return jsonify({'error': 'No file uploaded'}), 400
    file = request.files['logo']
    # Validate extension from filename
    ext = os.path.splitext(file.filename)[1].lower().lstrip('.')
    if ext not in ALLOWED_IMAGE_EXTENSIONS:
        return jsonify({'error': 'Invalid file type. Allowed: png, jpg, jpeg, gif, svg, webp'}), 400
    # Also validate MIME type reported by browser as a second check
    mime = (file.content_type or '').lower()
    allowed_mimes = {'image/png', 'image/jpeg', 'image/gif', 'image/svg+xml', 'image/webp',
                     'image/jpg', 'application/octet-stream'}
    if mime and mime not in allowed_mimes:
        return jsonify({'error': 'Invalid file content type'}), 400
    filename = secure_filename('firm_logo.' + ext)
    save_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)

    # ── 1. Save file to disk ─────────────────────────────────────────────────
    try:
        for old_ext in ALLOWED_IMAGE_EXTENSIONS:
            old_path = os.path.join(app.config['UPLOAD_FOLDER'], f'firm_logo.{old_ext}')
            if old_path != save_path and os.path.exists(old_path):
                try:
                    os.remove(old_path)
                except OSError as rm_exc:
                    logger.warning('Could not remove old logo file %s: %s', old_path, rm_exc)
        file.save(save_path)
    except OSError as exc:
        logger.error('Logo file save failed: %s', exc, exc_info=True)
        return jsonify({'error': f'Could not save logo file: {exc}'}), 500

    # ── 2. Persist filename in database ─────────────────────────────────────
    try:
        firm = FirmInfo.query.first()
        if not firm:
            firm = FirmInfo(firm_name='Your Law Firm')
            db.session.add(firm)
        firm.logo_filename = filename
        db.session.commit()
        return jsonify({'logo_filename': filename, 'logo_url': f'/static/uploads/{filename}'})
    except Exception as exc:
        logger.error('Logo DB update failed: %s', exc, exc_info=True)
        db.session.rollback()
        return jsonify({'error': f'Logo saved to disk but database update failed '
                                 f'({type(exc).__name__}). Please check server logs.'}), 500


# ── API: License Info ─────────────────────────────────────────────────────────

@app.route('/api/license-info', methods=['GET'])
def api_license_info():
    """Return current license status and details for display in the Setup modal."""
    license_path, public_key_b64 = _licensing.get_license_config(app.config)
    result = _licensing.get_cached_license_result(license_path, public_key_b64)
    data = result.license_data or {}
    return jsonify({
        'status': result.status.value,
        'message': result.message,
        'license_id': data.get('license_id'),
        'issued_to': data.get('issued_to'),
        'expires_at': data.get('expires_at'),
        'fingerprint': _licensing.get_cached_fingerprint(),
        'license_path': license_path,
    })


@app.route('/api/manager', methods=['GET', 'PUT'])
def api_manager():
    """Get or create/update the application Manager employee. Available without
    auth so it can be used from the login-page setup dialog."""
    manager = Employee.query.filter_by(is_manager=True).first()
    if request.method == 'GET':
        if manager:
            return jsonify(manager.to_dict())
        return jsonify({})
    # PUT – create or update
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data provided'}), 400
    if not manager:
        manager = Employee(is_manager=True, is_active=True, is_user=True)
        db.session.add(manager)
    for field in ('first_name', 'last_name', 'email', 'personal_email', 'title', 'phone_number', 'address'):
        if field in data:
            setattr(manager, field, data[field] or None)
    # Handle username – must be unique
    new_username = (data.get('username') or '').strip()
    if new_username:
        query = Employee.query.filter(Employee.username == new_username)
        if manager.id:
            query = query.filter(Employee.id != manager.id)
        if query.first():
            return jsonify({'error': 'Username already taken'}), 409
        manager.username = new_username
    # Handle password update
    if data.get('password'):
        manager.set_password(data['password'])
    db.session.flush()
    # Ensure a matching HR record exists for this manager (idempotent)
    _ensure_hr_record(manager)
    db.session.commit()
    return jsonify(manager.to_dict())


# ── Helper utilities ─────────────────────────────────────────────────────────

def _ensure_hr_record(employee):
    """Create an HrRecord for the given employee if one does not exist yet.

    This is called after any employee/manager is created so that every
    employee always has a corresponding HR record.
    """
    if not employee.id:
        return
    if not HrRecord.query.filter_by(employee_id=employee.id).first():
        hr_record = HrRecord(employee_id=employee.id)
        db.session.add(hr_record)


def _parse_date(val):
    """Parse a date string in YYYY-MM-DD format; returns None on failure."""
    if not val:
        return None
    try:
        return datetime.strptime(val, '%Y-%m-%d').date()
    except ValueError:
        return None


_IMPORT_MAX_ERRORS = 20


# ── API: Employees ────────────────────────────────────────────────────────────

@app.route('/api/employees', methods=['GET', 'POST'])
@login_required
def api_employees():
    if request.method == 'GET':
        return jsonify([e.to_dict() for e in Employee.query.filter(Employee.is_deleted == False).order_by(Employee.last_name, Employee.first_name).all()])
    data = request.get_json()
    if not data or not data.get('first_name') or not data.get('last_name'):
        return jsonify({'error': 'first_name and last_name are required'}), 400

    # Auto-generate a unique username from first/last name (ASCII letters, digits and dots only)
    def slugify(s):
        s = s.lower().strip()
        # Replace accented characters with ASCII equivalents via encode/decode
        try:
            import unicodedata
            s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')
        except Exception:
            pass
        # Keep only alphanumeric characters; replace spaces/hyphens with nothing
        s = re.sub(r'[^a-z0-9]', '', s)
        return s or 'employee'

    requested_username = (data.get('username') or '').strip()
    if requested_username:
        # Caller specified a username — validate uniqueness
        if Employee.query.filter_by(username=requested_username).first():
            return jsonify({'error': 'Username already taken'}), 409
        username = requested_username
    else:
        # Auto-generate a unique username from first/last name
        base = f"{slugify(data['first_name'])}.{slugify(data['last_name'])}"
        username = base
        counter = 2
        while Employee.query.filter_by(username=username).first():
            username = f"{base}.{counter}"
            counter += 1

    employee = Employee(
        first_name=data['first_name'].strip(),
        last_name=data['last_name'].strip(),
        username=username,
        email=data.get('email') or None,
        personal_email=data.get('personal_email') or None,
        title=data.get('title'),
        address=data.get('address') or None,
        phone_number=data.get('phone_number') or None,
        office_phone=data.get('office_phone') or None,
        social_insurance_number=data.get('social_insurance_number'),
        salary_type=data.get('salary_type'),
        salary=data.get('salary', 0),
        hiring_date=_parse_date(data.get('hiring_date')),
        leave_date=_parse_date(data.get('leave_date')),
        emergency_contact=data.get('emergency_contact') or None,
        emergency_phone=data.get('emergency_phone') or None,
        notes=data.get('notes') or None,
        is_active=data.get('is_active', True),
        is_manager=bool(data.get('is_manager', False)),
        is_user=bool(data.get('is_user', False)),
        timer_user=bool(data.get('timer_user', False)),
        is_accounting=bool(data.get('is_accounting', False)),
        hourly_rate=data.get('hourly_rate') or None,
        pin=data.get('pin') or None,
        group_name=data.get('group_name') or None,
        network_id=data.get('network_id') or None,
        supervisor=data.get('supervisor') or None,
        added_by=current_user.display_name,
    )
    # Managers must always be able to log in
    if employee.is_manager:
        employee.is_user = True
    if data.get('password'):
        employee.set_password(data['password'])
    employee.must_change_password = bool(data.get('must_change_password', False))
    db.session.add(employee)
    db.session.flush()  # obtain employee.id before creating hr_record
    # Automatically create an HR record for every new employee
    _ensure_hr_record(employee)
    db.session.commit()
    return jsonify(employee.to_dict()), 201

# ICI DEBUTE IMPORT COSTCODE
@app.route('/api/import/cost-codes', methods=['POST'])
@login_required
def api_import_cost_codes():
    """Import cost codes from a CSV/Excel file.

    Required columns: code, description
    Optional columns: charge_type, rate, is_active
    Optional form field: file_action (keep|rename|delete) – what to do with the
    uploaded file after a successful import.

    Existing codes are skipped (reported as failed rows); they are not updated.
    """
    if not current_user.is_manager:
        return jsonify({'error': 'access_denied', 'message': 'Manager access required.'}), 403
    if 'file' not in request.files or not request.files['file'].filename:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400

    file_obj = request.files['file']
    file_action = (request.form.get('file_action') or 'keep').strip().lower()

    # Save the file to the import upload folder so it can be renamed/deleted later.
    try:
        save_path, original_name, import_id = _save_import_file(file_obj)
    except Exception as exc:
        return jsonify({'success': False, 'error': f'Could not save uploaded file: {exc}'}), 400

    try:
        rows = _parse_uploaded_file(file_obj)
    except ValueError as exc:
        return jsonify({'success': False, 'error': str(exc)}), 400
    except Exception as exc:
        return jsonify({'success': False, 'error': f'Could not read file: {exc}'}), 400

    total = len(rows)
    imported = 0
    failed = 0
    errors = []

    try:
        for i, row in enumerate(rows, start=2):
            code_val = (row.get('code') or '').strip()
            description_val = (row.get('description') or '').strip()
            charge_type_val = (row.get('charge_type') or 'Professional Services').strip()
            rate_raw = (row.get('rate') or '0').strip()
            is_active_raw = (row.get('is_active') or 'true').strip().lower()

            if not code_val:
                errors.append(f'Row {i}: code is required')
                failed += 1
                continue

            if not description_val:
                errors.append(f'Row {i}: description is required')
                failed += 1
                continue

            try:
                rate_val = float(rate_raw) if rate_raw else 0.0
            except ValueError:
                errors.append(f'Row {i}: invalid rate "{rate_raw}"')
                failed += 1
                continue

            is_active_val = is_active_raw not in ('false', '0', 'no', 'inactive')

            if CostCode.query.filter_by(code=code_val).first():
                errors.append(f'Row {i}: cost code "{code_val}" already exists, skipped')
                failed += 1
                continue

            new_code = CostCode(
                code=code_val,
                description=description_val,
                charge_type=charge_type_val,
                rate=rate_val,
                is_active=is_active_val,
                created_at=datetime.utcnow()
            )
            db.session.add(new_code)
            imported += 1

        if imported > 0:
            db.session.commit()
        else:
            db.session.rollback()

        log_entry = ImportLog(
            import_id=import_id,
            filename=original_name,
            records_imported=imported,
            records_failed=failed,
            status='success' if failed == 0 else ('partial' if imported > 0 else 'failed'),
            error_message='; '.join(errors[:_IMPORT_MAX_ERRORS]) if errors else None
        )
        db.session.add(log_entry)
        db.session.commit()

    except Exception as exc:
        db.session.rollback()
        logger.exception('Import cost-codes failed: %s', exc)
        return jsonify({'success': False, 'error': f'Import failed: {exc}'}), 500

    file_result = _apply_import_file_action(save_path, original_name, file_action)

    _write_import_log('cost-codes', original_name, total, imported, failed,
                      errors[:_IMPORT_MAX_ERRORS], file_result)

    return jsonify({
        'success': True,
        'total': total,
        'imported': imported,
        'failed': failed,
        'errors': errors[:_IMPORT_MAX_ERRORS],
        'file_result': file_result
    })


@app.route('/api/employees/<int:employee_id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def api_employee_detail(employee_id):
    employee = Employee.query.get_or_404(employee_id)
    if request.method == 'GET':
        return jsonify(employee.to_dict())
    if request.method == 'PUT':
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400

        for field in ('first_name', 'last_name', 'email', 'personal_email', 'title', 'address',
                      'phone_number', 'office_phone', 'salary_type', 'emergency_contact', 'emergency_phone', 'notes',
                      'pin', 'group_name', 'network_id', 'supervisor'):
            if field in data:
                setattr(employee, field, data[field] or None)
        if 'salary' in data:
            employee.salary = data['salary'] or 0
        if 'hiring_date' in data:
            employee.hiring_date = _parse_date(data['hiring_date'])
        if 'leave_date' in data:
            employee.leave_date = _parse_date(data['leave_date'])
        if 'is_active' in data:
            employee.is_active = bool(data['is_active'])
        if 'is_manager' in data:
            employee.is_manager = bool(data['is_manager'])
        if 'is_user' in data:
            employee.is_user = bool(data['is_user'])
        if 'timer_user' in data:
            employee.timer_user = bool(data['timer_user'])
        if 'is_accounting' in data:
            employee.is_accounting = bool(data['is_accounting'])
        if 'hourly_rate' in data:
            employee.hourly_rate = data['hourly_rate'] or None
        # Managers must always be able to log in
        if employee.is_manager:
            employee.is_user = True
        if data.get('social_insurance_number'):
            employee.social_insurance_number = data['social_insurance_number']
        new_username = (data.get('username') or '').strip()
        if new_username and new_username != employee.username:
            if Employee.query.filter(
                Employee.username == new_username,
                Employee.id != employee_id
            ).first():
                return jsonify({'error': 'Username already taken'}), 409
            employee.username = new_username
        employee.changed_by = current_user.display_name
        db.session.commit()
        return jsonify(employee.to_dict())
    # Soft-delete: mark as deleted and inactive instead of removing the row
    employee.is_deleted = True
    employee.is_active = False
    employee.deleted_by = current_user.display_name
    db.session.commit()
    return '', 204


# ── Page: HR Records ──────────────────────────────────────────────────────────

@app.route('/hr-records')
@login_required
def hr_records_page():
    if not current_user.is_manager:
        flash('Access restricted to managers.', 'danger')
        return redirect(url_for('index'))
    employees = Employee.query.filter(Employee.is_active == True, Employee.is_deleted == False).order_by(Employee.last_name, Employee.first_name).all()
    records = HrRecord.query.filter(HrRecord.is_deleted == False).order_by(HrRecord.date_last_review.desc()).all()
    return render_template('hr_records.html', employees=employees, records=records)


# ── API: HR Records ───────────────────────────────────────────────────────────

@app.route('/api/hr-records', methods=['GET', 'POST'])
@login_required
def api_hr_records():
    if not current_user.is_manager:
        return jsonify({'error': 'access_denied', 'message': 'Manager access required.'}), 403
    if request.method == 'GET':
        employee_id = request.args.get('employee_id', type=int)
        q = HrRecord.query.filter(HrRecord.is_deleted == False)
        if employee_id:
            q = q.filter_by(employee_id=employee_id)
        return jsonify([r.to_dict() for r in q.order_by(HrRecord.date_last_review.desc()).all()])
    data = request.get_json()
    if not data or not data.get('employee_id'):
        return jsonify({'error': 'employee_id is required'}), 400
    Employee.query.get_or_404(data['employee_id'])
    record = HrRecord(
        employee_id=data['employee_id'],
        balance_pto=data.get('balance_pto', 0),
        date_last_review=_parse_date(data.get('date_last_review')),
        review_comment=data.get('review_comment') or None,
    )
    db.session.add(record)
    db.session.commit()
    return jsonify(record.to_dict()), 201


@app.route('/api/hr-records/<int:record_id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def api_hr_record_detail(record_id):
    if not current_user.is_manager:
        return jsonify({'error': 'access_denied', 'message': 'Manager access required.'}), 403
    record = HrRecord.query.get_or_404(record_id)
    if request.method == 'GET':
        return jsonify(record.to_dict())
    if request.method == 'PUT':
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        if 'balance_pto' in data:
            record.balance_pto = data['balance_pto'] if data['balance_pto'] is not None else 0
        if 'date_last_review' in data:
            record.date_last_review = _parse_date(data['date_last_review'])
        if 'review_comment' in data:
            record.review_comment = data['review_comment'] or None
        record.changed_by = current_user.display_name
        record.updated_at = datetime.now(UTC)
        db.session.commit()
        return jsonify(record.to_dict())
    # Soft-delete: mark as deleted instead of removing the row
    record.is_deleted = True
    db.session.commit()
    return '', 204


# ── API: Time logs (second_works) ─────────────────────────────────────────────

@app.route('/api/time-logs')
@login_required
def api_time_logs():
    """Return second_works records with matter and employee context.

    Managers can view all records or filter by employee_id / matter_id / client_id / date.
    Regular users / timer users see only their own records.
    """
    is_manager = getattr(current_user, 'is_manager', False)
    employee_id_filter = request.args.get('employee_id', type=int)
    matter_id_filter = request.args.get('matter_id', type=int)
    client_id_filter = request.args.get('client_id', type=int)
    date_from_str = request.args.get('date_from')
    date_to_str = request.args.get('date_to')

    q = db.session.query(SecondWork, Matter, Client, Employee).join(
        Matter, SecondWork.matter_id == Matter.id
    ).join(
        Client, Matter.client_id == Client.id
    ).join(
        Employee, SecondWork.employee_id == Employee.id
    )

    if not is_manager:
        q = q.filter(SecondWork.employee_id == current_user.id)
    elif employee_id_filter:
        q = q.filter(SecondWork.employee_id == employee_id_filter)

    if matter_id_filter:
        q = q.filter(SecondWork.matter_id == matter_id_filter)

    if client_id_filter:
        q = q.filter(Matter.client_id == client_id_filter)

    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d')
            q = q.filter(SecondWork.recorded_at >= date_from)
        except ValueError:
            pass

    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d')
            # Add 1 day and use strict `<` so all records on the end date are included.
            date_to_end = date_to + timedelta(days=1)
            q = q.filter(SecondWork.recorded_at < date_to_end)
        except ValueError:
            pass

    results = []
    for sw, matter, client, emp in q.order_by(SecondWork.recorded_at.desc()).limit(500).all():
        results.append({
            'id': sw.id,
            'employee_id': sw.employee_id,
            'employee_name': f'{emp.first_name or ""} {emp.last_name or ""}'.strip() or emp.username,
            'matter_id': sw.matter_id,
            'matter_number': matter.matter_number,
            'client_name': client.client_name,
            'seconds_worked': sw.seconds_worked,
            'hours_worked': round(sw.seconds_worked / 3600, 4),
            'expense_id': sw.expense_id,
            'recorded_at': sw.recorded_at.isoformat() if sw.recorded_at else None,
        })
    return jsonify(results)


# ── Page: Time Logs ───────────────────────────────────────────────────────────

@app.route('/time-logs')
@login_required
def time_logs_page():
    is_manager = getattr(current_user, 'is_manager', False)
    is_timer = getattr(current_user, 'timer_user', False)
    if not is_manager and not is_timer and not getattr(current_user, 'is_user', False):
        flash('Access denied.', 'danger')
        return redirect(url_for('index'))
    employees = Employee.query.filter_by(is_active=True).order_by(Employee.last_name, Employee.first_name).all() if is_manager else []
    clients = Client.query.filter_by(is_active=True).order_by(Client.client_number).all() if is_manager else []
    return render_template('time_logs.html', employees=employees, clients=clients, is_manager=is_manager)


# ── API: Clients ──────────────────────────────────────────────────────────────

@app.route('/api/clients', methods=['GET', 'POST'])
@login_required
def api_clients():
    if request.method == 'GET':
        show_inactive = request.args.get('show_inactive', 'false').lower() == 'true'
        q = Client.query.filter(Client.is_deleted == False)
        if not show_inactive:
            q = q.filter(Client.is_active == True)
        return jsonify([c.to_dict() for c in q.order_by(Client.client_number).all()])
    data = request.get_json()
    if not data or not data.get('client_number') or not data.get('client_name'):
        return jsonify({'error': 'client_number and client_name are required'}), 400
    if Client.query.filter_by(client_number=data['client_number']).first():
        return jsonify({'error': 'Client number already exists'}), 409
    client = Client(
        client_number=data['client_number'],
        client_name=data['client_name'],
        street=data.get('street') or None,
        city=data.get('city') or None,
        state=data.get('state') or None,
        postal_code=data.get('postal_code') or None,
        country=data.get('country') or None,
        contact_name=data.get('contact_name') or None,
        phone=data.get('phone') or None,
        email=data.get('email') or None,
        is_active=data.get('is_active', True),
        created_by=current_user.display_name,
    )
    db.session.add(client)
    db.session.commit()
    return jsonify(client.to_dict()), 201


@app.route('/api/clients/<int:client_id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def api_client_detail(client_id):
    client = Client.query.get_or_404(client_id)
    if request.method == 'GET':
        return jsonify(client.to_dict())
    if request.method == 'PUT':
        data = request.get_json()
        if 'client_number' in data and data['client_number']:
            new_num = data['client_number'].strip()
            if new_num != client.client_number:
                if Client.query.filter(
                    Client.client_number == new_num
                ).filter(
                    Client.id != client_id
                ).first():
                    return jsonify({'error': 'Client number already exists'}), 409
                client.client_number = new_num
        if 'client_name' in data:
            client.client_name = data['client_name']
        for field in ('street', 'city', 'state', 'postal_code', 'country', 'contact_name',
                      'phone', 'email'):
            if field in data:
                setattr(client, field, data[field] or None)
        if 'is_active' in data:
            was_active = client.is_active
            client.is_active = bool(data['is_active'])
            if was_active and not client.is_active:
                client.disabled_by = current_user.display_name
            elif not was_active and client.is_active:
                client.reenabled_by = current_user.display_name
        db.session.commit()
        return jsonify(client.to_dict())
    # Soft-delete: mark as deleted and inactive instead of removing the row
    client.is_deleted = True
    client.is_active = False
    client.deleted_by = current_user.display_name
    db.session.commit()
    return '', 204


@app.route('/api/clients/<int:client_id>/matters', methods=['GET', 'POST'])
@login_required
def api_client_matters(client_id):
    client = Client.query.get_or_404(client_id)
    if request.method == 'GET':
        show_inactive = request.args.get('show_inactive', 'false').lower() == 'true'
        q = Matter.query.filter(Matter.client_id == client_id, Matter.is_deleted == False)
        if not show_inactive:
            q = q.filter(Matter.is_active == True)
        return jsonify([m.to_dict() for m in q.order_by(Matter.matter_number).all()])
    data = request.get_json()
    if not data or not data.get('matter_number'):
        return jsonify({'error': 'matter_number is required'}), 400
    if Matter.query.filter_by(client_id=client_id, matter_number=data['matter_number']).first():
        return jsonify({'error': 'Matter number already exists for this client'}), 409
    matter = Matter(
        client_id=client_id,
        matter_number=data['matter_number'],
        matter_description=data.get('matter_description'),
        is_active=data.get('is_active', True),
        attorney1_name=data.get('attorney1_name') or None,
        attorney1_start_date=_parse_date(data.get('attorney1_start_date')),
        attorney2_name=data.get('attorney2_name') or None,
        attorney2_start_date=_parse_date(data.get('attorney2_start_date')),
        attorney3_name=data.get('attorney3_name') or None,
        attorney3_start_date=_parse_date(data.get('attorney3_start_date')),
        attorney4_name=data.get('attorney4_name') or None,
        attorney4_start_date=_parse_date(data.get('attorney4_start_date')),
        attorney5_name=data.get('attorney5_name') or None,
        attorney5_start_date=_parse_date(data.get('attorney5_start_date')),
        attorney6_name=data.get('attorney6_name') or None,
        attorney6_start_date=_parse_date(data.get('attorney6_start_date')),
        created_by=current_user.display_name,
    )
    db.session.add(matter)
    db.session.commit()
    return jsonify(matter.to_dict()), 201


@app.route('/api/client-matters/<int:matter_id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def api_matter_detail(matter_id):
    matter = Matter.query.get_or_404(matter_id)
    if request.method == 'GET':
        return jsonify(matter.to_dict())
    if request.method == 'PUT':
        data = request.get_json()
        if 'matter_number' in data and data['matter_number']:
            new_num = data['matter_number'].strip()
            if new_num != matter.matter_number:
                if Matter.query.filter(
                    Matter.client_id == matter.client_id
                ).filter(
                    Matter.matter_number == new_num
                ).filter(
                    Matter.id != matter_id
                ).first():
                    return jsonify({'error': 'Matter number already exists for this client'}), 409
                matter.matter_number = new_num
        if 'matter_description' in data:
            matter.matter_description = data['matter_description']
        if 'is_active' in data:
            was_active = matter.is_active
            matter.is_active = bool(data['is_active'])
            if was_active and not matter.is_active:
                matter.disabled_by = current_user.display_name
            elif not was_active and matter.is_active:
                matter.reenabled_by = current_user.display_name
        for field in ('attorney1_name', 'attorney2_name', 'attorney3_name',
                      'attorney4_name', 'attorney5_name', 'attorney6_name'):
            if field in data:
                setattr(matter, field, data[field] or None)
        for n in range(1, 7):
            key = f'attorney{n}_start_date'
            if key in data:
                setattr(matter, key, _parse_date(data[key]))
        db.session.commit()
        return jsonify(matter.to_dict())
    # Soft-delete: mark as deleted and inactive instead of removing the row
    matter.is_deleted = True
    matter.is_active = False
    matter.deleted_by = current_user.display_name
    db.session.commit()
    return '', 204


# ── API: Expenses ─────────────────────────────────────────────────────────────

@app.route('/api/matters/<int:matter_id>/expenses', methods=['GET', 'POST'])
@login_required
def api_matter_expenses(matter_id):
    Matter.query.get_or_404(matter_id)
    if request.method == 'GET':
        show_billed = request.args.get('show_billed', 'false').lower() == 'true'
        q = Expense.query.filter(Expense.matter_id == matter_id, Expense.is_deleted == False)
        if not show_billed:
            q = q.filter(Expense.is_billed == False)
        return jsonify([e.to_dict() for e in q.order_by(Expense.expense_date.desc()).all()])
    data = request.get_json()
    if not data or not data.get('description') or data.get('amount') is None:
        return jsonify({'error': 'description and amount are required'}), 400
    expense_date = data.get('expense_date')
    if expense_date:
        try:
            expense_date = datetime.fromisoformat(expense_date).date()
        except ValueError:
            expense_date = datetime.utcnow().date()
    else:
        expense_date = datetime.utcnow().date()
    expense = Expense(
        matter_id=matter_id,
        code=data.get('code'),
        employee_id=current_user.id,
        username=current_user.username,
        description=data['description'],
        amount=data['amount'],
        expense_date=expense_date,
        is_billed=False
    )
    db.session.add(expense)
    db.session.commit()
    return jsonify(expense.to_dict()), 201


@app.route('/api/clients/<int:client_id>/expenses', methods=['GET'])
@login_required
def api_client_expenses(client_id):
    """Return all expenses for every matter belonging to a client."""
    client = Client.query.get_or_404(client_id)
    show_billed = request.args.get('show_billed', 'false').lower() == 'true'
    matter_ids = [m.id for m in client.matters]
    if not matter_ids:
        return jsonify([])
    q = Expense.query.filter(Expense.matter_id.in_(matter_ids), Expense.is_deleted == False)
    if not show_billed:
        q = q.filter(Expense.is_billed == False)
    return jsonify([e.to_dict() for e in q.order_by(Expense.expense_date.desc()).all()])


@app.route('/api/expenses/<int:expense_id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def api_expense_detail(expense_id):
    expense = Expense.query.get_or_404(expense_id)
    if request.method == 'GET':
        return jsonify(expense.to_dict())
    if request.method == 'PUT':
        data = request.get_json()
        for field in ('description', 'amount', 'code', 'expense_date', 'is_billed', 'invoice_id'):
            if field in data:
                setattr(expense, field, data[field])
        expense.updated_at = datetime.utcnow()
        db.session.commit()
        return jsonify(expense.to_dict())
    # Soft-delete: mark as deleted instead of removing the row
    expense.is_deleted = True
    db.session.commit()
    return '', 204


# ── API: Cost codes ───────────────────────────────────────────────────────────

@app.route('/api/cost-codes', methods=['GET', 'POST'])
@login_required
def cost_codes():
    if request.method == 'GET':
        return jsonify([c.to_dict() for c in CostCode.query.order_by(CostCode.code).all()])
    # Write operations are manager-only
    if not current_user.is_manager:
        return jsonify({'error': 'access_denied', 'message': 'Manager access required.'}), 403
    data = request.get_json()
    code = CostCode(
        code=data['code'],
        description=data['description'],
        charge_type=data.get('charge_type') or None,
        rate=data.get('rate', 0.00),
        account_code=data.get('account_code') or None,
        is_active=data.get('is_active', True)
    )
    db.session.add(code)
    db.session.commit()
    return jsonify(code.to_dict()), 201


@app.route('/api/cost-codes/<int:code_id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def api_cost_code_detail(code_id):
    cost_code = CostCode.query.get_or_404(code_id)
    if request.method == 'GET':
        return jsonify(cost_code.to_dict())
    # Write operations are manager-only
    if not current_user.is_manager:
        return jsonify({'error': 'access_denied', 'message': 'Manager access required.'}), 403
    if request.method == 'PUT':
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        for field in ('description', 'charge_type', 'account_code'):
            if field in data:
                setattr(cost_code, field, data[field] or None)
        if 'code' in data and data['code']:
            # Ensure uniqueness if code is being changed
            existing = CostCode.query.filter(
                CostCode.code == data['code'],
                CostCode.id != code_id
            ).first()
            if existing:
                return jsonify({'error': 'Cost code already exists'}), 409
            cost_code.code = data['code']
        if 'rate' in data:
            cost_code.rate = data['rate'] if data['rate'] is not None else 0
        if 'is_active' in data:
            cost_code.is_active = bool(data['is_active'])
        cost_code.changed_by = current_user.display_name
        cost_code.updated_at = datetime.utcnow()
        db.session.commit()
        return jsonify(cost_code.to_dict())
    db.session.delete(cost_code)
    db.session.commit()
    return '', 204


# ── API: Invoices ─────────────────────────────────────────────────────────────

@app.route('/api/invoices', methods=['GET', 'POST'])
@login_required
def api_invoices():
    if request.method == 'GET':
        invs = Invoice.query.order_by(Invoice.invoice_date.desc()).all()
        return jsonify([i.to_dict() for i in invs])
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data provided'}), 400

    # Resolve matter and/or client
    matter_id = data.get('matter_id') or None
    client_id = data.get('client_id') or None

    if matter_id:
        matter = Matter.query.get_or_404(matter_id)
        client = matter.client
    elif client_id:
        client = Client.query.get_or_404(client_id)
        matter = None
        matter_id = None
    else:
        return jsonify({'error': 'matter_id or client_id is required'}), 400

    # Validate trust authorization before accepting an apply_trust request.
    # A RETRAIT is created in the name of the invoice; the same rule that applies
    # to direct trust withdrawals must also apply here.
    if data.get('apply_trust', False) and matter_id:
        today_auth = datetime.now(UTC).date()
        trust_auths_check = TrustAuthorization.query.filter_by(matter_id=matter_id).all()
        if not any(a.is_active_on(today_auth) for a in trust_auths_check):
            return jsonify({
                'error': 'no_authorization',
                'message': 'Aucune autorisation client active pour ce dossier. '
                           'Veuillez ajouter une autorisation signée avant '
                           'd\'appliquer des fonds en fiducie sur une facture.'
            }), 403

    # Generate invoice number if not provided
    invoice_number = data.get('invoice_number')
    if not invoice_number:
        import uuid
        if matter:
            invoice_number = f'INV-{client.client_number}-{matter.matter_number}-{uuid.uuid4().hex[:6].upper()}'
        else:
            invoice_number = f'INV-{client.client_number}-{uuid.uuid4().hex[:8].upper()}'

    try:
        invoice_date = datetime.fromisoformat(data['invoice_date']).date()
    except (KeyError, ValueError):
        invoice_date = datetime.utcnow().date()

    due_date = None
    if data.get('due_date'):
        try:
            due_date = datetime.fromisoformat(data['due_date']).date()
        except ValueError:
            pass

    gst_rate = float(data.get('gst_rate', 0.0))
    qst_rate = float(data.get('qst_rate', 0.0))
    tax2_compound = bool(data.get('tax2_compound', False))
    expense_ids = data.get('expense_ids', [])

    # Calculate totals from selected expenses
    if matter_id:
        selected_expenses = Expense.query.filter(
            Expense.id.in_(expense_ids),
            Expense.matter_id == matter_id,
            Expense.is_billed == False
        ).all() if expense_ids else []
    else:
        # All-matters invoice: expenses may belong to any matter of this client
        matter_ids = [m.id for m in client.matters]
        selected_expenses = Expense.query.filter(
            Expense.id.in_(expense_ids),
            Expense.matter_id.in_(matter_ids),
            Expense.is_billed == False
        ).all() if expense_ids else []

    subtotal = sum(float(e.amount) for e in selected_expenses)
    gst_amount = _round_half_up(subtotal * gst_rate / 100)
    qst_base = subtotal + gst_amount if tax2_compound else subtotal
    qst_amount = _round_half_up(qst_base * qst_rate / 100)
    total_amount = _round_half_up(subtotal + gst_amount + qst_amount)

    invoice = Invoice(
        matter_id=matter_id,
        client_id=client_id if not matter_id else None,
        invoice_number=invoice_number,
        invoice_date=invoice_date,
        due_date=due_date,
        subtotal=subtotal,
        gst_rate=gst_rate,
        gst_amount=gst_amount,
        qst_rate=qst_rate,
        qst_amount=qst_amount,
        total_amount=total_amount,
        status=data.get('status', 'draft'),
        notes=data.get('notes')
    )
    db.session.add(invoice)
    db.session.flush()

    # Mark selected expenses as billed
    for expense in selected_expenses:
        expense.is_billed = True
        expense.invoice_id = invoice.id
        expense.invoice_number = invoice.invoice_number
        expense.invoice_date = invoice_date

    # Apply available credit notes to reduce invoice total (only when requested)
    credit_applied = 0.0
    if client and data.get('apply_credit', False):
        unapplied_credits = CreditNote.query.filter(
            CreditNote.client_id == client.id,
        ).order_by(CreditNote.created_at).all()
        # Filter to only credits with remaining balance
        unapplied_credits = [cn for cn in unapplied_credits if cn.remaining > 0]
        remaining_total = float(total_amount)
        for cn in unapplied_credits:
            if remaining_total <= 0:
                break
            avail = cn.remaining
            apply = min(avail, remaining_total)
            cn.applied_amount = float(cn.applied_amount or 0) + apply
            cn.applied_invoice_id = invoice.id
            remaining_total -= apply
            credit_applied += apply
        if credit_applied > 0:
            invoice.credit_applied = round(credit_applied, 2)
            invoice.total_amount = round(float(total_amount) - credit_applied, 2)

    # Apply trust balance to reduce invoice total (only when requested)
    trust_applied = 0.0
    if data.get('apply_trust', False) and matter_id:
        trust_transactions = TransactionsFiducie.query.filter_by(matter_id=matter_id).all()
        trust_balance = sum(
            float(t.montant) if t.type_trans == 'DEPOT' else -float(t.montant)
            for t in trust_transactions if not t.est_annulee
        )
        current_total = float(invoice.total_amount)
        if trust_balance > 0 and current_total > 0:
            trust_applied = min(trust_balance, current_total)
            # Create trust withdrawal
            trust_withdrawal = TransactionsFiducie(
                matter_id=matter_id,
                type_trans='RETRAIT',
                montant=trust_applied,
                beneficiaire=client.client_name if client else '',
                motif=f'Paiement facture {invoice.invoice_number}',
                invoice_number=invoice.invoice_number,
                est_annulee=False,
                created_by=current_user.display_name
            )
            db.session.add(trust_withdrawal)
            invoice.trust_applied = round(trust_applied, 2)
            invoice.total_amount = round(current_total - trust_applied, 2)
            if invoice.total_amount <= 0:
                invoice.status = 'paid'

    db.session.commit()
    # Auto-post journal entry for non-draft invoices
    if invoice.status != 'draft':
        _post_invoice_journal(invoice, client)
        db.session.commit()
    return jsonify(invoice.to_dict()), 201

@app.route('/api/invoices/<int:invoice_id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def api_invoice_detail(invoice_id):
    invoice = Invoice.query.get_or_404(invoice_id)
    if request.method == 'GET':
        result = invoice.to_dict()
        result['expenses'] = [e.to_dict() for e in invoice.expenses]
        return jsonify(result)
    if request.method == 'PUT':
        data = request.get_json()
        old_status = invoice.status
        for key, value in data.items():
            if hasattr(invoice, key) and key not in ['id', 'created_at']:
                setattr(invoice, key, value)
        invoice.updated_at = datetime.now(UTC)
        db.session.commit()
        # Auto-post journal entries on relevant status transitions
        new_status = invoice.status
        resolved_client = invoice.resolved_client
        if old_status != 'paid' and new_status == 'paid':
            _post_payment_journal(invoice, resolved_client,
                                  float(invoice.total_amount or 0))
            db.session.commit()
        elif old_status == 'draft' and new_status not in ('draft', 'cancelled'):
            _post_invoice_journal(invoice, resolved_client)
            db.session.commit()
        return jsonify(invoice.to_dict())
    db.session.delete(invoice)
    db.session.commit()
    return '', 204


@app.route('/api/invoices/<int:invoice_id>/cancel', methods=['POST'])
@login_required
def api_invoice_cancel(invoice_id):
    """Cancel an invoice: reset all associated expenses to un-billed state
    (clear invoice_id and invoice_number), then mark the invoice as 'cancelled'.
    """
    invoice = Invoice.query.get_or_404(invoice_id)
    if invoice.status == 'cancelled':
        return jsonify({'error': 'Invoice is already cancelled'}), 409
    # Unbill all associated expenses so they can be re-invoiced in the future
    for expense in invoice.expenses:
        expense.is_billed = False
        expense.invoice_id = None
        expense.invoice_number = None
        expense.invoice_date = None
    #invoice.status = 'cancelled'
    invoice.updated_at = datetime.now(UTC)    
    db.session.commit()
    return jsonify(invoice.to_dict())


@app.route('/api/invoices/post-to-gl', methods=['POST'])
@login_required
def api_invoices_post_to_gl():
    """Post invoices and payments that have no journal entry yet to the GL."""
    if not current_user.is_manager:
        return jsonify({'error': 'access_denied', 'message': 'Manager access required.'}), 403
    # Find invoices not yet present in the journal (no JournalEntry with source_type='invoice')
    posted_invoice_ids = {
        je.source_id
        for je in JournalEntry.query.filter_by(source_type='invoice').all()
    }
    posted_payment_ids = {
        je.source_id
        for je in JournalEntry.query.filter_by(source_type='payment').all()
    }
    invoices = Invoice.query.filter(Invoice.status != 'cancelled').all()
    posted_count = 0
    for inv in invoices:
        client = inv.resolved_client
        if inv.id not in posted_invoice_ids:
            try:
                _post_invoice_journal(inv, client)
                posted_count += 1
            except Exception as exc:
                logger.warning("Could not post invoice %s to GL: %s", inv.id, exc)
        if inv.status == 'paid' and inv.id not in posted_payment_ids:
            try:
                _post_payment_journal(inv, client, float(inv.total_amount or 0))
                posted_count += 1
            except Exception as exc:
                logger.warning("Could not post payment for invoice %s to GL: %s", inv.id, exc)
    db.session.commit()
    return jsonify({'posted': posted_count, 'message': f'{posted_count} entr{"ée" if posted_count == 1 else "ées"} soumise{"" if posted_count == 1 else "s"} au GL.'})


@app.route('/import')
@login_required
def import_page():
    if not current_user.is_manager:
        flash('Access restricted to managers.', 'danger')
        return redirect(url_for('index'))
    return render_template('import.html')


# ── API: Credit Notes ────────────────────────────────────────────────────────

@app.route('/api/credit-notes', methods=['GET', 'POST'])
@login_required
def api_credit_notes():
    if request.method == 'GET':
        client_id = request.args.get('client_id')
        q = CreditNote.query
        if client_id:
            q = q.filter_by(client_id=int(client_id))
        notes = q.order_by(CreditNote.created_at.desc()).all()
        return jsonify([n.to_dict() for n in notes])
    data = request.get_json()
    if not data or not data.get('client_id') or not data.get('amount'):
        return jsonify({'error': 'client_id and amount are required'}), 400
    note = CreditNote(
        client_id=int(data['client_id']),
        amount=float(data['amount']),
        reason=data.get('reason', ''),
        added_by=current_user.display_name,
    )
    db.session.add(note)
    db.session.commit()
    return jsonify(note.to_dict()), 201


@app.route('/api/credit-notes/<int:note_id>', methods=['DELETE'])
@login_required
def api_credit_note_delete(note_id):
    note = CreditNote.query.get_or_404(note_id)
    if float(note.applied_amount or 0) > 0:
        return jsonify({'error': 'Cannot delete a credit note that has been partially applied'}), 400
    db.session.delete(note)
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/clients/<int:client_id>/credit-balance', methods=['GET'])
@login_required
def api_client_credit_balance(client_id):
    """Return the total unapplied credit balance for a client."""
    notes = CreditNote.query.filter_by(client_id=client_id).all()
    balance = sum(n.remaining for n in notes)
    return jsonify({'client_id': client_id, 'credit_balance': round(balance, 2)})


# ── API: Import ───────────────────────────────────────────────────────────────

def _parse_uploaded_file(file_obj):
    """Parse an uploaded CSV or Excel file and return a list of dicts."""
    filename = file_obj.filename or ''
    ext = os.path.splitext(filename)[1].lower()
    if ext in ('.xlsx', '.xls'):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_obj.read()), data_only=True)
            ws = wb.active
            headers = [str(cell.value).strip() if cell.value is not None else '' for cell in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(v is not None for v in row):
                    rows.append({headers[i]: (str(row[i]).strip() if row[i] is not None else '') for i in range(len(headers))})
            return rows
        except ImportError:
            raise ValueError('Excel support requires openpyxl. Please upload a CSV file instead.')
    else:
        content = file_obj.read().decode('utf-8-sig')
        # Normalize line endings: \r\n → \n, then bare \r → \n
        content = content.replace('\r\n', '\n').replace('\r', '\n')
        reader = csv.DictReader(io.StringIO(content))
        # Strip whitespace from header names to tolerate files with padded columns
        if reader.fieldnames:
            reader.fieldnames = [f.strip() for f in reader.fieldnames]
        return [dict(row) for row in reader]


def calculate_file_hash(filepath):
    """Return the SHA-256 hex digest of a file's contents.

    Args:
        filepath: Path to the file to hash.

    Returns:
        str: The SHA-256 hash as a hexadecimal string.
    """
    hasher = hashlib.sha256()
    with open(filepath, 'rb') as f:
        for chunk in iter(lambda: f.read(65536), b''):
            hasher.update(chunk)
    return hasher.hexdigest()


def _save_import_file(file_obj):
    """Save an uploaded import file to IMPORT_UPLOAD_FOLDER and return its saved path.

    A UUID prefix is added to the on-disk filename to prevent collisions when
    multiple files with the same name are uploaded concurrently.  The original
    (sanitised) filename is returned separately so it can be used for logging
    and in user-facing messages.

    Returns (save_path, original_name, import_id) where *import_id* is the
    32-character UUID hex string that prefixes the on-disk filename.
    """
    import uuid as _uuid
    original_name = secure_filename(file_obj.filename or 'import')
    import_id = _uuid.uuid4().hex
    unique_name = f'{import_id}_{original_name}'
    save_path = os.path.join(app.config['IMPORT_UPLOAD_FOLDER'], unique_name)
    file_obj.seek(0)
    file_obj.save(save_path)
    file_obj.seek(0)
    return save_path, original_name, import_id


def _apply_import_file_action(save_path, original_name, action):
    """Apply a post-import file action to the saved upload.

    Args:
        save_path:     Absolute path to the saved (UUID-prefixed) file on disk.
        original_name: The sanitised original filename (used in return messages).
        action:        One of 'keep' (do nothing), 'rename' (append today's date
                       before the extension), or 'delete' (the server copy is
                       kept for audit purposes; the client-side browser is
                       responsible for deleting the original source file via the
                       File System Access API to prevent re-importing duplicates).

    Returns:
        str – a human-readable description of the action taken.
    """
    if action == 'delete':
        # The server copy is kept for audit purposes.
        # The client-side browser will delete the original source file.
        return f'File "{original_name}" kept on server.'

    if action == 'rename':
        try:
            base, ext = os.path.splitext(save_path)
            # Include seconds + microseconds to avoid collisions on same-day renames
            dated_path = f'{base}_{datetime.utcnow().strftime("%Y%m%d_%H%M%S_%f")}{ext}'
            os.rename(save_path, dated_path)
            return f'File "{original_name}" renamed with today\'s date.'
        except OSError as exc:
            return f'Could not rename "{original_name}": {exc}'

    # Default: keep
    return f'File "{original_name}" kept on server.'

def _write_import_log(import_type, filename, total, imported, failed, errors, file_result):
    """Write a plain-text log file for an import operation.

    The log is stored in ``app.config['IMPORT_LOG_FOLDER']`` and named
    ``import_YYYY-MM-DD_HHMMSS.log``.  Each call creates a new file so that
    successive imports are never mixed together.

    Args:
        import_type:  Human-readable label, e.g. ``'costs'`` or ``'matters'``.
        filename:     Original upload filename.
        total:        Total rows in the file.
        imported:     Rows successfully imported.
        failed:       Rows that failed.
        errors:       List of error strings (may be truncated).
        file_result:  Result string from :func:`_apply_import_file_action`.
    """
    log_dir = app.config.get('IMPORT_LOG_FOLDER', '')
    if not log_dir:
        return
    try:
        os.makedirs(log_dir, exist_ok=True)
        now = datetime.now(timezone.utc)
        log_name = f'import_{now.strftime("%Y-%m-%d_%H%M%S")}.log'
        log_path = os.path.join(log_dir, log_name)
        status = 'success' if failed == 0 else ('partial' if imported > 0 else 'failed')
        lines = [
            f'LawLedger Import Log',
            f'====================',
            f'Date/Time (UTC) : {now.strftime("%Y-%m-%d %H:%M:%S")}',
            f'Import type     : {import_type}',
            f'File            : {filename}',
            f'Status          : {status}',
            f'Total rows      : {total}',
            f'Imported        : {imported}',
            f'Failed          : {failed}',
            f'File action     : {file_result}',
        ]
        if errors:
            lines.append('')
            lines.append('Errors:')
            for err in errors:
                lines.append(f'  - {err}')
        lines.append('')
        with open(log_path, 'w', encoding='utf-8') as fh:
            fh.write('\n'.join(lines))
    except OSError:
        pass  # Log writing is best-effort; never block the import response.


@app.route('/api/import/costs', methods=['POST'])
@login_required
def api_import_costs():
    """Import soft costs (expenses) from a CSV/Excel file.

    Required columns: client_number, matter_number, expense_code, amount
    Optional columns: username, expense_date
    Legacy columns also accepted: client_matter_number, expenses_type, date

    The uploaded file is always kept as an archive in IMPORT_FILES_DIR
    (configured via the IMPORT_FILES_DIR environment variable).
    """
    if not current_user.is_manager:
        return jsonify({'error': 'access_denied', 'message': 'Manager access required.'}), 403
    if 'file' not in request.files or not request.files['file'].filename:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400

    file_obj = request.files['file']

    # Save the file to the import upload folder so it can be renamed/deleted later.
    try:
        save_path, original_name, import_id = _save_import_file(file_obj)
    except Exception as exc:
        return jsonify({'success': False, 'error': f'Could not save uploaded file: {exc}'}), 400

    # Anti-duplicate: reject if the same file content was already successfully imported.
    # A previous import with status='failed' (all rows rejected) does NOT block re-importing
    # so that users can fix their data and try again with the same file.
    file_hash = calculate_file_hash(save_path)
    existing = ImportLog.query.filter(
        ImportLog.file_hash == file_hash,
        ImportLog.status.in_(['success', 'partial'])
    ).first()
    if existing:
        if os.path.exists(save_path):
            os.remove(save_path)
        import_date_str = existing.import_date.strftime('%Y-%m-%d %H:%M') if existing.import_date else 'date inconnue'
        return jsonify({
            'success': False,
            'error': f'Fichier déjà importé le {import_date_str} (ID d\'importation : {existing.import_id})'
        }), 400

    try:
        rows = _parse_uploaded_file(file_obj)
    except ValueError as exc:
        return jsonify({'success': False, 'error': str(exc)}), 400
    except Exception as exc:
        return jsonify({'success': False, 'error': f'Could not read file: {exc}'}), 400

    total = len(rows)
    imported = 0
    failed = 0
    errors = []

    try:
        for i, row in enumerate(rows, start=2):
            client_number_raw = (row.get('client_number') or '').strip()
            matter_number_raw = (row.get('matter_number') or '').strip()
            # Legacy fallback: client_matter_number
            legacy_matter = (row.get('client_matter_number') or '').strip()
            expense_code_raw = (row.get('expense_code') or row.get('expenses_type') or '').strip()
            amount_raw = (row.get('amount') or '').strip()
            date_raw = (row.get('expense_date') or row.get('date') or '').strip()
            username_raw = (row.get('username') or '').strip()
            quantity_raw = (row.get('quantity') or '').strip()
            user_pin_raw = (row.get('user_pin') or '').strip()

            if not expense_code_raw:
                errors.append(f'Row {i}: expense_code is required')
                failed += 1
                continue

            if not amount_raw:
                errors.append(f'Row {i}: amount is required')
                failed += 1
                continue

            if not client_number_raw and not matter_number_raw and not legacy_matter:
                errors.append(f'Row {i}: client_number and matter_number are required')
                failed += 1
                continue

            try:
                amount = float(amount_raw)
            except ValueError:
                errors.append(f'Row {i}: invalid amount "{amount_raw}"')
                failed += 1
                continue

            # Parse optional quantity (defaults to 1)
            quantity = 1.0
            if quantity_raw:
                try:
                    quantity = float(quantity_raw)
                except ValueError:
                    errors.append(f'Row {i}: invalid quantity "{quantity_raw}"')
                    failed += 1
                    continue

            # Find matter by client_number + matter_number, or legacy client_matter_number
            matter = None
            if client_number_raw and matter_number_raw:
                client = Client.query.filter_by(client_number=client_number_raw).first()
                if client:
                    matter = Matter.query.filter_by(
                        client_id=client.id, matter_number=matter_number_raw, is_active=True
                    ).first()
                    if not matter:
                        matter = Matter.query.filter_by(
                            client_id=client.id, matter_number=matter_number_raw
                        ).first()
                if not matter:
                    errors.append(
                        f'Row {i}: matter "{matter_number_raw}" for client "{client_number_raw}" not found'
                    )
                    failed += 1
                    continue
            else:
                # Fall back to legacy matter_number lookup
                lookup = matter_number_raw or legacy_matter
                matter = Matter.query.filter_by(matter_number=lookup, is_active=True).first()
                if not matter:
                    matter = Matter.query.filter_by(matter_number=lookup).first()
                if not matter:
                    errors.append(f'Row {i}: matter "{lookup}" not found')
                    failed += 1
                    continue

            # Look up cost code by expense_code (string code)
            cost_code = CostCode.query.filter_by(code=expense_code_raw).first()

            # Resolve employee by user_pin first, then username, fall back to current user
            employee = None
            if user_pin_raw:
                employee = Employee.query.filter_by(pin=user_pin_raw).first()
                if employee is None:
                    errors.append(f'Row {i}: user_pin "{user_pin_raw}" not found, using current user instead')
            if employee is None and username_raw:
                employee = Employee.query.filter_by(username=username_raw).first()
                if employee is None:
                    errors.append(f'Row {i}: username "{username_raw}" not found, using current user instead')
            if employee is None:
                employee = current_user

            # Parse date
            #expense_date = datetime.utcnow().date()
            expense_date = datetime.now(UTC).date()
            if date_raw:
                for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d'):
                    try:
                        expense_date = datetime.strptime(date_raw, fmt).date()
                        break
                    except ValueError:
                        continue

            # Derive description from the cost code lookup; fall back to the expense_code value
            if cost_code:
                description = cost_code.description or expense_code_raw
            else:
                description = expense_code_raw

            expense = Expense(
                matter_id=matter.id,
                code=expense_code_raw,
                employee_id=employee.id,
                username=employee.username,
                description=description,
                amount=amount,
                quantity=quantity,
                user_pin=user_pin_raw or None,
                expense_date=expense_date,
                is_billed=False,
                import_id=import_id
            )
            db.session.add(expense)
            imported += 1

        if imported > 0:
            db.session.commit()
        else:
            db.session.rollback()

        log_entry = ImportLog(
            import_id=import_id,
            filename=original_name,
            file_hash=file_hash,
            records_imported=imported,
            records_failed=failed,
            status='success' if failed == 0 else ('partial' if imported > 0 else 'failed'),
            error_message='; '.join(errors[:_IMPORT_MAX_ERRORS]) if errors else None
        )
        db.session.add(log_entry)
        db.session.commit()

    except Exception as exc:
        db.session.rollback()
        logger.exception('Import costs failed: %s', exc)
        return jsonify({'success': False, 'error': f'Import failed: {exc}'}), 500

    # Always keep the server copy in IMPORT_FILES_DIR (configured via .env).
    file_result = f'File "{original_name}" archived on server.'

    _write_import_log('costs', original_name, total, imported, failed,
                      errors[:_IMPORT_MAX_ERRORS], file_result)

    return jsonify({
        'success': True,
        'total': total,
        'imported': imported,
        'failed': failed,
        'errors': errors[:_IMPORT_MAX_ERRORS],
        'file_result': file_result
    })


@app.route('/api/import/matters', methods=['POST'])
@login_required
def api_import_matters():
    """Import client matters from a CSV/Excel file.

    Expected columns: matter_number (or fmatter_number), client_name, is_active
    Optional columns: client_number, matter_description,
                      street, city, state, postal_code, country,
                      contact_name, phone, email
    Optional form field: file_action (keep|rename|delete) – what to do with the
    uploaded file after a successful import.
    """
    if not current_user.is_manager:
        return jsonify({'error': 'access_denied', 'message': 'Manager access required.'}), 403
    if 'file' not in request.files or not request.files['file'].filename:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400

    file_obj = request.files['file']
    file_action = (request.form.get('file_action') or 'keep').strip().lower()

    # Save the file to the import upload folder so it can be renamed/deleted later.
    try:
        save_path, original_name, import_id = _save_import_file(file_obj)
    except Exception as exc:
        return jsonify({'success': False, 'error': f'Could not save uploaded file: {exc}'}), 400

    try:
        rows = _parse_uploaded_file(file_obj)
    except ValueError as exc:
        return jsonify({'success': False, 'error': str(exc)}), 400
    except Exception as exc:
        return jsonify({'success': False, 'error': f'Could not read file: {exc}'}), 400

    total = len(rows)
    imported = 0
    failed = 0
    errors = []

    try:
        for i, row in enumerate(rows, start=2):
            # Accept both matter_number and fmatter_number (fmatter_number is the
            # column name used by the firm's legacy export format)
            matter_number = (row.get('matter_number') or row.get('fmatter_number') or '').strip()
            client_name = (row.get('client_name') or '').strip()
            client_number = (row.get('client_number') or '').strip()
            matter_description = (row.get('matter_description') or '').strip() or None
            is_active_raw = (row.get('is_active') or 'true').strip().lower()
            is_active = is_active_raw not in ('false', '0', 'no', 'inactive')

            # Extra client detail fields present in the extended format
            client_fields = {
                'street':       (row.get('street') or '').strip() or None,
                'city':         (row.get('city') or '').strip() or None,
                'state':        (row.get('state') or '').strip() or None,
                'postal_code':  (row.get('postal_code') or '').strip() or None,
                'country':      (row.get('country') or '').strip() or None,
                'contact_name': (row.get('contact_name') or '').strip() or None,
                'phone':        (row.get('phone') or '').strip() or None,
                'email':        (row.get('email') or '').strip() or None,
            }

            if not matter_number or not client_name:
                errors.append(f'Row {i}: matter_number and client_name are required')
                failed += 1
                continue

            # Find or create client — match ONLY by client_number when provided,
            # never by name, to avoid merging two different clients with the same name.
            client = None
            if client_number:
                client = Client.query.filter_by(client_number=client_number).first()
            if not client:
                # Only fall back to name lookup when NO client_number was supplied in the row.
                if not client_number:
                    client = Client.query.filter_by(client_name=client_name).first()
            if not client:
                if not client_number:
                    import uuid
                    client_number = 'C-' + uuid.uuid4().hex[:6].upper()
                client = Client(
                    client_number=client_number,
                    client_name=client_name,
                    is_active=True
                )
                db.session.add(client)
            # Populate / update client detail fields when the CSV provides them.
            # Only non-empty values are written; absent or blank columns leave the
            # existing client data unchanged.
            for field, value in client_fields.items():
                if value is not None:
                    setattr(client, field, value)
            db.session.flush()

            # Build the unique key: client_number + matter_number
            # Duplicate detection uses ONLY the numeric keys — never client name.
            effective_client_number = client.client_number or client_number or ''
            client_matter_key = f"{effective_client_number}_{matter_number}"

            # Check for duplicate matter using the unique client_matter_key only.
            if Matter.query.filter_by(client_matter_key=client_matter_key).first():
                errors.append(f'Row {i}: matter "{matter_number}" already exists for client "{effective_client_number}"')
                failed += 1
                continue

            # Secondary guard: same client_id + matter_number (handles legacy rows without a key).
            if Matter.query.filter_by(client_id=client.id, matter_number=matter_number).first():
                errors.append(f'Row {i}: matter "{matter_number}" already exists for client "{effective_client_number}"')
                failed += 1
                continue

            matter = Matter(
                client_id=client.id,
                matter_number=matter_number,
                matter_description=matter_description,
                is_active=is_active,
                client_matter_key=client_matter_key
            )
            db.session.add(matter)
            imported += 1

        if imported > 0:
            db.session.commit()
        else:
            db.session.rollback()

        log_entry = ImportLog(
            import_id=import_id,
            filename=original_name,
            records_imported=imported,
            records_failed=failed,
            status='success' if failed == 0 else ('partial' if imported > 0 else 'failed'),
            error_message='; '.join(errors[:_IMPORT_MAX_ERRORS]) if errors else None
        )
        db.session.add(log_entry)
        db.session.commit()

    except Exception as exc:
        db.session.rollback()
        logger.exception('Import matters failed: %s', exc)
        return jsonify({'success': False, 'error': f'Import failed: {exc}'}), 500

    file_result = _apply_import_file_action(save_path, original_name, file_action)

    _write_import_log('matters', original_name, total, imported, failed,
                      errors[:_IMPORT_MAX_ERRORS], file_result)

    return jsonify({
        'success': True,
        'total': total,
        'imported': imported,
        'failed': failed,
        'errors': errors[:_IMPORT_MAX_ERRORS],
        'file_result': file_result
    })


# ── API: Export ───────────────────────────────────────────────────────────────

@app.route('/api/export/client-matters')
@login_required
def api_export_client_matters():
    """Export active client matters in CSV, TSV, or Excel format."""
    include_inactive = request.args.get('include_inactive', 'false').lower() == 'true'
    fmt = request.args.get('format', 'csv').lower()
    separator = request.args.get('separator', 'comma')
    # Map named separators to actual characters
    sep_map = {'comma': ',', 'semicolon': ';', 'tab': '\t', 'pipe': '|',
               'percent': '%', 'slash': '/', 'dollar': '$', 'plus': '+',
               'ampersand': '&', 'backslash': '\\', 'hash': '#',
               'closeparen': ')', 'openparen': '('}
    sep_char = sep_map.get(separator, separator)
    if len(sep_char) != 1:
        sep_char = ','

    query = (
        db.session.query(Matter, Client)
        .join(Client, Matter.client_id == Client.id)
    )
    if not include_inactive:
        query = query.filter(Matter.is_active == True, Client.is_active == True)
    query = query.order_by(Client.client_number, Matter.matter_number)
    rows_data = query.all()

    headers = ['client_number', 'client_name', 'matter_number', 'matter_description', 'status']

    if fmt == 'xlsx':
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Client Matters'
            ws.append(headers)
            for matter, client in rows_data:
                ws.append([
                    client.client_number,
                    client.client_name,
                    matter.matter_number,
                    matter.matter_description or '',
                    'Active' if matter.is_active else 'Inactive'
                ])
            xls_io = io.BytesIO()
            wb.save(xls_io)
            xls_io.seek(0)
            return Response(
                xls_io.read(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={'Content-Disposition': 'attachment; filename="active_client_matters.xlsx"'}
            )
        except ImportError:
            return jsonify({'error': 'Excel export requires openpyxl. Please install it or choose CSV/TSV.'}), 500

    # CSV / custom separator
    output = io.StringIO()
    writer = csv.writer(output, delimiter=sep_char)
    writer.writerow(headers)
    for matter, client in rows_data:
        writer.writerow([
            client.client_number,
            client.client_name,
            matter.matter_number,
            matter.matter_description or '',
            'Active' if matter.is_active else 'Inactive'
        ])

    csv_bytes = output.getvalue().encode('utf-8')
    ext = 'tsv' if sep_char == '\t' else 'csv'
    mime = 'text/tab-separated-values' if sep_char == '\t' else 'text/csv'
    filename = f'active_client_matters.{ext}'
    return Response(
        csv_bytes,
        mimetype=mime,
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


@app.route('/api/export/employees')
@login_required
def api_export_employees():
    """Export employees in CSV, TSV, or Excel format.

    Exported columns: last_name, first_name, pin, email (work), office_phone,
    address, network_id, personal_email.
    """
    include_inactive = request.args.get('include_inactive', 'false').lower() == 'true'
    fmt = request.args.get('format', 'csv').lower()
    separator = request.args.get('separator', 'comma')
    sep_map = {'comma': ',', 'semicolon': ';', 'tab': '\t', 'pipe': '|',
               'percent': '%', 'slash': '/', 'dollar': '$', 'plus': '+',
               'ampersand': '&', 'backslash': '\\', 'hash': '#',
               'closeparen': ')', 'openparen': '('}
    sep_char = sep_map.get(separator, separator)
    if len(sep_char) != 1:
        sep_char = ','

    query = Employee.query.order_by(Employee.last_name, Employee.first_name)
    if not include_inactive:
        query = query.filter_by(is_active=True)
    employees = query.all()

    headers = ['last_name', 'first_name', 'pin', 'email', 'office_phone',
               'address', 'network_id', 'personal_email']

    if fmt == 'xlsx':
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Employees'
            ws.append(headers)
            for emp in employees:
                ws.append([
                    emp.last_name or '',
                    emp.first_name or '',
                    emp.pin or '',
                    emp.email or '',
                    emp.office_phone or '',
                    emp.address or '',
                    emp.network_id or '',
                    emp.personal_email or '',
                ])
            xls_io = io.BytesIO()
            wb.save(xls_io)
            xls_io.seek(0)
            return Response(
                xls_io.read(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={'Content-Disposition': 'attachment; filename="employees.xlsx"'}
            )
        except ImportError:
            return jsonify({'error': 'Excel export requires openpyxl. Please install it or choose CSV/TSV.'}), 500

    output = io.StringIO()
    writer = csv.writer(output, delimiter=sep_char)
    writer.writerow(headers)
    for emp in employees:
        writer.writerow([
            emp.last_name or '',
            emp.first_name or '',
            emp.pin or '',
            emp.email or '',
            emp.office_phone or '',
            emp.address or '',
            emp.network_id or '',
            emp.personal_email or '',
        ])

    csv_bytes = output.getvalue().encode('utf-8')
    ext = 'tsv' if sep_char == '\t' else 'csv'
    mime = 'text/tab-separated-values' if sep_char == '\t' else 'text/csv'
    filename = f'employees.{ext}'
    return Response(
        csv_bytes,
        mimetype=mime,
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


# ── Page & API: Suppliers (Fournisseurs) ─────────────────────────────────────

@app.route('/suppliers')
@login_required
def suppliers_page():
    if not current_user.is_manager:
        flash('Access restricted to managers.', 'danger')
        return redirect(url_for('index'))
    return render_template('suppliers.html')


@app.route('/api/suppliers', methods=['GET', 'POST'])
@login_required
def api_suppliers():
    if not current_user.is_manager:
        return jsonify({'error': 'access_denied', 'message': 'Manager access required.'}), 403
    if request.method == 'GET':
        search = request.args.get('search', '').strip()
        q = Supplier.query.filter(Supplier.is_deleted == False)
        if search:
            q = q.filter(
                db.or_(
                    Supplier.name.ilike(f'%{search}%'),
                    Supplier.service_provided.ilike(f'%{search}%'),
                    Supplier.phone.ilike(f'%{search}%'),
                )
            )
        suppliers = q.order_by(Supplier.name).all()
        return jsonify([s.to_dict() for s in suppliers])
    # POST – create new supplier
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'name_required', 'message': 'Supplier name is required.'}), 400
    supplier = Supplier(
        name=name,
        account_number=data.get('account_number', ''),
        address=data.get('address', ''),
        phone=data.get('phone', ''),
        email=data.get('email', ''),
        service_provided=data.get('service_provided', ''),
        notes=data.get('notes', ''),
        accounting_code=data.get('accounting_code', ''),
        is_active=True,
        added_by=current_user.display_name,
    )
    db.session.add(supplier)
    db.session.commit()
    return jsonify(supplier.to_dict()), 201


@app.route('/api/suppliers/<int:supplier_id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def api_supplier_detail(supplier_id):
    if not current_user.is_manager:
        return jsonify({'error': 'access_denied', 'message': 'Manager access required.'}), 403
    supplier = db.session.get(Supplier, supplier_id)
    if not supplier or supplier.is_deleted:
        return jsonify({'error': 'not_found'}), 404
    if request.method == 'GET':
        return jsonify(supplier.to_dict())
    if request.method == 'PUT':
        data = request.get_json() or {}
        if 'name' in data:
            name = data['name'].strip()
            if not name:
                return jsonify({'error': 'name_required', 'message': 'Supplier name is required.'}), 400
            supplier.name = name
        for field in ('account_number', 'address', 'phone', 'email', 'service_provided', 'notes', 'accounting_code'):
            if field in data:
                setattr(supplier, field, data[field])
        if 'is_active' in data:
            supplier.is_active = bool(data['is_active'])
        db.session.commit()
        return jsonify(supplier.to_dict())
    # DELETE – soft delete
    supplier.is_deleted = True
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/suppliers/<int:supplier_id>/payments', methods=['GET', 'POST'])
@login_required
def api_supplier_payments(supplier_id):
    if not current_user.is_manager:
        return jsonify({'error': 'access_denied', 'message': 'Manager access required.'}), 403
    supplier = db.session.get(Supplier, supplier_id)
    if not supplier or supplier.is_deleted:
        return jsonify({'error': 'not_found'}), 404
    if request.method == 'GET':
        payments = SupplierPayment.query.filter_by(supplier_id=supplier_id).filter(
            SupplierPayment.is_deleted == False
        ).order_by(
            SupplierPayment.invoice_date.desc(), SupplierPayment.id.desc()
        ).all()
        return jsonify([p.to_dict() for p in payments])
    # POST – record a payment
    data = request.get_json() or {}
    try:
        amount = float(data.get('amount') or 0)
    except (ValueError, TypeError):
        return jsonify({'error': 'invalid_amount', 'message': 'Invalid amount.'}), 400
    if amount <= 0:
        return jsonify({'error': 'invalid_amount', 'message': 'Amount must be greater than zero.'}), 400
    invoice_date = None
    if data.get('invoice_date'):
        try:
            invoice_date = datetime.strptime(data['invoice_date'], '%Y-%m-%d').date()
        except ValueError:
            pass
    payment_date = None
    if data.get('payment_date'):
        try:
            payment_date = datetime.strptime(data['payment_date'], '%Y-%m-%d').date()
        except ValueError:
            pass
    payment = SupplierPayment(
        supplier_id=supplier_id,
        invoice_number=data.get('invoice_number', ''),
        invoice_date=invoice_date,
        amount=amount,
        description=data.get('description', ''),
        payment_date=payment_date,
        payment_method=data.get('payment_method', ''),
        cheque_number=data.get('cheque_number', ''),
        bank_transaction=data.get('bank_transaction', ''),
        created_by=current_user.display_name,
        is_paid=payment_date is not None,
        date_paid=payment_date,
        paid_by=current_user.display_name if payment_date is not None else None,
    )
    db.session.add(payment)
    db.session.commit()
    if payment_date:
        _post_supplier_payment_journal(payment, supplier)
        payment.is_posted = True
        payment.posted_by = current_user.display_name
        db.session.commit()
    return jsonify(payment.to_dict()), 201


@app.route('/api/supplier-payments/<int:payment_id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def api_supplier_payment_detail(payment_id):
    if not current_user.is_manager:
        return jsonify({'error': 'access_denied', 'message': 'Manager access required.'}), 403
    payment = db.session.get(SupplierPayment, payment_id)
    if not payment:
        return jsonify({'error': 'not_found'}), 404
    if getattr(payment, 'is_deleted', False):
        return jsonify({'error': 'not_found'}), 404
    if request.method == 'GET':
        return jsonify(payment.to_dict())
    if request.method == 'PUT':
        data = request.get_json() or {}
        if 'amount' in data:
            try:
                payment.amount = float(data['amount'])
            except (ValueError, TypeError):
                return jsonify({'error': 'invalid_amount'}), 400
        for field in ('invoice_number', 'description', 'payment_method',
                      'cheque_number', 'bank_transaction'):
            if field in data:
                setattr(payment, field, data[field])
        if 'invoice_date' in data:
            try:
                payment.invoice_date = datetime.strptime(data['invoice_date'], '%Y-%m-%d').date() if data['invoice_date'] else None
            except ValueError:
                pass  # Invalid date format from client; leave field unchanged
        if 'payment_date' in data:
            try:
                new_payment_date = datetime.strptime(data['payment_date'], '%Y-%m-%d').date() if data['payment_date'] else None
                payment.payment_date = new_payment_date
                # Mark the invoice as paid and record the date it was paid
                if new_payment_date:
                    payment.is_paid = True
                    payment.date_paid = new_payment_date
                    payment.paid_by = current_user.display_name
            except ValueError:
                pass  # Invalid date format from client; leave field unchanged
        db.session.commit()
        return jsonify(payment.to_dict())
    # DELETE – soft delete
    payment.is_deleted = True
    payment.deleted_by = current_user.display_name if current_user.is_authenticated else 'system'
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/suppliers/payments/post-to-gl', methods=['POST'])
@login_required
def api_supplier_payments_post_to_gl():
    """Post unposted supplier payments to the GL journal and mark them as posted."""
    if not current_user.is_manager:
        return jsonify({'error': 'access_denied', 'message': 'Manager access required.'}), 403
    payments = SupplierPayment.query.filter(
        SupplierPayment.is_deleted == False,
        SupplierPayment.is_paid == True,
        SupplierPayment.is_posted == False,
    ).all()
    if not payments:
        return jsonify({'posted': 0, 'message': 'Aucune entrée non soumise trouvée.'})
    posted_count = 0
    for payment in payments:
        supplier = payment.supplier
        try:
            _post_supplier_payment_journal(payment, supplier)
            payment.is_posted = True
            payment.posted_by = current_user.display_name
            posted_count += 1
        except Exception as exc:
            logger.warning("Could not post supplier payment %s to GL: %s", payment.id, exc)
    db.session.commit()
    return jsonify({'posted': posted_count, 'message': f'{posted_count} entr{"ée" if posted_count == 1 else "ées"} soumise{"" if posted_count == 1 else "s"} au GL.'})


@app.route('/api/suppliers/payments/unpaid', methods=['GET'])
@login_required
def api_suppliers_payments_unpaid():
    """Return all unpaid supplier invoices (is_paid=False) as JSON."""
    if not current_user.is_manager:
        return jsonify({'error': 'access_denied', 'message': 'Manager access required.'}), 403
    payments = SupplierPayment.query.filter(
        SupplierPayment.is_deleted == False,
        SupplierPayment.is_paid == False,
    ).order_by(SupplierPayment.invoice_date.asc(), SupplierPayment.id.asc()).all()
    return jsonify([p.to_dict() for p in payments])


@app.route('/suppliers/unpaid/print')
@login_required
def suppliers_unpaid_print():
    """Print view of all unpaid supplier invoices (is_paid=False)."""
    if not current_user.is_manager:
        flash('Access restricted to managers.', 'danger')
        return redirect(url_for('index'))
    firm = FirmInfo.query.first()
    payments = SupplierPayment.query.filter(
        SupplierPayment.is_deleted == False,
        SupplierPayment.is_paid == False,
    ).order_by(SupplierPayment.invoice_date.asc(), SupplierPayment.id.asc()).all()
    now_str = datetime.utcnow().strftime('%Y-%m-%d %H:%M')
    return render_template('suppliers_unpaid_print.html', firm=firm, payments=payments, now=now_str)


# ── Calendar / Agenda Module ─────────────────────────────────────────────────

@app.route('/calendar')
@login_required
def calendar_page():
    return render_template('calendar.html')


@app.route('/api/calendar/events', methods=['GET', 'POST'])
@login_required
def api_calendar_events():
    if request.method == 'GET':
        date_from = request.args.get('date_from', '').strip()
        date_to = request.args.get('date_to', '').strip()
        matter_id = request.args.get('matter_id', '').strip()
        show_done = request.args.get('show_done', '').lower() in ('1', 'true', 'yes')

        q = CalendarEvent.query.filter(
            CalendarEvent.is_deleted == False,
            CalendarEvent.created_by == current_user.username,
        )
        if matter_id:
            try:
                q = q.filter(CalendarEvent.matter_id == int(matter_id))
            except ValueError:
                pass
        if not show_done:
            q = q.filter(CalendarEvent.is_done == False)
        if date_from:
            try:
                df = datetime.strptime(date_from, '%Y-%m-%d').date()
                q = q.filter(CalendarEvent.event_date >= df)
            except ValueError:
                pass
        if date_to:
            try:
                dt = datetime.strptime(date_to, '%Y-%m-%d').date()
                q = q.filter(CalendarEvent.event_date <= dt)
            except ValueError:
                pass
        events = q.order_by(CalendarEvent.event_date.asc(), CalendarEvent.event_time.asc()).all()
        return jsonify([e.to_dict() for e in events])

    # POST – create new event
    data = request.get_json() or {}
    title = (data.get('title') or '').strip()
    if not title:
        return jsonify({'error': 'title_required', 'message': 'Event title is required.'}), 400
    event_date_str = (data.get('event_date') or '').strip()
    if not event_date_str:
        return jsonify({'error': 'date_required', 'message': 'Event date is required.'}), 400
    try:
        event_date = datetime.strptime(event_date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'invalid_date', 'message': 'Invalid date format. Use YYYY-MM-DD.'}), 400

    matter_id_val = None
    if data.get('matter_id'):
        try:
            matter_id_val = int(data['matter_id'])
            if not db.session.get(Matter, matter_id_val):
                matter_id_val = None
        except (ValueError, TypeError):
            matter_id_val = None

    room_id_val = None
    if data.get('room_id'):
        try:
            room_id_val = int(data['room_id'])
            room = db.session.get(RoomConfig, room_id_val)
            if not room or not room.is_active:
                room_id_val = None
        except (ValueError, TypeError):
            room_id_val = None

    duration_minutes_val = None
    if data.get('duration_minutes'):
        try:
            duration_minutes_val = int(data['duration_minutes'])
        except (ValueError, TypeError):
            duration_minutes_val = None

    event = CalendarEvent(
        matter_id=matter_id_val,
        title=title,
        event_type=data.get('event_type', '') or None,
        event_date=event_date,
        event_time=data.get('event_time', '') or None,
        location=data.get('location', '') or None,
        notes=data.get('notes', '') or None,
        is_done=bool(data.get('is_done', False)),
        assigned_to=data.get('assigned_to', '') or None,
        created_by=current_user.username,
        room_id=room_id_val,
        duration_minutes=duration_minutes_val,
    )
    db.session.add(event)
    db.session.commit()
    return jsonify(event.to_dict()), 201


@app.route('/api/calendar/events/<int:event_id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def api_calendar_event_detail(event_id):
    event = db.session.get(CalendarEvent, event_id)
    if not event or event.is_deleted:
        return jsonify({'error': 'not_found'}), 404

    if request.method == 'GET':
        return jsonify(event.to_dict())

    if request.method == 'PUT':
        data = request.get_json() or {}
        if 'title' in data:
            title = data['title'].strip()
            if not title:
                return jsonify({'error': 'title_required', 'message': 'Event title is required.'}), 400
            event.title = title
        if 'event_date' in data and data['event_date']:
            try:
                event.event_date = datetime.strptime(data['event_date'], '%Y-%m-%d').date()
            except ValueError:
                return jsonify({'error': 'invalid_date'}), 400
        if 'matter_id' in data:
            try:
                mid = int(data['matter_id']) if data['matter_id'] else None
                event.matter_id = mid if (mid and db.session.get(Matter, mid)) else None
            except (ValueError, TypeError):
                event.matter_id = None
        if 'room_id' in data:
            try:
                rid = int(data['room_id']) if data['room_id'] else None
                if rid:
                    room = db.session.get(RoomConfig, rid)
                    event.room_id = rid if (room and room.is_active) else None
                else:
                    event.room_id = None
            except (ValueError, TypeError):
                event.room_id = None
        for field in ('event_type', 'event_time', 'location', 'notes', 'assigned_to'):
            if field in data:
                setattr(event, field, data[field] or None)
        if 'duration_minutes' in data:
            try:
                event.duration_minutes = int(data['duration_minutes']) if data['duration_minutes'] else None
            except (ValueError, TypeError):
                event.duration_minutes = None
        if 'is_done' in data:
            event.is_done = bool(data['is_done'])
        db.session.commit()
        return jsonify(event.to_dict())

    # DELETE – soft delete
    event.is_deleted = True
    db.session.commit()
    return jsonify({'success': True})


# ── Calendar Room Configuration ────────────────────────────────────────────────

@app.route('/api/calendar/rooms', methods=['GET', 'POST'])
@login_required
def api_calendar_rooms():
    """List all room configurations or save/update them (manager only for POST)."""
    if request.method == 'GET':
        rooms = RoomConfig.query.order_by(RoomConfig.room_index).all()
        return jsonify([r.to_dict() for r in rooms])

    # POST – bulk upsert: client sends list of {room_index, room_name, is_active}
    if not current_user.is_manager:
        return jsonify({'error': 'access_denied', 'message': 'Manager access required.'}), 403
    data = request.get_json() or []
    if not isinstance(data, list):
        return jsonify({'error': 'Expected a list of room config items.'}), 400
    for item in data:
        try:
            idx = int(item.get('room_index', 0))
        except (ValueError, TypeError):
            continue
        if idx < 1 or idx > 10:
            continue
        room = RoomConfig.query.filter_by(room_index=idx).first()
        if room is None:
            room = RoomConfig(room_index=idx)
            db.session.add(room)
        room.room_name = (item.get('room_name') or '').strip() or f'Salle {idx}'
        room.is_active = bool(item.get('is_active', True))
    db.session.commit()
    rooms = RoomConfig.query.order_by(RoomConfig.room_index).all()
    return jsonify([r.to_dict() for r in rooms])


@app.route('/api/calendar/rooms/<int:room_id>', methods=['PUT', 'DELETE'])
@login_required
def api_calendar_room_detail(room_id):
    """Update or delete a single room configuration (manager only)."""
    if not current_user.is_manager:
        return jsonify({'error': 'access_denied', 'message': 'Manager access required.'}), 403
    room = db.session.get(RoomConfig, room_id)
    if not room:
        return jsonify({'error': 'not_found'}), 404
    if request.method == 'PUT':
        data = request.get_json() or {}
        if 'room_name' in data:
            room.room_name = data['room_name'].strip() or room.room_name
        if 'is_active' in data:
            room.is_active = bool(data['is_active'])
        db.session.commit()
        return jsonify(room.to_dict())
    # DELETE
    db.session.delete(room)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/calendar/rooms/<int:room_id>/reservations', methods=['GET'])
@login_required
def api_room_reservations(room_id):
    """Return all non-deleted calendar events reserved for a given room."""
    room = db.session.get(RoomConfig, room_id)
    if not room:
        return jsonify({'error': 'not_found'}), 404
    date_from = request.args.get('date_from', '').strip()
    date_to   = request.args.get('date_to', '').strip()
    q = CalendarEvent.query.filter(
        CalendarEvent.is_deleted == False,
        CalendarEvent.room_id == room_id,
    )
    if date_from:
        try:
            df = datetime.strptime(date_from, '%Y-%m-%d').date()
            q = q.filter(CalendarEvent.event_date >= df)
        except ValueError:
            pass
    if date_to:
        try:
            dt = datetime.strptime(date_to, '%Y-%m-%d').date()
            q = q.filter(CalendarEvent.event_date <= dt)
        except ValueError:
            pass
    events = q.order_by(CalendarEvent.event_date.asc(), CalendarEvent.event_time.asc()).all()
    return jsonify([e.to_dict() for e in events])


# ── Salary Module ─────────────────────────────────────────────────────────────

@app.route('/salary')
@login_required
def salary_page():
    """Salary management page (manager only)."""
    if not current_user.is_manager:
        flash('Access restricted to managers.', 'danger')
        return redirect(url_for('index'))
    firm = FirmInfo.query.first()
    return render_template('salary.html', firm=firm)


@app.route('/api/salary/config', methods=['GET', 'POST'])
@login_required
def api_salary_config():
    """List all salary configuration fields or save the configuration."""
    if not current_user.is_manager:
        return jsonify({'error': 'access_denied', 'message': 'Manager access required.'}), 403
    if request.method == 'GET':
        configs = SalaryConfig.query.order_by(SalaryConfig.field_index).all()
        return jsonify([c.to_dict() for c in configs])
    # POST – bulk upsert: client sends list of {field_index, field_name, account_code}
    data = request.get_json() or []
    if not isinstance(data, list):
        return jsonify({'error': 'Expected a list of config items.'}), 400
    for item in data:
        idx = int(item.get('field_index', 0))
        if idx < 1 or idx > 10:
            continue
        cfg = SalaryConfig.query.filter_by(field_index=idx).first()
        if cfg is None:
            cfg = SalaryConfig(field_index=idx)
            db.session.add(cfg)
        cfg.field_name = (item.get('field_name') or '').strip() or f'Champ {idx}'
        cfg.account_code = (item.get('account_code') or '').strip()
        cfg.is_active = bool(item.get('is_active', True))
    db.session.commit()
    configs = SalaryConfig.query.order_by(SalaryConfig.field_index).all()
    return jsonify([c.to_dict() for c in configs])


@app.route('/api/salary/config/<int:config_id>', methods=['PUT', 'DELETE'])
@login_required
def api_salary_config_detail(config_id):
    """Update or delete a single salary config."""
    if not current_user.is_manager:
        return jsonify({'error': 'access_denied', 'message': 'Manager access required.'}), 403
    cfg = db.session.get(SalaryConfig, config_id)
    if not cfg:
        return jsonify({'error': 'not_found'}), 404
    if request.method == 'PUT':
        data = request.get_json() or {}
        if 'field_name' in data:
            cfg.field_name = data['field_name'].strip() or cfg.field_name
        if 'account_code' in data:
            cfg.account_code = data['account_code']
        if 'is_active' in data:
            cfg.is_active = bool(data['is_active'])
        db.session.commit()
        return jsonify(cfg.to_dict())
    # DELETE
    db.session.delete(cfg)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/salary/entries', methods=['GET', 'POST'])
@login_required
def api_salary_entries():
    """List salary entries or add a new one."""
    if not current_user.is_manager:
        return jsonify({'error': 'access_denied', 'message': 'Manager access required.'}), 403
    if request.method == 'GET':
        date_from = request.args.get('date_from', '').strip()
        date_to = request.args.get('date_to', '').strip()
        q = SalaryEntry.query.filter(SalaryEntry.is_deleted == False)
        if date_from:
            q = q.filter(SalaryEntry.entry_date >= date_from)
        if date_to:
            q = q.filter(SalaryEntry.entry_date <= date_to)
        entries = q.order_by(SalaryEntry.entry_date.desc(), SalaryEntry.id.desc()).all()
        return jsonify([e.to_dict() for e in entries])
    # POST
    data = request.get_json() or {}
    config_id = data.get('config_id')
    if not config_id:
        return jsonify({'error': 'config_id is required.'}), 400
    cfg = db.session.get(SalaryConfig, int(config_id))
    if not cfg:
        return jsonify({'error': 'not_found', 'message': 'Salary config not found.'}), 404
    try:
        amount = float(data.get('amount') or 0)
    except (ValueError, TypeError):
        return jsonify({'error': 'invalid_amount'}), 400
    if amount <= 0:
        return jsonify({'error': 'invalid_amount', 'message': 'Amount must be greater than zero.'}), 400
    entry_date = None
    if data.get('entry_date'):
        try:
            entry_date = datetime.strptime(data['entry_date'], '%Y-%m-%d').date()
        except ValueError:
            pass
    if not entry_date:
        entry_date = datetime.utcnow().date()
    entry = SalaryEntry(
        config_id=cfg.id,
        entry_date=entry_date,
        amount=amount,
        description=data.get('description', ''),
        created_by=current_user.username,
    )
    db.session.add(entry)
    db.session.commit()
    return jsonify(entry.to_dict()), 201


@app.route('/api/salary/entries/<int:entry_id>', methods=['DELETE'])
@login_required
def api_salary_entry_detail(entry_id):
    """Soft-delete a salary entry."""
    if not current_user.is_manager:
        return jsonify({'error': 'access_denied', 'message': 'Manager access required.'}), 403
    entry = db.session.get(SalaryEntry, entry_id)
    if not entry or entry.is_deleted:
        return jsonify({'error': 'not_found'}), 404
    entry.is_deleted = True
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/salary/entries/post-to-gl', methods=['POST'])
@login_required
def api_salary_post_to_gl():
    """Post unposted salary entries to the GL journal and mark them as posted."""
    if not current_user.is_manager:
        return jsonify({'error': 'access_denied', 'message': 'Manager access required.'}), 403
    data = request.get_json() or {}
    date_from_str = data.get('date_from', '').strip()
    date_to_str = data.get('date_to', '').strip()

    q = SalaryEntry.query.filter(
        SalaryEntry.is_deleted == False,
        SalaryEntry.is_posted == False,
    )
    if date_from_str:
        try:
            q = q.filter(SalaryEntry.entry_date >= datetime.strptime(date_from_str, '%Y-%m-%d').date())
        except ValueError:
            pass
    if date_to_str:
        try:
            q = q.filter(SalaryEntry.entry_date <= datetime.strptime(date_to_str, '%Y-%m-%d').date())
        except ValueError:
            pass
    entries = q.order_by(SalaryEntry.entry_date, SalaryEntry.id).all()
    if not entries:
        return jsonify({'posted': 0, 'message': 'No unposted entries found for this period.'})

    actor = current_user.display_name if current_user.is_authenticated else 'system'
    posted_count = 0
    for entry in entries:
        cfg = entry.config
        account_code = cfg.account_code if cfg else ''
        if not account_code:
            continue
        amount = float(entry.amount or 0)
        if amount <= 0:
            continue
        try:
            _create_journal_entry(
                entry_date=entry.entry_date,
                description=f'Salaire – {cfg.field_name if cfg else ""}',
                source_type='salary',
                source_id=entry.id,
                lines=[
                    {
                        'account_code': account_code,
                        'debit': amount,
                        'credit': 0,
                    },
                    {
                        'account_code': '2100',
                        'debit': 0,
                        'credit': amount,
                    },
                ],
                created_by=actor,
            )
            entry.is_posted = True
            entry.posted_by = actor
            posted_count += 1
        except Exception as exc:
            logger.warning("Could not post salary entry %s to GL: %s", entry.id, exc)
    db.session.commit()
    return jsonify({'posted': posted_count, 'message': f'{posted_count} entr{"ée" if posted_count == 1 else "ées"} soumise{"" if posted_count == 1 else "s"} au GL.'})




def _run_license_diagnostic():
    """Run a one-time licence diagnostic at startup."""
    print("\n" + "="*50)
    print("   DIAGNOSTIC FINAL DE LA LICENCE")
    print("="*50)

    # 1. Calcul du fingerprint réel de ce PC
    reels_fp = None
    try:
        reels_guid = _licensing.get_machine_guid()
        reels_fp = _licensing.compute_fingerprint(reels_guid)
        print(f"TON PC (Réel)     : {reels_fp}")
    except Exception as e:
        print(f"ERREUR CALCUL PC  : {e}")

    # 2. Lecture et Vérification Signature
    # Use the same path as the running app: read LICENSE_FILE from the
    # already-loaded .env (load_dotenv() ran at module level above), falling
    # back to the default config/license.json location.
    fichier = os.path.expandvars(
        os.environ.get(
            'LICENSE_FILE',
            os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config', 'license.json'),
        )
    )
    print(f"FICHIER LICENCE   : {fichier}")
    if os.path.exists(fichier):
        try:
            with open(fichier, "r") as f:
                data = json.load(f)
                dans_fichier = data.get('device_fingerprint')
                print(f"DANS LE FICHIER   : {dans_fichier}")

                # TEST DU FINGERPRINT
                if reels_fp == dans_fichier:
                    print("✅ MATCH : L'ordinateur correspond à la licence.")
                else:
                    print("❌ ERREUR : L'ordinateur ne correspond PAS.")

                # TEST DE LA SIGNATURE (LE POINT CRITIQUE)
                cle_publique = _licensing._DEFAULT_PUBLIC_KEY_B64
                if _licensing.verify_signature(data, cle_publique):
                    print("✅ SIGNATURE : La signature est VALIDE !")
                else:
                    print("❌ SIGNATURE : La signature est REJETÉE (Clé incorrecte) !")

        except Exception as e:
            print(f"ERREUR LECTURE   : {e}")
    else:
        print(f"ERREUR           : Le fichier licence est absent à : {fichier}")
        print(f"  → Vérifiez la variable LICENSE_FILE dans votre fichier .env")
    print("="*50 + "\n")

# BIND_HOST is the address the server listens on.  Behind an IIS reverse
# proxy this must stay 127.0.0.1 (the default) so that IIS can reach the
# backend.  HOST is the *external-facing* address used only for building
# URLs in emails (password-reset links, etc.).

if __name__ == '__main__':
    # Force UTF-8 output for Windows console / NSSM service
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')
    except AttributeError:
        pass  # Already wrapped or not a real terminal

    bind_host = os.environ.get('BIND_HOST', '127.0.0.1')
    port  = int(os.environ.get('PORT', '5000'))
    debug = os.environ.get('DEBUG', '').lower() in ('1', 'true', 'yes')

    print(f"\n{'='*60}")
    print(f"LawLedger Application Starting")
    print(f"{'='*60}")
    print(f"Listening on: http://{bind_host}:{port}")
    print(f"{'='*60}\n")

    with app.app_context():
        _apply_schema_migrations()

    _run_license_diagnostic()

    use_waitress = os.environ.get('USE_WAITRESS', '').lower() in ('1', 'true', 'yes')
    if use_waitress:
        try:
            from waitress import serve
            print("Starting with Waitress WSGI server")
            serve(app, host=bind_host, port=port, threads=100)
        except ImportError:
            print("Waitress not installed; falling back to Flask dev server. Install with: pip install waitress")
            app.run(host=bind_host, port=port, debug=debug)
    else:
        app.run(host=bind_host, port=port, debug=debug)
